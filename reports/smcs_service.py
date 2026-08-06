from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import SMCSCode


EXPLICIT_CODE_PATTERN = re.compile(
    r"\bSMCS(?:\s+CODE)?\s*[:#-]?\s*([A-Z0-9]{2,5})\b",
    re.IGNORECASE,
)


def _normalize(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


@transaction.atomic
def import_smcs_workbook(path: str | Path) -> dict:
    source_path = Path(path)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook.active
    created = 0
    updated = 0
    skipped = 0
    errors = []
    seen = set()
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        description = str(values[0] or "").strip()
        raw_code = values[1] if len(values) > 1 else None
        code = str(raw_code or "").strip().upper()
        if code.endswith(".0"):
            code = code[:-2]
        if not code and not description:
            continue
        if not re.fullmatch(r"[A-Z0-9]{2,20}", code) or not description:
            errors.append(f"Row {row_number}: code and description are required.")
            continue
        if code in seen:
            skipped += 1
            continue
        seen.add(code)
        item, was_created = SMCSCode.objects.update_or_create(
            code=code,
            defaults={
                "description": description,
                "source": "CAT SMCS Codes",
                "source_file": source_path.name,
                "validation_status": "To Review",
                "is_active": True,
                "imported_at": timezone.now(),
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1
    return {
        "rows_read": worksheet.max_row - 1,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


def _description_candidates() -> list[tuple[SMCSCode, str]]:
    candidates = []
    queryset = SMCSCode.objects.filter(
        is_active=True,
        validation_status__in=["To Review", "Validated"],
    ).only("id", "code", "description")
    for item in queryset:
        normalized = _normalize(item.description)
        # Very short generic descriptions create unsafe matches in comments.
        if len(normalized) >= 8 or len(normalized.split()) >= 2:
            candidates.append((item, normalized))
    return candidates


def resolve_event_smcs(events: list[dict]) -> dict:
    codes = {
        item.code: item
        for item in SMCSCode.objects.filter(
            is_active=True,
            validation_status__in=["To Review", "Validated"],
        )
    }
    descriptions = _description_candidates()
    grouped = defaultdict(lambda: {
        "event_ids": set(),
        "equipment": set(),
        "downtime_hours": 0.0,
        "explicit_matches": 0,
        "description_matches": 0,
        "comments_count": 0,
    })
    matched_events = set()
    for event in events:
        comment = str(event.get("Comment") or "").strip()
        if not comment:
            continue
        normalized_comment = f" {_normalize(comment)} "
        matches: dict[str, str] = {}
        for raw_code in EXPLICIT_CODE_PATTERN.findall(comment):
            code = raw_code.upper()
            if code in codes:
                matches[code] = "Explicit SMCS code"
        for item, description in descriptions:
            if f" {description} " in normalized_comment:
                matches.setdefault(item.code, "Exact description")
        for code, match_method in matches.items():
            bucket = grouped[code]
            event_id = str(event.get("Event ID") or "")
            if event_id in bucket["event_ids"]:
                continue
            bucket["event_ids"].add(event_id)
            equipment = event.get("Serial Number") or event.get("Equipment")
            if equipment:
                bucket["equipment"].add(str(equipment))
            bucket["downtime_hours"] += float(event.get("Duration") or 0)
            bucket["comments_count"] += 1
            if match_method == "Explicit SMCS code":
                bucket["explicit_matches"] += 1
            else:
                bucket["description_matches"] += 1
            matched_events.add(event_id)

    rows = []
    for code, values in grouped.items():
        item = codes[code]
        methods = []
        if values["explicit_matches"]:
            methods.append("Explicit SMCS code")
        if values["description_matches"]:
            methods.append("Exact description")
        rows.append({
            "SMCS Code": code,
            "SMCS Description": item.description,
            "Downtime Hours": round(values["downtime_hours"], 2),
            "Event Count": len(values["event_ids"]),
            "Affected Equipment": len(values["equipment"]),
            "Comments Count": values["comments_count"],
            "Match Method": " + ".join(methods),
            "Event IDs": sorted(values["event_ids"]),
        })
    rows.sort(key=lambda row: (-row["Downtime Hours"], row["SMCS Code"]))
    total_events = len(events)
    return {
        "rows": rows,
        "coverage": {
            "event_count": total_events,
            "matched_event_count": len(matched_events),
            "unmatched_event_count": max(total_events - len(matched_events), 0),
            "event_coverage_percentage": round(
                len(matched_events) / total_events * 100 if total_events else 0,
                2,
            ),
        },
        "matching_rule": (
            "SMCS codes explicitly labelled in comments plus exact CAT SMCS "
            "description phrases. No AI-inferred code is used."
        ),
    }
