from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone

from .models import AIConfigSection, KnowledgeSynonym, OpenAIModelPricing, OpenAIUsageLog
from .openai_client_service import create_tracked_response


class OpenAIUsageDashboardTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("usage-admin", password="password", is_staff=True)
        OpenAIModelPricing.objects.create(
            model_name="test-model",
            effective_from=timezone.now() - timezone.timedelta(days=1),
            input_cost_per_million_tokens=Decimal("1"),
            cached_input_cost_per_million_tokens=Decimal("0.5"),
            output_cost_per_million_tokens=Decimal("2"),
        )

    def test_tracked_response_writes_internal_usage(self):
        response = SimpleNamespace(
            model="test-model",
            id="resp-test",
            usage=SimpleNamespace(
                input_tokens=100,
                output_tokens=50,
                total_tokens=150,
                input_tokens_details=SimpleNamespace(cached_tokens=20),
                output_tokens_details=SimpleNamespace(reasoning_tokens=10),
            ),
        )
        client = SimpleNamespace(
            responses=SimpleNamespace(create=lambda **kwargs: response),
        )
        create_tracked_response(
            client,
            model="test-model",
            input=[],
            section="performance",
            feature="Intent Extraction",
        )
        log = OpenAIUsageLog.objects.get()
        self.assertEqual(log.total_tokens, 150)
        self.assertEqual(log.cached_input_tokens, 20)
        self.assertIsNotNone(log.estimated_cost)

    def test_dashboard_is_admin_only(self):
        response = self.client.get("/api/admin/openai-usage/dashboard/")
        self.assertEqual(response.status_code, 302)
        self.client.force_login(self.admin)
        response = self.client.get("/api/admin/openai-usage/dashboard/")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])

    def test_synonym_excel_export_is_available_to_admin(self):
        self.client.force_login(self.admin)
        response = self.client.get("/knowledge-base/synonyms/export/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_synonym_template_and_import(self):
        from openpyxl import load_workbook

        AIConfigSection.objects.get_or_create(
            code="performance",
            defaults={"name": "Performance"},
        )
        initial_count = KnowledgeSynonym.objects.count()
        self.client.force_login(self.admin)
        template = self.client.get("/knowledge-base/synonyms/template/")
        self.assertEqual(template.status_code, 200)
        workbook = load_workbook(BytesIO(template.content))
        worksheet = workbook["Synonyms Import"]
        worksheet.cell(2, 2, "availability")
        worksheet.cell(2, 3, "physical availability")
        output = BytesIO()
        workbook.save(output)
        workbook_bytes = output.getvalue()
        upload = SimpleUploadedFile(
            "synonyms.xlsx",
            workbook_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            "/knowledge-base/synonyms/import/",
            {"file": upload},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["summary"]["created"], 1)
        self.assertEqual(KnowledgeSynonym.objects.count(), initial_count + 1)
        self.assertTrue(KnowledgeSynonym.objects.filter(
            canonical_term="availability",
            synonym="physical availability",
        ).exists())
        duplicate_upload = SimpleUploadedFile(
            "synonyms.xlsx",
            workbook_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        duplicate_response = self.client.post(
            "/knowledge-base/synonyms/import/",
            {"file": duplicate_upload},
        )
        self.assertEqual(duplicate_response.status_code, 200)
        self.assertEqual(duplicate_response.json()["summary"]["created"], 0)
        self.assertEqual(duplicate_response.json()["summary"]["error_count"], 1)
        self.assertEqual(KnowledgeSynonym.objects.count(), initial_count + 1)
