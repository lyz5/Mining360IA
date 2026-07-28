import csv
import json
import re
import time
from urllib.parse import unquote, urlparse

from django.db import models, transaction
from django.db.models.deletion import ProtectedError, RestrictedError
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.apps import apps
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import authenticate, login as django_login, logout as django_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from datetime import date, datetime
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.db.models.functions import TruncDate

try:  # pragma: no cover - optional dependency
    import snowflake.connector as snowflake_connector
except ImportError:  # pragma: no cover
    snowflake_connector = None

from .live_sources import (
    add_live_source,
    add_live_view,
    delete_live_source,
    delete_live_view,
    execute_live_view,
    get_live_source,
    get_live_view,
    list_live_sources,
    set_live_source_verification,
    update_live_source,
    update_live_view,
)
from .sqlserver import connect
from .powerbi import (
    RLS_ROLE_OPTIONS,
    discover_dataset_measures_rest,
    execute_dataset_dax,
    get_access_token,
    generate_report_embed_token,
    get_workspace_report,
    env_value,
    list_workspace_reports_with_refresh,
    list_workspace_reports,
    resolve_workspace_dataset_id,
    resolve_dataset_roles,
)
from .resource_library import (
    get_resource,
    get_resource_path,
    list_resource_facets,
    list_resources,
    read_text_resource,
    save_uploaded_resource,
)
from .semantic_engine import build_availability_matrix_question, build_availability_question
from .power_automate import execute_dax_via_flow, get_flow_url
from .openai_assistant import (
    chat_semantic_response_with_openai,
    interpret_semantic_answer_with_openai,
    is_openai_configured,
    parse_semantic_question_with_openai,
)
from .openai_service import generate_chat_response, extract_intent as openai_extract_intent
from .intent_extractor_service import extract_intent
from .dax_generator_service import generate_dax_from_intent, validate_intent, IntentValidationError
from .knowledge_resolution_service import (
    get_cached_trace,
    resolve_knowledge_question,
    trace_as_basic_pdf,
    trace_as_markdown,
)
from .synonym_resolution_service import resolve_synonyms
from .synonym_utils import default_match_type, normalize_synonym_key
from .ai_config_service import (
    build_section_catalog,
    get_active_section_objects,
    get_active_sections,
    get_dax_template,
    get_filter_mapping,
    get_metric_mapping,
    get_prompt_template,
    get_question_examples,
    get_section_by_code,
    get_synonyms,
)
from .data_quality import (
    DataQualityContext,
    build_context_summary,
    compute_score,
    run_checks,
    serialize_result,
    available_checks,
)
from .models import (
    AIConfigSection,
    AIDaxTemplate,
    AIFilterMapping,
    AIBusinessRule,
    AIBusinessVocabulary,
    AIDebugRun,
    AIFewShotExample,
    AIKPITarget,
    AIMetricMapping,
    AIPowerBIPage,
    AIPromptTemplate,
    AIQuestionExample,
    AIRecommendedAction,
    AISemanticColumn,
    AISemanticMeasure,
    AISemanticRelationship,
    AISemanticTable,
    AISynonym,
    AIVisualMapping,
    BusinessPerformanceConfig,
    BusinessPerformanceMapping,
    DataBrowser,
    DataBrowserColumn,
    DataQualityRun,
    KnowledgeAILog,
    KnowledgeBusinessGlossary,
    KnowledgeBusinessRule,
    KnowledgeKPIDictionary,
    KnowledgeMiningTerminology,
    KnowledgePrompt,
    KnowledgeQuestion,
    KnowledgeRecommendedAction,
    KnowledgeSynonym,
    KnowledgeUserFeedback,
    KPIPageMapping,
    PowerBIReport,
    ReportingReportPreference,
    SystemDatabaseConfig,
    SystemIntegrationConfig,
    SystemManagedTable,
    SystemParameter,
)
from .system_configuration_service import (
    ensure_portable_configuration,
    integration_payload,
    save_integration,
    schema_payload,
    test_integration,
)
from .business_performance_service import (
    BusinessPerformanceError,
    BusinessPerformanceService,
    MappingNotConfigured,
)
from .models import PlatformUser
from .access_control import has_module_access, is_platform_admin, user_module_access, wants_json
from .powerbi_interaction_orchestrator import process_user_question
from .ad_auth import exchange_code, fetch_me, login_url, search_directory_users
from .data_browsers import (
    DataBrowserValidationError,
    delete_browser_records,
    inspect_import_file,
    import_browser_records,
    import_browser_records_batch,
    get_import_job_status,
    insert_browser_record,
    lookup_options,
    preview_browser_data,
    start_import_job,
    sync_browser_sql,
    update_browser_record,
    validate_browser_definition,
)


def _json_safe(value):
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _normalize_snowflake_account(value: str) -> str:
    account = (value or "").strip()
    if not account:
        return ""
    account = account.removeprefix("https://").removeprefix("http://")
    account = account.replace(".snowflakecomputing.com", "")
    account = account.rstrip("/")
    return account


def _source_inventory(source):
    summary = {
        "label": "Inventory",
        "count": None,
        "custom_views": len(source.views),
    }

    engine = source.engine.lower()
    if engine == "sql server":
        summary["label"] = "Databases"
        try:
            with connect(
                server=source.server,
                database=source.database or "master",
                user=source.user or None,
                password=source.password or None,
                port=source.port or None,
            ) as connection:
                cursor = connection.cursor()
                cursor.execute(
                    """
                    SELECT COUNT(*)
                    FROM sys.databases
                    WHERE database_id > 4
                      AND state = 0
                    """
                )
                row = cursor.fetchone() or (0,)
            summary["count"] = int(row[0] or 0)
        except Exception:
            summary["count"] = None
        return summary

    if engine == "snowflake":
        summary["label"] = "Schemas"
        if snowflake_connector is None:
            return summary
        account = source.connection_details.get("account") or source.server
        warehouse = source.connection_details.get("warehouse")
        role = source.connection_details.get("role") or None
        database = source.database or source.connection_details.get("database") or None
        if not account or not warehouse or not source.user:
            return summary
        try:
            connect_kwargs = {
                "account": _normalize_snowflake_account(account),
                "user": source.user,
                "password": source.password or None,
                "warehouse": warehouse,
                "role": role,
            }
            if database:
                connect_kwargs["database"] = database
            connection = snowflake_connector.connect(**connect_kwargs)
            try:
                cursor = connection.cursor()
                if database:
                    cursor.execute(
                        """
                        SELECT COUNT(*)
                        FROM INFORMATION_SCHEMA.SCHEMATA
                        WHERE SCHEMA_NAME <> 'INFORMATION_SCHEMA'
                        """
                    )
                else:
                    cursor.execute("SHOW SCHEMAS IN ACCOUNT")
                row = cursor.fetchone() or (0,)
            finally:
                connection.close()
            summary["count"] = int(row[0] or 0)
        except Exception:
            summary["count"] = None

    return summary


def _source_databases(source):
    databases = []
    if source.engine.lower() != "sql server":
        return databases

    connection_targets = [source.database or "master"]
    if connection_targets[0].lower() != "master":
        connection_targets.append("master")

    for database_name in connection_targets:
        try:
            with connect(
                server=source.server,
                database=database_name,
                user=source.user or None,
                password=source.password or None,
                port=source.port or None,
            ) as connection:
                cursor = connection.cursor()
                cursor.execute(
                    """
                    SELECT name
                    FROM sys.databases
                    WHERE database_id > 4
                      AND state = 0
                    ORDER BY name
                    """
                )
                databases = [str(row[0]) for row in cursor.fetchall() or [] if row and row[0]]
                break
        except Exception:
            databases = []

    selected = (source.database or "").strip()
    if selected and selected not in databases:
        databases.insert(0, selected)
    return databases


def _source_catalog(source):
    catalog = {
            "tables": [],
            "views": [],
            "custom_views": [
                {
                    "key": view.key,
                    "name": view.name,
                    "description": view.description,
                    "sql": view.sql,
                    "preview_url": reverse("source-object-preview", args=[source.key, "custom", view.key]),
                    "edit_url": reverse("source-custom-view-edit", args=[source.key, view.key]),
                    "delete_url": reverse("source-custom-view-delete", args=[source.key, view.key]),
                }
                for view in source.views
            ],
        }

    if source.engine.lower() != "sql server":
        return catalog

    try:
        with connect(
            server=source.server,
            database=source.database or "master",
            user=source.user or None,
            password=source.password or None,
            port=source.port or None,
        ) as connection:
            cursor = connection.cursor()
            cursor.execute(
                """
                SELECT s.name AS schema_name, t.name AS table_name
                FROM sys.tables t
                INNER JOIN sys.schemas s ON s.schema_id = t.schema_id
                WHERE t.is_ms_shipped = 0
                ORDER BY s.name, t.name
                """
            )
            catalog["tables"] = [
                {
                    "schema": row[0],
                    "name": row[1],
                    "qualified_name": f"{row[0]}.{row[1]}",
                    "preview_url": reverse("source-object-preview", args=[source.key, "table", f"{row[0]}.{row[1]}"]),
                }
                for row in cursor.fetchall() or []
            ]

            cursor.execute(
                """
                SELECT s.name AS schema_name, v.name AS view_name
                FROM sys.views v
                INNER JOIN sys.schemas s ON s.schema_id = v.schema_id
                WHERE v.is_ms_shipped = 0
                ORDER BY s.name, v.name
                """
            )
            catalog["views"] = [
                {
                    "schema": row[0],
                    "name": row[1],
                    "qualified_name": f"{row[0]}.{row[1]}",
                    "preview_url": reverse("source-object-preview", args=[source.key, "view", f"{row[0]}.{row[1]}"]),
                }
                for row in cursor.fetchall() or []
            ]
    except Exception:
        pass

    return catalog


def _normalize_preview_limit(value):
    text = str(value or "").strip().lower()
    if not text or text == "1000":
        return 1000
    if text == "all":
        return None
    try:
        parsed = int(text)
    except ValueError:
        return 1000
    return max(parsed, 1)


def _normalize_preview_operator(value):
    operator = str(value or "equal").strip().lower()
    return operator if operator in {"equal", "like"} else "equal"


def _normalize_preview_joiner(value):
    joiner = str(value or "AND").strip().upper()
    return joiner if joiner in {"AND", "OR"} else "AND"


def _parse_preview_filters(raw_filters):
    if not raw_filters:
        return []

    if isinstance(raw_filters, str):
        try:
            payload = json.loads(raw_filters)
        except Exception:
            return []
    elif isinstance(raw_filters, list):
        payload = raw_filters
    else:
        return []

    if not isinstance(payload, list):
        return []

    filters = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        column = str(item.get("column") or item.get("column_name") or "").strip()
        value = item.get("value")
        if not column or value is None or str(value).strip() == "":
            continue
        filters.append(
            {
                "column": column,
                "operator": _normalize_preview_operator(item.get("operator")),
                "value": str(value),
                "joiner": _normalize_preview_joiner(item.get("joiner")),
                "type": str(item.get("type") or "").strip().lower(),
            }
        )
    return filters


def _escape_sql_identifier(value: str) -> str:
    return f"[{str(value).replace(']', ']]')}]"


def _quote_sql_literal(value: str, force_text: bool = False) -> str:
    text = str(value or "").replace("'", "''")
    return f"N'{text}'" if force_text else f"'{text}'"


def _preview_sql_literal(value: str, preview_type: str) -> str:
    text = str(value or "").strip()
    if preview_type == "number" and text:
        try:
            float(text)
        except ValueError:
            return _quote_sql_literal(text, force_text=True)
        return text
    return _quote_sql_literal(text, force_text=True)


def _normalize_object_identifier(value: str) -> str:
    return unquote(str(value or "").strip()).rstrip("/").strip()


def _build_preview_filter_clause(filter_data: dict) -> str:
    column = _escape_sql_identifier(filter_data["column"])
    operator = _normalize_preview_operator(filter_data.get("operator"))
    value = str(filter_data.get("value") or "").strip()
    preview_type = str(filter_data.get("type") or "").strip().lower()

    if preview_type == "date":
        return f"CONVERT(date, {column}) = {_quote_sql_literal(value[:10], force_text=True)}"

    if operator == "like":
        return f"{column} LIKE {_quote_sql_literal(f'%{value}%', force_text=True)}"

    if preview_type == "number":
        return f"{column} = {_preview_sql_literal(value, 'number')}"

    return f"{column} = {_quote_sql_literal(value, force_text=True)}"


def _build_preview_sql(object_sql: str, filters: list[dict], limit_value):
    preview_limit = _normalize_preview_limit(limit_value)
    select_prefix = "SELECT *" if preview_limit is None else f"SELECT TOP ({preview_limit}) *"
    sql = f"{select_prefix}\nFROM {object_sql}"

    active_filters = [item for item in filters if str(item.get("value") or "").strip()]
    if active_filters:
        clauses = []
        for index, filter_data in enumerate(active_filters):
            clause = _build_preview_filter_clause(filter_data)
            if index == 0:
                clauses.append(f"({clause})")
            else:
                joiner = _normalize_preview_joiner(filter_data.get("joiner"))
                clauses.append(f"{joiner} ({clause})")
        sql += "\nWHERE " + " ".join(clauses)
    return sql


def _table_preview_sql(qualified_name: str, limit="1000") -> str:
    parts = [part.strip("[] ") for part in qualified_name.split(".", 1)]
    if len(parts) != 2 or not all(parts):
        raise ValueError("Invalid table identifier.")
    schema, table = parts
    object_sql = f"{_escape_sql_identifier(schema)}.{_escape_sql_identifier(table)}"
    return _build_preview_sql(object_sql, [], limit)


def _preview_column_type(value) -> str:
    from datetime import date as _date, datetime as _datetime
    from decimal import Decimal as _decimal

    if value is None:
        return "text"
    if isinstance(value, bool):
        return "text"
    if isinstance(value, (_datetime, _date)):
        return "date"
    if isinstance(value, (int, float, _decimal)):
        return "number"

    text = str(value).strip()
    if not text:
        return "text"
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        return "date"
    return "text"


def _fetch_query_preview(source, sql: str, object_name: str) -> dict:
    with connect(
        server=source.server,
        database=source.database or "master",
        user=source.user or None,
        password=source.password or None,
        port=source.port or None,
    ) as connection:
        cursor = connection.cursor()
        rows = cursor.execute(sql).fetchall()
        columns = [column[0] for column in cursor.description or []]

    column_types = []
    for column_index, _column in enumerate(columns):
        detected_type = "text"
        for row in rows:
            sample_value = row[column_index] if column_index < len(row) else None
            detected_type = _preview_column_type(sample_value)
            if detected_type != "text" or sample_value not in (None, ""):
                break
        column_types.append(detected_type)

    records = [
        {column: _json_safe(value) for column, value in zip(columns, row)}
        for row in rows
    ]
    row_values = [
        [_json_safe(value) for value in row]
        for row in rows
    ]
    return {
        "sql": sql,
        "columns": columns,
        "column_types": column_types,
        "rows": records,
        "row_values": row_values,
        "row_count": len(records),
        "qualified_name": object_name,
    }


def _preview_payload_from_url(source, preview_url: str) -> dict:
    if not preview_url:
        raise ValueError("No preview target selected.")

    object_sql, object_name, target = _resolve_preview_object(source, preview_url)

    sql = f"SELECT *\nFROM {object_sql}"
    preview = _fetch_query_preview(source, sql, object_name)
    return {
        "kind": target["kind"],
        "identifier": target["identifier"],
        "sql": sql,
        "preview": preview,
        "object_name": object_name,
    }


def _resolve_preview_target_from_url(source, preview_url: str) -> dict:
    if not preview_url:
        raise ValueError("No preview target selected.")

    parsed = urlparse(preview_url)
    path = parsed.path or preview_url
    object_sql, object_name, target = _resolve_preview_object(source, preview_url)

    return {
        "kind": target["kind"],
        "identifier": target["identifier"],
        "object_name": object_name,
    }


def _resolve_preview_object(source, preview_url: str) -> tuple[str, str, dict]:
    parsed = urlparse(preview_url)
    path = parsed.path or preview_url
    match = re.search(
        r"/data-sources/(?P<source_key>[^/]+)/preview/(?P<kind>[^/]+)/(?P<identifier>.+?)/?$",
        path,
    )
    if match and match.group("source_key").lower() != source.key.lower():
        raise ValueError("Preview source mismatch.")

    if match:
        kind = match.group("kind").strip().lower()
        identifier = _normalize_object_identifier(match.group("identifier"))
    else:
        table_match = re.search(
            r"/data-sources/(?P<source_key>[^/]+)/tables/(?P<identifier>.+?)/preview/?$",
            path,
        )
        if not table_match:
            raise ValueError("Unsupported preview target.")
        if table_match.group("source_key").lower() != source.key.lower():
            raise ValueError("Preview source mismatch.")
        kind = "table"
        identifier = _normalize_object_identifier(table_match.group("identifier"))

    if kind in {"table", "tables", "view", "views"}:
        object_sql = ".".join(_escape_sql_identifier(part.strip("[] ")) for part in identifier.split(".", 1))
        object_name = identifier
    elif kind in {"custom", "custom-view", "custom_view"}:
        _, view = get_live_view(source.key, identifier)
        object_sql = f"(\n{view.sql}\n) AS dq_source"
        object_name = view.name
    else:
        raise ValueError("Unknown preview object type.")

    return object_sql, object_name, {
        "kind": kind,
        "identifier": identifier,
        "object_name": object_name,
    }


def _preview_export_rows(source, preview_url: str, filters: list[dict] | None = None) -> dict:
    object_sql, object_name, target = _resolve_preview_object(source, preview_url)
    sql = _build_preview_sql(object_sql, filters or [], "all")
    preview = _fetch_query_preview(source, sql, object_name)
    return {
        "kind": target["kind"],
        "identifier": target["identifier"],
        "object_name": object_name,
        "sql": sql,
        "preview": preview,
    }


def _rows_from_preview_snapshot(preview_snapshot: dict) -> list[dict]:
    if not isinstance(preview_snapshot, dict):
        return []

    rows = preview_snapshot.get("rows")
    if isinstance(rows, list):
        return rows

    columns = preview_snapshot.get("columns") if isinstance(preview_snapshot.get("columns"), list) else []
    row_values = preview_snapshot.get("row_values")
    if not isinstance(row_values, list) or not columns:
        return []

    normalized_rows = []
    for row in row_values:
        if not isinstance(row, (list, tuple)):
            continue
        normalized_rows.append({column: value for column, value in zip(columns, row)})
    return normalized_rows


def _data_quality_context_from_request(source, preview_url: str, request_payload: dict | None = None) -> tuple[DataQualityContext, dict]:
    request_payload = request_payload if isinstance(request_payload, dict) else {}
    embedded_preview = request_payload.get("preview") if isinstance(request_payload, dict) else None
    if isinstance(embedded_preview, dict) and embedded_preview.get("columns") is not None:
        target = _resolve_preview_target_from_url(source, preview_url)
        rows = _rows_from_preview_snapshot(embedded_preview)
        preview = {
            "sql": embedded_preview.get("sql") or "",
            "columns": embedded_preview.get("columns") or [],
            "column_types": embedded_preview.get("column_types") or [],
            "rows": rows,
            "row_values": embedded_preview.get("row_values") or [],
            "row_count": embedded_preview.get("row_count") or len(rows),
            "qualified_name": embedded_preview.get("qualified_name") or target["object_name"],
            "filters": embedded_preview.get("filters") or [],
            "limit": embedded_preview.get("limit") or "1000",
        }
        payload = {
            "kind": target["kind"],
            "identifier": target["identifier"],
            "sql": embedded_preview.get("sql") or "",
            "preview": preview,
            "object_name": target["object_name"],
        }
    else:
        payload = _preview_payload_from_url(source, preview_url)
        preview = payload["preview"]
    previous_run = (
        DataQualityRun.objects.filter(
            source_key=source.key,
            object_kind=payload["kind"],
            object_name=payload["object_name"],
            status="Completed",
        )
        .order_by("-created_at")
        .first()
    )
    previous_summary = {}
    if previous_run and isinstance(previous_run.summary, dict):
        previous_summary = previous_run.summary

    context = DataQualityContext(
        source_key=source.key,
        source_name=source.name,
        object_kind=payload["kind"],
        object_name=payload["object_name"],
        columns=preview.get("columns", []),
        column_types=preview.get("column_types", []),
        records=preview.get("rows", []),
        preview_url=preview_url,
        previous_summary=previous_summary,
        metadata=(request_payload or {}).get("metadata", {}) if isinstance(request_payload, dict) else {},
    )
    return context, payload


def _run_data_quality(source, preview_url: str, control_key: str | None = None, request_payload: dict | None = None):
    context, payload = _data_quality_context_from_request(source, preview_url, request_payload)
    keys = [control_key] if control_key else None
    results = run_checks(context, keys)
    summary = build_context_summary(context)
    score = compute_score(results)
    stored_payload = dict(request_payload or {})
    stored_payload["preview_url"] = preview_url
    run = DataQualityRun.objects.create(
        source_key=source.key,
        source_name=source.name,
        object_kind=context.object_kind,
        object_name=context.object_name,
        run_type="single" if control_key else "all",
        status="Completed",
        score=score,
        total_rows=context.row_count,
        controls_count=len(results),
        summary=summary,
        results=[serialize_result(result) for result in results],
        request_payload=stored_payload,
        finished_at=timezone.now(),
    )
    return run, context, payload, results, summary, score


def _is_ajax_request(request) -> bool:
    return (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in request.headers.get("Accept", "")
    )


def _user_is_platform_admin(user) -> bool:
    return is_platform_admin(user)


def _platform_role_payload(data) -> dict:
    is_admin = data.get("is_platform_admin") == "on"
    scope = {}
    if not is_admin:
        for field, scope_key in (("bp_scope_countries", "country"), ("bp_scope_customers", "customer")):
            values = [item.strip() for item in data.get(field, "").split(",") if item.strip()]
            if values:
                scope[scope_key] = values
        if data.get("bp_rls_role", "").strip():
            scope["rls_role"] = data.get("bp_rls_role", "").strip()
    return {
        "is_platform_admin": is_admin,
        "can_access_reporting": is_admin or data.get("can_access_reporting") == "on",
        "can_access_ai": is_admin or data.get("can_access_ai") == "on",
        "can_access_data": is_admin or data.get("can_access_data") == "on",
        "can_access_sources": is_admin or data.get("can_access_sources") == "on",
        "business_performance_role": "Administrator" if is_admin else data.get("business_performance_role", ""),
        "business_performance_scope": scope,
    }


def login_page(request):
    if request.user.is_authenticated:
        return redirect(request.GET.get("next") or "dashboard")
    next_url = request.GET.get("next") or request.POST.get("next") or ""
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        authenticated_user = authenticate(request, username=username, password=password)
        if authenticated_user and authenticated_user.is_active:
            platform_user = getattr(authenticated_user, "platformuser", None)
            if platform_user and platform_user.is_active:
                django_login(request, authenticated_user)
                return redirect(next_url or "dashboard")
            if authenticated_user.is_superuser or authenticated_user.is_staff:
                django_login(request, authenticated_user)
                return redirect(next_url or "dashboard")
            messages.error(request, "Your Mining360 profile is inactive or not authorized.")
            return render(request, "reports/login.html", {"next": next_url, "username": username})
        messages.error(request, "Invalid local username or password.")
    return render(request, "reports/login.html", {"next": next_url, "username": request.POST.get("username", "")})


def auth_start(request):
    request.session["login_next"] = request.GET.get("next") or request.POST.get("next") or "/"
    return redirect(login_url(request))


def auth_callback(request):
    expected_state = request.session.get("azure_ad_state")
    received_state = request.GET.get("state")
    if not expected_state or received_state != expected_state:
        messages.error(request, "Invalid Microsoft sign-in state.")
        return redirect("login")
    if request.GET.get("error"):
        messages.error(request, request.GET.get("error_description") or request.GET.get("error"))
        return redirect("login")
    code = request.GET.get("code")
    if not code:
        messages.error(request, "Microsoft sign-in did not return an authorization code.")
        return redirect("login")

    try:
        token = exchange_code(request, code)
        profile = fetch_me(token["access_token"])
    except Exception as exc:
        messages.error(request, str(exc))
        return redirect("login")

    upn = (profile.get("userPrincipalName") or profile.get("mail") or "").strip().lower()
    azure_id = profile.get("id", "")
    if not upn or not azure_id:
        messages.error(request, "Microsoft profile is missing userPrincipalName or id.")
        return redirect("login")

    first_platform_user = not PlatformUser.objects.exists()
    platform_user = PlatformUser.objects.filter(azure_ad_id=azure_id).first() or PlatformUser.objects.filter(user_principal_name=upn).first()
    if not platform_user and first_platform_user:
        platform_user = PlatformUser.objects.create(
            azure_ad_id=azure_id,
            user_principal_name=upn,
            email=profile.get("mail") or upn,
            display_name=profile.get("displayName") or upn,
            job_title=profile.get("jobTitle") or "",
            is_active=True,
            is_platform_admin=True,
            can_access_reporting=True,
            can_access_ai=True,
            can_access_data=True,
            can_access_sources=True,
            business_performance_role="Administrator",
        )
    if not platform_user or not platform_user.is_active:
        messages.error(request, "Your account is not authorized for Mining360.")
        return redirect("login")

    user, _ = User.objects.get_or_create(
        username=upn,
        defaults={
            "email": platform_user.email or upn,
            "first_name": platform_user.display_name[:150],
            "is_staff": platform_user.is_platform_admin,
            "is_superuser": platform_user.is_platform_admin,
        },
    )
    user.email = platform_user.email or upn
    user.first_name = platform_user.display_name[:150]
    user.is_active = platform_user.is_active
    user.is_staff = platform_user.is_platform_admin
    user.is_superuser = platform_user.is_platform_admin
    user.save()
    platform_user.django_user = user
    platform_user.azure_ad_id = azure_id
    platform_user.user_principal_name = upn
    platform_user.email = profile.get("mail") or upn
    platform_user.display_name = profile.get("displayName") or platform_user.display_name
    platform_user.job_title = profile.get("jobTitle") or platform_user.job_title
    platform_user.save()
    django_login(request, user)
    return redirect(request.session.pop("login_next", None) or "dashboard")


def logout_page(request):
    django_logout(request)
    return redirect("login")


@login_required
def users_manage(request):
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    query = request.GET.get("q", "").strip()
    directory_results = []
    search_error = ""
    if query:
        try:
            directory_results = search_directory_users(query)
            for person in directory_results:
                person["primary_email"] = person.get("mail") or person.get("userPrincipalName") or ""
        except Exception as exc:
            search_error = str(exc)
    return render(
        request,
        "reports/users.html",
        {
            "active_section": "users",
            "query": query,
            "directory_results": directory_results,
            "search_error": search_error,
            "platform_users": PlatformUser.objects.all(),
            "role_choices": PlatformUser.ROLE_CHOICES,
            "bp_role_choices": PlatformUser.BUSINESS_PERFORMANCE_ROLES,
            "sidebar_stats": [
                {"label": "Users", "value": PlatformUser.objects.count()},
                {"label": "Auth", "value": "AD"},
            ],
        },
    )


@login_required
def users_add(request):
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    if request.method != "POST":
        return redirect("users-manage")
    azure_id = request.POST.get("azure_ad_id", "").strip()
    upn = request.POST.get("user_principal_name", "").strip().lower()
    display_name = request.POST.get("display_name", "").strip() or upn
    email = request.POST.get("email", "").strip() or upn
    job_title = request.POST.get("job_title", "").strip()
    role_payload = _platform_role_payload(request.POST)
    if not azure_id or not upn:
        messages.error(request, "Missing Azure AD user id or user principal name.")
        return redirect("users-manage")
    platform_user, _ = PlatformUser.objects.update_or_create(
        azure_ad_id=azure_id,
        defaults={
            "user_principal_name": upn,
            "email": email,
            "display_name": display_name,
            "job_title": job_title,
            "is_active": True,
            **role_payload,
        },
    )
    messages.success(request, f"{platform_user.display_name} added to Mining360.")
    return redirect("users-manage")


@login_required
def users_toggle(request, user_id):
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    platform_user = get_object_or_404(PlatformUser, id=user_id)
    if request.method == "POST":
        platform_user.is_active = not platform_user.is_active
        platform_user.save()
        messages.success(request, f"{platform_user.display_name} status updated.")
    return redirect("users-manage")


@login_required
def users_roles_update(request, user_id):
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    if request.method != "POST":
        return redirect("users-manage")
    platform_user = get_object_or_404(PlatformUser, id=user_id)
    for field, value in _platform_role_payload(request.POST).items():
        setattr(platform_user, field, value)
    platform_user.save()
    if platform_user.django_user_id:
        platform_user.django_user.is_staff = platform_user.is_platform_admin
        platform_user.django_user.is_superuser = platform_user.is_platform_admin
        platform_user.django_user.save(update_fields=["is_staff", "is_superuser"])
    messages.success(request, f"{platform_user.display_name} roles updated.")
    return redirect("users-manage")


def dashboard(request):
    module_access = user_module_access(request.user)
    modules = [
        {
            "name": "Reporting",
            "description": "Open and control the focused Power BI reports.",
            "url_name": "reporting",
            "module": "reporting",
        },
        {
            "name": "AI",
            "description": "OpenAI-powered analysis and assistance workflows.",
            "url_name": "ai-home",
            "module": "ai",
        },
        {
            "name": "Sources",
            "description": "Catalog and monitor Mining360 data connections.",
            "url_name": "data-sources",
            "module": "sources",
        },
        {
            "name": "Data",
            "description": "Analyze datasets, run checks and prepare operational data workflows.",
            "url_name": "data-home",
            "module": "data",
        },
        {
            "name": "Resources",
            "description": "Search CAT reference documents and knowledge resources.",
            "url_name": "resources",
        },
    ]
    if _user_is_platform_admin(request.user):
        modules.insert(
            2,
            {
                "name": "IA Config",
                "description": "Configure intent extraction, glossary, mappings and DAX templates.",
                "url_name": "ia-config-home",
            },
        )
    modules = [
        item
        for item in modules
        if (
            (not item.get("module") or module_access.get(item["module"]) or _user_is_platform_admin(request.user))
        )
    ]
    return render(
        request,
        "reports/dashboard.html",
        {
            "modules": modules,
            "active_section": "dashboard",
            "sidebar_stats": [
                {"label": "Modules", "value": len(modules)},
                {"label": "Mode", "value": "Admin"},
            ],
        },
    )


def data_home(request):
    cards = [
        {
            "name": "Data Quality Center",
            "description": "Run quality controls, inspect failed records and export impacted rows.",
            "url_name": "data-quality-center",
        },
        {
            "name": "Source Data Preview",
            "description": "Open registered sources and explore tables, views and custom views.",
            "url_name": "data-sources",
        },
    ]
    browsers = DataBrowser.objects.prefetch_related("columns").all()
    return render(
        request,
        "reports/data_home.html",
        {
            "cards": cards,
            "browsers": [_browser_payload(browser, include_columns=False) for browser in browsers],
            "browser_state": [_browser_payload(browser, include_columns=True) for browser in browsers],
            "active_section": "data",
            "sidebar_stats": [
                {"label": "Workflows", "value": len(cards)},
                {"label": "Focus", "value": "Data"},
            ],
        },
    )


def _request_payload(request) -> dict:
    if request.body:
        try:
            payload = json.loads(request.body.decode("utf-8"))
            if isinstance(payload, dict):
                return payload
        except Exception:
            pass
    return request.POST.dict()


def _browser_column_payload(column: DataBrowserColumn) -> dict:
    return {
        "id": column.id,
        "display_name": column.display_name,
        "sql_name": column.sql_name,
        "data_type": column.data_type,
        "length": column.length,
        "is_required": column.is_required,
        "is_unique": column.is_unique,
        "default_value": column.default_value,
        "display_order": column.display_order,
        "is_visible": column.is_visible,
        "is_lookup": column.is_lookup,
        "lookup_source_name": column.lookup_source_name,
        "lookup_value_column": column.lookup_value_column,
        "lookup_label_column": column.lookup_label_column,
        "lookup_filter": column.lookup_filter,
    }


def _browser_payload(browser: DataBrowser, include_columns: bool = True) -> dict:
    payload = {
        "id": browser.id,
        "name": browser.name,
        "display_order": browser.display_order,
        "description": browser.description,
        "table_name": browser.table_name,
        "source_view_name": browser.source_view_name,
        "is_active": browser.is_active,
        "show_browser_record_id": browser.show_browser_record_id,
        "show_eventchain_id": browser.show_eventchain_id,
        "last_synced_at": browser.last_synced_at.isoformat() if browser.last_synced_at else "",
        "last_sync_status": browser.last_sync_status,
        "last_sync_message": browser.last_sync_message,
        "created_at": browser.created_at.isoformat() if browser.created_at else "",
        "updated_at": browser.updated_at.isoformat() if browser.updated_at else "",
    }
    if include_columns:
        payload["columns"] = [_browser_column_payload(column) for column in browser.columns.all()]
    return payload


def _bool_payload(payload: dict, key: str, default: bool = False) -> bool:
    value = payload.get(key, default)
    if isinstance(value, bool):
        return value
    return str(value).lower() in {"1", "true", "yes", "on"}


def _int_payload(payload: dict, key: str, default: int | None = None) -> int | None:
    value = payload.get(key)
    if value in (None, ""):
        return default
    return int(value)


def _apply_browser_payload(browser: DataBrowser, payload: dict) -> DataBrowser:
    browser.name = str(payload.get("name", browser.name or "")).strip()
    browser.description = str(payload.get("description", browser.description or "")).strip()
    browser.table_name = str(payload.get("table_name", browser.table_name or "")).strip()
    browser.source_view_name = str(payload.get("source_view_name", browser.source_view_name or "")).strip()
    browser.is_active = _bool_payload(payload, "is_active", browser.is_active)
    browser.show_browser_record_id = _bool_payload(
        payload, "show_browser_record_id", browser.show_browser_record_id
    )
    browser.show_eventchain_id = _bool_payload(
        payload, "show_eventchain_id", browser.show_eventchain_id
    )
    if not browser.name:
        raise DataBrowserValidationError("Browser name is required.")
    if not browser.table_name:
        raise DataBrowserValidationError("SQL table name is required.")
    if not browser.source_view_name:
        raise DataBrowserValidationError("Source view name is required.")
    return browser


def _apply_column_payload(column: DataBrowserColumn, payload: dict) -> DataBrowserColumn:
    column.display_name = str(payload.get("display_name", column.display_name or "")).strip()
    column.sql_name = str(payload.get("sql_name", column.sql_name or "")).strip()
    column.data_type = str(payload.get("data_type", column.data_type or "Text")).strip()
    column.length = _int_payload(payload, "length", column.length)
    if "allow_null" in payload:
        column.is_required = not _bool_payload(payload, "allow_null", not column.is_required)
    else:
        column.is_required = _bool_payload(payload, "is_required", column.is_required)
    column.is_unique = _bool_payload(payload, "is_unique", column.is_unique)
    column.default_value = str(payload.get("default_value", column.default_value or "")).strip()
    column.display_order = _int_payload(payload, "display_order", column.display_order or 0) or 0
    column.is_visible = _bool_payload(payload, "is_visible", column.is_visible)
    column.is_lookup = _bool_payload(payload, "is_lookup", column.is_lookup)
    column.lookup_source_name = str(payload.get("lookup_source_name", column.lookup_source_name or "")).strip()
    column.lookup_value_column = str(payload.get("lookup_value_column", column.lookup_value_column or "")).strip()
    column.lookup_label_column = str(payload.get("lookup_label_column", column.lookup_label_column or "")).strip()
    column.lookup_filter = str(payload.get("lookup_filter", column.lookup_filter or "")).strip()
    if not column.display_name:
        raise DataBrowserValidationError("Column display name is required.")
    if not column.sql_name:
        raise DataBrowserValidationError("Column SQL name is required.")
    if column.data_type not in dict(DataBrowserColumn.DATA_TYPES):
        raise DataBrowserValidationError("Invalid column data type.")
    return column


def _json_error(message: str, status: int = 400):
    return JsonResponse({"ok": False, "error": message}, status=status)


@require_http_methods(["GET", "POST"])
def data_browsers_api(request):
    if request.method == "GET":
        browsers = DataBrowser.objects.prefetch_related("columns").all()
        return JsonResponse({"ok": True, "browsers": [_browser_payload(browser, include_columns=False) for browser in browsers]})
    try:
        browser = _apply_browser_payload(DataBrowser(), _request_payload(request))
        browser.display_order = (
            DataBrowser.objects.aggregate(max_order=models.Max("display_order")).get("max_order") or 0
        ) + 1
        browser.save()
        validate_browser_definition(browser)
        sync_browser_sql(browser)
    except Exception as exc:
        if "browser" in locals() and browser.pk:
            browser.delete()
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "browser": _browser_payload(browser)}, status=201)


@require_http_methods(["POST"])
def data_browsers_reorder_api(request):
    payload = _request_payload(request)
    raw_ids = payload.get("browser_ids")
    if not isinstance(raw_ids, list):
        return _json_error("browser_ids must be a list.")
    try:
        browser_ids = [int(browser_id) for browser_id in raw_ids]
    except (TypeError, ValueError):
        return _json_error("Every Browser ID must be an integer.")
    if len(browser_ids) != len(set(browser_ids)):
        return _json_error("The Browser order contains duplicate IDs.")

    browsers = list(DataBrowser.objects.all())
    existing_ids = {browser.id for browser in browsers}
    if set(browser_ids) != existing_ids:
        return _json_error("The submitted order must contain every Browser exactly once.")

    browsers_by_id = {browser.id: browser for browser in browsers}
    with transaction.atomic():
        for position, browser_id in enumerate(browser_ids, start=1):
            browsers_by_id[browser_id].display_order = position
        DataBrowser.objects.bulk_update(browsers, ["display_order"])
    return JsonResponse({
        "ok": True,
        "browsers": [
            _browser_payload(browsers_by_id[browser_id], include_columns=False)
            for browser_id in browser_ids
        ],
    })


@require_http_methods(["GET", "PUT", "DELETE"])
def data_browser_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    if request.method == "GET":
        return JsonResponse({"ok": True, "browser": _browser_payload(browser)})
    if request.method == "DELETE":
        browser.delete()
        return JsonResponse({"ok": True})
    try:
        browser = _apply_browser_payload(browser, _request_payload(request))
        validate_browser_definition(browser)
        browser.save()
        sync_browser_sql(browser)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "browser": _browser_payload(browser)})


@require_http_methods(["POST"])
def data_browser_columns_api(request, browser_id):
    browser = get_object_or_404(DataBrowser, id=browser_id)
    try:
        column = _apply_column_payload(DataBrowserColumn(browser=browser), _request_payload(request))
        column.save()
        validate_browser_definition(browser)
        sync_browser_sql(browser)
    except Exception as exc:
        if "column" in locals() and column.pk:
            column.delete()
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "column": _browser_column_payload(column)}, status=201)


@require_http_methods(["POST"])
def data_browser_columns_reorder_api(request, browser_id):
    browser = get_object_or_404(DataBrowser, id=browser_id)
    payload = _request_payload(request)
    raw_ids = payload.get("column_ids")
    if not isinstance(raw_ids, list):
        return _json_error("column_ids must be a list.")

    try:
        column_ids = [int(column_id) for column_id in raw_ids]
    except (TypeError, ValueError):
        return _json_error("Every column ID must be an integer.")

    if len(column_ids) != len(set(column_ids)):
        return _json_error("The column order contains duplicate IDs.")

    columns = list(browser.columns.all())
    existing_ids = {column.id for column in columns}
    if set(column_ids) != existing_ids:
        return _json_error("The submitted order must contain every Browser column exactly once.")

    columns_by_id = {column.id: column for column in columns}
    with transaction.atomic():
        for position, column_id in enumerate(column_ids, start=1):
            column = columns_by_id[column_id]
            column.display_order = position
        DataBrowserColumn.objects.bulk_update(columns, ["display_order"])

    ordered_columns = [columns_by_id[column_id] for column_id in column_ids]
    return JsonResponse({
        "ok": True,
        "columns": [_browser_column_payload(column) for column in ordered_columns],
    })


@require_http_methods(["PUT", "DELETE"])
def data_browser_column_api(request, browser_id, column_id):
    browser = get_object_or_404(DataBrowser, id=browser_id)
    column = get_object_or_404(DataBrowserColumn, browser=browser, id=column_id)
    if request.method == "DELETE":
        column.delete()
        return JsonResponse({"ok": True})
    try:
        column = _apply_column_payload(column, _request_payload(request))
        column.save()
        validate_browser_definition(browser)
        sync_browser_sql(browser)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "column": _browser_column_payload(column)})


@require_http_methods(["POST"])
def data_browser_sync_sql_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    try:
        result = sync_browser_sql(browser)
    except Exception as exc:
        browser.last_sync_status = "Failed"
        browser.last_sync_message = str(exc)
        browser.last_synced_at = timezone.now()
        browser.save(update_fields=["last_sync_status", "last_sync_message", "last_synced_at", "updated_at"])
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "result": result, "browser": _browser_payload(browser)})


@require_http_methods(["POST"])
def data_browser_import_preview_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return _json_error("Import file is required.")
    try:
        result = inspect_import_file(uploaded_file)
    except Exception as exc:
        return _json_error(str(exc))
    browser_payload = _browser_payload(browser)
    browser_payload["columns"] = [
        {
            "id": column.id,
            "display_name": column.display_name,
            "sql_name": column.sql_name,
            "data_type": column.data_type,
            "is_required": column.is_required,
            "is_unique": column.is_unique,
            "default_value": column.default_value,
            "is_lookup": column.is_lookup,
        }
        for column in browser.columns.all()
    ]
    return JsonResponse({"ok": True, "import_preview": result, "browser": browser_payload})


@require_http_methods(["POST"])
def data_browser_import_batch_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    payload = _request_payload(request)
    token = str(payload.get("token") or "").strip()
    if not token:
        return _json_error("Import token is required.")
    try:
        mapping_raw = payload.get("mapping", "{}")
        mapping = json.loads(mapping_raw) if isinstance(mapping_raw, str) else (mapping_raw if isinstance(mapping_raw, dict) else {})
    except Exception:
        mapping = {}
    mapping["_audit_user"] = request.user.get_username() if request.user.is_authenticated else "System"
    try:
        result = import_browser_records_batch(
            browser,
            token,
            mapping=mapping,
            start=int(payload.get("start", 0) or 0),
            batch_size=int(payload.get("batch_size", 50) or 50),
        )
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "import": result})


@require_http_methods(["POST"])
def data_browser_import_start_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    payload = _request_payload(request)
    token = str(payload.get("token") or "").strip()
    if not token:
        return _json_error("Import token is required.")
    try:
        mapping_raw = payload.get("mapping", {})
        mapping = json.loads(mapping_raw) if isinstance(mapping_raw, str) else (mapping_raw if isinstance(mapping_raw, dict) else {})
        mapping["_audit_user"] = request.user.get_username() if request.user.is_authenticated else "System"
        result = start_import_job(browser, token, mapping=mapping)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "import": result})


@require_http_methods(["GET"])
def data_browser_import_status_api(request, job_token):
    try:
        result = get_import_job_status(job_token)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "status": result})


@require_http_methods(["GET"])
def data_browser_data_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    try:
        result = preview_browser_data(
            browser,
            limit=request.GET.get("limit", "100"),
            filters=request.GET.get("filters", "[]"),
        )
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "data": result})


@require_http_methods(["GET"])
def data_browser_export_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    try:
        payload = preview_browser_data(
            browser,
            limit=request.GET.get("limit", "1000"),
            filters=request.GET.get("filters", "[]"),
        )
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    try:
        from io import BytesIO
        from openpyxl import Workbook
    except Exception as exc:
        return JsonResponse({"ok": False, "error": f"Excel export unavailable: {exc}"}, status=500)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (browser.name or "Browser")[:31]
    columns = list(payload.get("columns") or [])
    rows = list(payload.get("rows") or [])
    if columns:
        sheet.append(columns)
        for record in rows:
            sheet.append([_json_safe(record.get(column)) for column in columns])
        if not rows:
            sheet.append(["No rows returned"])
    else:
        sheet.append(["No rows returned"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"{browser.table_name or browser.name or 'browser'}_export.xlsx".replace(" ", "_")
    response = FileResponse(buffer, as_attachment=True, filename=filename)
    response["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@require_http_methods(["POST"])
def data_browser_records_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    try:
        actor = request.user.get_username() if request.user.is_authenticated else "System"
        result = insert_browser_record(browser, _request_payload(request), actor=actor)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "record": result}, status=201)


@require_http_methods(["PUT", "DELETE"])
def data_browser_record_api(request, browser_id, record_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    try:
        if request.method == "PUT":
            actor = request.user.get_username() if request.user.is_authenticated else "System"
            result = update_browser_record(browser, record_id, _request_payload(request), actor=actor)
        else:
            result = delete_browser_records(browser, [record_id])
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "record": result})


@require_http_methods(["DELETE", "POST"])
def data_browser_records_bulk_delete_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    payload = _request_payload(request)
    try:
        result = delete_browser_records(browser, payload.get("record_ids") or payload.get("ids") or [])
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "result": result})


@require_http_methods(["POST"])
def data_browser_import_api(request, browser_id):
    browser = get_object_or_404(DataBrowser.objects.prefetch_related("columns"), id=browser_id)
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return _json_error("Import file is required.")
    try:
        mapping_raw = request.POST.get("mapping", "{}")
        try:
            mapping = json.loads(mapping_raw) if mapping_raw else {}
        except Exception:
            mapping = {}
        result = import_browser_records(browser, uploaded_file, mapping=mapping)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "import": result})


@require_http_methods(["GET"])
def data_browser_lookup_options_api(request, browser_id, column_id):
    browser = get_object_or_404(DataBrowser, id=browser_id)
    column = get_object_or_404(DataBrowserColumn, browser=browser, id=column_id)
    try:
        raw_limit = str(request.GET.get("limit", "500") or "500").strip().lower()
        limit = None if raw_limit == "all" else int(raw_limit)
        result = lookup_options(column, limit=limit)
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "lookup": result})


def _latest_data_quality_run(source_key: str | None = None, object_name: str | None = None):
    queryset = DataQualityRun.objects.all()
    if source_key:
        queryset = queryset.filter(source_key=source_key)
    if object_name:
        queryset = queryset.filter(object_name=object_name)
    return queryset.order_by("-created_at").first()


def _data_quality_controls_state(source_key: str | None = None, object_name: str | None = None):
    latest_run = _latest_data_quality_run(source_key, object_name)
    latest_results = latest_run.results if latest_run and isinstance(latest_run.results, list) else []
    results_by_key = {item.get("key"): item for item in latest_results if isinstance(item, dict)}
    controls = []
    for control in available_checks():
        result = results_by_key.get(control.key, {})
        controls.append(
            {
                "key": control.key,
                "name": control.name,
                "category": control.category,
                "description": control.description,
                "status": result.get("status", "Not run"),
                "impacted_records": result.get("impacted_records", 0),
                "error_percentage": result.get("error_percentage", 0),
                "execution_ms": result.get("execution_ms", 0),
                "affected_columns": result.get("affected_columns", []),
                "latest_run_id": latest_run.id if latest_run else None,
                "has_records": bool(result.get("records")),
            }
        )
    history = DataQualityRun.objects.all()
    if source_key:
        history = history.filter(source_key=source_key)
    if object_name:
        history = history.filter(object_name=object_name)
    return latest_run, controls, history[:12]


def reporting_home(request):
    reports = []
    error = None

    try:
        reports = list_workspace_reports_with_refresh()
    except Exception as exc:
        error = str(exc)
    preferences = {
        item.report_id: item
        for item in ReportingReportPreference.objects.filter(
            report_id__in=[str(getattr(report, "id", "") or "") for report in reports]
        )
    }
    reports = [
        report
        for report in reports
        if preferences.get(str(getattr(report, "id", "") or "")) is None
        or preferences[str(getattr(report, "id", "") or "")].is_visible
    ]
    reports.sort(
        key=lambda report: (
            preferences.get(str(getattr(report, "id", "") or "")).display_order
            if preferences.get(str(getattr(report, "id", "") or ""))
            else 100000,
            str(getattr(report, "display_name", "") or "").lower(),
        )
    )
    completed_count = sum(
        1 for report in reports
        if str(getattr(report, "refresh_status", "") or "").lower() == "completed"
    )
    failed_count = sum(
        1 for report in reports
        if str(getattr(report, "refresh_status", "") or "").lower() == "failed"
    )
    no_refresh_count = len(reports) - completed_count - failed_count
    report_cards = []
    for report in reports:
        name = str(report.display_name or "").lower()
        if "fuel" in name:
            visual_class = "fuel"
        elif "logistic" in name or "aftermarket" in name or "parts" in name:
            visual_class = "logistics"
        elif "operator" in name:
            visual_class = "operator"
        elif "sos" in name:
            visual_class = "sos"
        elif "connected" in name:
            visual_class = "connected"
        elif "prime mover" in name:
            visual_class = "prime"
        elif "monthly" in name or "customer" in name:
            visual_class = "customer"
        elif "lcc" in name or "poca" in name:
            visual_class = "financial"
        else:
            visual_class = "performance"
        report_cards.append({
            "id": report.id,
            "name": report.name,
            "display_name": report.display_name,
            "dataset_id": report.dataset_id,
            "web_url": report.web_url,
            "embed_url": report.embed_url,
            "report_type": report.report_type,
            "last_refresh": report.last_refresh,
            "refresh_status": report.refresh_status,
            "visual_class": visual_class,
        })
    reports = report_cards

    return render(
        request,
        "reports/home.html",
        {
            "reports": reports,
            "report_count": len(reports),
            "completed_count": completed_count,
            "failed_count": failed_count,
            "no_refresh_count": no_refresh_count,
            "error": error,
            "workspace_name": "Efficience Mine Workspace",
            "active_section": "reporting",
            "sidebar_stats": [
                {"label": "Reports", "value": len(reports)},
                {"label": "Embed", "value": "On"},
            ],
        },
    )


@ensure_csrf_cookie
def data_sources(request):
    sources = list_live_sources()
    source_cards = []
    for source in sources:
        source_cards.append(
            {
                "source": source,
                "inventory": _source_inventory(source),
            }
        )

    return render(
        request,
        "reports/data_sources.html",
        {
            "active_section": "sources",
            "source_count": len(sources),
            "source_cards": source_cards,
            "sidebar_stats": [
                {"label": "Sources", "value": len(sources)},
                {"label": "Mode", "value": "Live"},
            ],
        },
    )


def data_quality_center(request):
    source_key = (request.GET.get("source_key") or "").strip()
    object_name = (request.GET.get("object_name") or "").strip()
    preview_url = (request.GET.get("preview_url") or "").strip()
    category = (request.GET.get("category") or "").strip()
    query = (request.GET.get("q") or "").strip().lower()

    latest_run, controls, history = _data_quality_controls_state(source_key or None, object_name or None)
    if not preview_url and latest_run and isinstance(latest_run.request_payload, dict):
        preview_url = str(latest_run.request_payload.get("preview_url") or "").strip()

    if category:
        controls = [item for item in controls if item["category"].lower() == category.lower()]
    if query:
        controls = [
            item
            for item in controls
            if query in item["name"].lower()
            or query in item["description"].lower()
            or query in item["status"].lower()
        ]

    score = float(latest_run.score) if latest_run else 100.0
    run_status = latest_run.status if latest_run else "No run"
    last_run_at = latest_run.created_at if latest_run else None
    source_label = latest_run.source_name if latest_run else (source_key or "All sources")
    object_label = latest_run.object_name if latest_run else (object_name or "No object selected")

    return render(
        request,
        "reports/data_quality_center.html",
        {
            "active_section": "sources",
            "dq_controls": controls,
            "dq_history": history,
            "dq_score": score,
            "dq_last_run": last_run_at,
            "dq_run_status": run_status,
            "dq_source_label": source_label,
            "dq_object_label": object_label,
            "dq_filter_query": query,
            "dq_filter_category": category,
            "dq_source_key": source_key,
            "dq_object_name": object_name,
            "dq_preview_url": preview_url,
            "dq_categories": sorted({item["category"] for item in controls}),
            "sidebar_stats": [
                {"label": "Score", "value": f"{score:.0f}%"},
                {"label": "Controls", "value": len(available_checks())},
            ],
        },
    )


def _extract_connection_details(engine: str, data) -> tuple[dict[str, str], dict[str, object], str | None]:
    normalized_engine = (engine or "SQL Server").strip()
    details: dict[str, str] = {}
    connection_args: dict[str, object] = {}
    error = None

    if normalized_engine == "Snowflake":
        account = data.get("account", "").strip()
        warehouse = data.get("warehouse", "").strip()
        role = data.get("role", "").strip()
        database = data.get("database", "").strip()
        default_schema = data.get("default_schema", "").strip()
        user = data.get("user", "").strip()
        password = data.get("password", "").strip()

        if not account or not warehouse or not user:
            error = "Account, warehouse and user are required for Snowflake."
        else:
            details = {
                "account": _normalize_snowflake_account(account),
                "warehouse": warehouse,
                "role": role,
                "default_schema": default_schema,
            }
            connection_args = {
                "server": _normalize_snowflake_account(account),
                "database": database,
                "user": user,
                "password": password,
                "port": 0,
            }
    else:
        server = data.get("server", "").strip()
        database = data.get("database", "").strip()
        user = data.get("user", "").strip()
        password = data.get("password", "").strip()
        port_raw = data.get("port", "").strip()
        port = int(port_raw) if port_raw else 0

        if not server:
            error = "Instance/server is required for SQL Server."
        else:
            details = {
                "port": str(port or 0),
            }
            connection_args = {
                "server": server,
                "database": database or "master",
                "database_raw": database,
                "user": user,
                "password": password,
                "port": port,
            }

    return details, connection_args, error


@ensure_csrf_cookie
def source_add(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        engine = request.POST.get("engine", "SQL Server").strip() or "SQL Server"
        owner = request.POST.get("owner", "").strip()
        description = request.POST.get("description", "").strip()
        is_active = request.POST.get("is_active") == "on"
        error = None
        connection_details, connection_args, connection_error = _extract_connection_details(engine, request.POST)

        if not name:
            error = "Source name is required."
        elif connection_error:
            error = connection_error
        else:
            source = add_live_source(
                name=name,
                engine=engine,
                server=str(connection_args["server"]),
                database=str(connection_args.get("database_raw", connection_args.get("database", "")) or ""),
                user=str(connection_args["user"]),
                password=str(connection_args["password"]),
                port=int(connection_args["port"]),
                owner=owner,
                description=description,
                connection_details=connection_details,
                status="Active" if is_active else "Unknown",
                status_class="success" if is_active else "neutral",
                verification_message="Created manually. Verify the source to test the connection.",
            )
            return redirect("data-sources")

        return render(
            request,
            "reports/source_add.html",
            {
                "active_section": "sources",
                "error": error,
                "form": request.POST,
                "sidebar_stats": [
                    {"label": "Mode", "value": "Create"},
                    {"label": "Validate", "value": "Live"},
                ],
            },
        )

    return render(
        request,
        "reports/source_add.html",
        {
            "active_section": "sources",
            "sidebar_stats": [
                {"label": "Mode", "value": "Create"},
                {"label": "Validate", "value": "Live"},
            ],
        },
    )


@ensure_csrf_cookie
def source_edit(request, source_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    if request.method == "POST":
        password = request.POST.get("password", "").strip()
        connection_details, connection_args, connection_error = _extract_connection_details(source.engine, request.POST)
        if connection_error:
            return render(
                request,
                "reports/source_edit.html",
                {
                    "active_section": "sources",
                    "source": source,
                    "error": connection_error,
                    "sidebar_stats": [
                        {"label": "Views", "value": len(source.views)},
                        {"label": "Engine", "value": source.engine},
                    ],
                },
            )
        updated = update_live_source(
            source.key,
            server=str(connection_args["server"]),
            database=str(connection_args.get("database_raw", connection_args.get("database", "")) or ""),
            user=str(connection_args["user"]),
            password=password or None,
            port=int(connection_args["port"]),
            owner=request.POST.get("owner", source.owner).strip(),
            description=request.POST.get("description", source.description).strip(),
            connection_details=connection_details or source.connection_details,
            is_active=request.POST.get("is_active") == "on",
        )
        return redirect("source-detail", source_key=updated.key)

    return render(
        request,
        "reports/source_edit.html",
        {
            "active_section": "sources",
            "source": source,
            "sidebar_stats": [
                {"label": "Views", "value": len(source.views)},
                {"label": "Engine", "value": source.engine},
            ],
        },
    )


@ensure_csrf_cookie
@require_http_methods(["POST"])
def source_database_update(request, source_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    database = request.POST.get("database", "").strip()
    if not database:
        if _is_ajax_request(request):
            return JsonResponse({"ok": False, "error": "Database is required."}, status=400)
        messages.error(request, "Database is required.")
        return redirect("source-detail", source_key=source.key)

    updated = update_live_source(source.key, database=database)
    inventory = _source_inventory(updated)
    catalog = _source_catalog(updated)
    if _is_ajax_request(request):
        return JsonResponse(
            {
                "ok": True,
                "source": {
                    "key": updated.key,
                    "engine": updated.engine,
                    "database": updated.database,
                    "status": updated.verification_status,
                    "status_class": updated.status_class,
                    "last_verified": updated.last_verified,
                },
                "inventory": inventory,
                "catalog": catalog,
            }
        )

    messages.success(request, f"Database changed to {updated.database}.")
    return redirect("source-detail", source_key=source.key)


@ensure_csrf_cookie
@require_http_methods(["POST"])
def source_custom_view_add(request, source_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    name = request.POST.get("name", "").strip()
    description = request.POST.get("description", "").strip()
    sql = request.POST.get("sql", "").strip()

    if not name or not sql:
        error = "Custom view name and SQL are required."
        if _is_ajax_request(request):
            return JsonResponse({"ok": False, "error": error}, status=400)
        messages.error(request, error)
        return redirect("source-detail", source_key=source.key)

    view = add_live_view(source.key, name=name, description=description, sql=sql)
    inventory = _source_inventory(get_live_source(source.key))
    catalog = _source_catalog(get_live_source(source.key))
    if _is_ajax_request(request):
        return JsonResponse(
            {
                "ok": True,
                "view": {
                    "key": view.key,
                    "name": view.name,
                    "description": view.description,
                },
                "source": {
                    "key": source.key,
                    "engine": source.engine,
                    "database": source.database,
                },
                "inventory": inventory,
                "catalog": catalog,
            }
        )

    messages.success(request, f"Custom view '{view.name}' created.")
    return redirect("source-detail", source_key=source.key)


@ensure_csrf_cookie
@require_http_methods(["GET", "POST", "PUT"])
def source_custom_view_edit(request, source_key, view_key):
    try:
        source, view = get_live_view(source_key, view_key)
    except KeyError:
        raise Http404("Custom view not found")

    if request.method == "GET":
        return JsonResponse(
            {
                "ok": True,
                "view": {
                    "key": view.key,
                    "name": view.name,
                    "description": view.description,
                    "sql": view.sql,
                },
            }
        )

    payload = _request_payload(request)
    name = str(payload.get("name", view.name)).strip()
    description = str(payload.get("description", view.description)).strip()
    sql = str(payload.get("sql", view.sql)).strip()
    if not name or not sql:
        error = "Custom view name and SQL are required."
        if _is_ajax_request(request):
            return JsonResponse({"ok": False, "error": error}, status=400)
        messages.error(request, error)
        return redirect("source-detail", source_key=source.key)

    updated_view = update_live_view(source.key, view.key, name=name, description=description, sql=sql)
    updated_source = get_live_source(source.key)
    inventory = _source_inventory(updated_source)
    catalog = _source_catalog(updated_source)
    if _is_ajax_request(request):
        return JsonResponse(
            {
                "ok": True,
                "view": {
                    "key": updated_view.key,
                    "name": updated_view.name,
                    "description": updated_view.description,
                    "sql": updated_view.sql,
                },
                "source": {
                    "key": updated_source.key,
                    "engine": updated_source.engine,
                    "database": updated_source.database,
                },
                "inventory": inventory,
                "catalog": catalog,
            }
        )

    messages.success(request, f"Custom view '{updated_view.name}' updated.")
    return redirect("source-detail", source_key=source.key)


@ensure_csrf_cookie
@require_http_methods(["POST", "DELETE"])
def source_custom_view_delete(request, source_key, view_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    deleted = delete_live_view(source.key, view_key)
    if not deleted:
        raise Http404("Custom view not found")

    updated_source = get_live_source(source.key)
    inventory = _source_inventory(updated_source)
    catalog = _source_catalog(updated_source)
    if _is_ajax_request(request):
        return JsonResponse(
            {
                "ok": True,
                "deleted": True,
                "source": {
                    "key": updated_source.key,
                    "engine": updated_source.engine,
                    "database": updated_source.database,
                },
                "inventory": inventory,
                "catalog": catalog,
            }
        )

    messages.success(request, "Custom view deleted.")
    return redirect("source-detail", source_key=source.key)


def source_verify(request, source_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    is_ajax = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in request.headers.get("Accept", "")
    )

    if source.engine.lower() != "sql server":
        if snowflake_connector is None:
            updated = set_live_source_verification(
                source.key,
                False,
                "Snowflake connector is not installed.",
            )
        else:
            try:
                connect_kwargs = {
                    "account": _normalize_snowflake_account(source.connection_details.get("account") or source.server),
                    "user": source.user,
                    "password": source.password or None,
                    "warehouse": source.connection_details.get("warehouse"),
                    "role": source.connection_details.get("role") or None,
                }
                database = source.database or source.connection_details.get("database") or None
                if database:
                    connect_kwargs["database"] = database
                connection = snowflake_connector.connect(**connect_kwargs)
                try:
                    connection.cursor().execute("SELECT 1").fetchone()
                finally:
                    connection.close()
                updated = set_live_source_verification(source.key, True, "Connection OK")
            except Exception as exc:
                updated = set_live_source_verification(source.key, False, str(exc))
        if is_ajax:
            return JsonResponse(
                {
                    "ok": updated.verification_status.lower() == "active",
                    "source": {
                        "key": updated.key,
                        "status": updated.verification_status,
                        "status_class": updated.status_class,
                        "last_verified": updated.last_verified,
                        "message": updated.verification_message,
                    },
                },
                status=200 if updated.verification_status.lower() == "active" else 400,
            )
        return redirect("source-detail", source_key=source.key)

    try:
        with connect(
            server=source.server,
            database=source.database or "master",
            user=source.user or None,
            password=source.password or None,
            port=source.port or None,
        ) as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT 1")
            cursor.fetchone()
        updated = set_live_source_verification(source.key, True, "Connection OK")
        if is_ajax:
            return JsonResponse(
                {
                    "ok": True,
                    "source": {
                        "key": updated.key,
                        "status": updated.verification_status,
                        "status_class": updated.status_class,
                        "last_verified": updated.last_verified,
                        "message": updated.verification_message,
                    },
                }
            )
    except Exception as exc:
        updated = set_live_source_verification(source.key, False, str(exc))
        if is_ajax:
            return JsonResponse(
                {
                    "ok": False,
                    "source": {
                        "key": updated.key,
                        "status": updated.verification_status,
                        "status_class": updated.status_class,
                        "last_verified": updated.last_verified,
                        "message": updated.verification_message,
                    },
                },
                status=400,
            )

    return redirect("source-detail", source_key=source.key)


@ensure_csrf_cookie
def source_detail(request, source_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    inventory = _source_inventory(source)
    catalog = _source_catalog(source)
    available_databases = _source_databases(source)
    dq_available_controls = [
        {
            "key": control.key,
            "name": control.name,
            "category": control.category,
            "description": control.description,
        }
        for control in available_checks()
    ]

    return render(
        request,
        "reports/source_detail.html",
        {
            "active_section": "sources",
            "source": source,
            "inventory": inventory,
            "catalog": catalog,
            "available_databases": available_databases,
            "dq_available_controls": dq_available_controls,
            "table_preview": None,
            "selected_table": "",
            "sidebar_stats": [
                {"label": "Views", "value": len(source.views)},
                {"label": "Engine", "value": source.engine},
            ],
        },
    )


def source_table_preview(request, source_key, table_name):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    if source.engine.lower() != "sql server":
        return JsonResponse({"ok": False, "error": "Table preview is only available for SQL Server sources."}, status=400)

    if not _is_ajax_request(request):
        raise Http404("Not found")

    filters = _parse_preview_filters(request.GET.get("filters", ""))

    try:
        table_name = _normalize_object_identifier(table_name)
        schema, table = [part.strip("[] ") for part in table_name.split(".", 1)]
        object_sql = f"{_escape_sql_identifier(schema)}.{_escape_sql_identifier(table)}"
        preview = _fetch_query_preview(
            source,
            _build_preview_sql(object_sql, filters, request.GET.get("limit", "1000")),
            table_name,
        )
        return JsonResponse({"ok": True, "preview": preview})
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


def source_object_preview(request, source_key, kind, identifier):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    if not _is_ajax_request(request):
        raise Http404("Not found")

    kind = (kind or "").strip().lower()
    limit = request.GET.get("limit", "1000")
    filters = _parse_preview_filters(request.GET.get("filters", ""))

    try:
        identifier = _normalize_object_identifier(identifier)
        if kind in {"table", "tables", "view", "views"}:
            object_sql = ".".join(_escape_sql_identifier(part.strip("[] ")) for part in identifier.split(".", 1))
            preview = _fetch_query_preview(source, _build_preview_sql(object_sql, filters, limit), identifier)
        elif kind in {"custom", "custom-view", "custom_view"}:
            view_source, view = get_live_view(source.key, identifier)
            object_sql = f"(\n{view.sql}\n) AS preview_view"
            preview = _fetch_query_preview(view_source, _build_preview_sql(object_sql, filters, limit), view.name)
        else:
            raise ValueError("Unknown object type.")
        return JsonResponse({"ok": True, "preview": preview})
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


def source_preview_export(request, source_key):
    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    preview_url = (request.GET.get("preview_url") or "").strip()
    if not preview_url:
        raise Http404("Preview target not found")

    filters = _parse_preview_filters(request.GET.get("filters", ""))

    try:
        payload = _preview_export_rows(source, preview_url, filters)
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    try:
        from io import BytesIO
        from openpyxl import Workbook
    except Exception as exc:
        return JsonResponse({"ok": False, "error": f"Excel export unavailable: {exc}"}, status=500)

    workbook = Workbook()
    sheet = workbook.active
    title = (payload.get("object_name") or "Preview")[:31]
    sheet.title = title or "Preview"

    preview = payload.get("preview") or {}
    columns = list(preview.get("columns") or [])
    rows = list(preview.get("rows") or [])
    if columns:
        sheet.append(columns)
        if rows:
            for record in rows:
                sheet.append([_json_safe(record.get(column)) for column in columns])
        else:
            sheet.append(["No rows returned"])
    else:
        sheet.append(["No rows returned"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"{source.key}_{title or 'preview'}.xlsx".replace(" ", "_")
    response = FileResponse(buffer, as_attachment=True, filename=filename)
    response["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


def source_delete(request, source_key):
    if request.method != "POST":
        raise Http404("Source not found")

    try:
        source = get_live_source(source_key)
    except KeyError:
        raise Http404("Source not found")

    is_ajax = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or "application/json" in request.headers.get("Accept", "")
    )

    deleted = delete_live_source(source.key)
    if is_ajax:
        return JsonResponse(
            {
                "ok": bool(deleted),
                "deleted": bool(deleted),
                "source": {"key": source.key, "name": source.name},
            },
            status=200 if deleted else 404,
        )

    return redirect("data-sources")


def source_view(request, source_key, view_key):
    try:
        source, view = get_live_view(source_key, view_key)
    except KeyError:
        raise Http404("View not found")

    limit = request.GET.get("limit", "200")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    quick_range = request.GET.get("quick_range", "")
    error = None
    result = None

    def _valid_date(value: str) -> str:
        if not value:
            return ""
        return date.fromisoformat(value).isoformat()

    try:
        normalized_from = _valid_date(date_from)
        normalized_to = _valid_date(date_to)
    except ValueError:
        error = "Invalid date filter format. Use YYYY-MM-DD."
        normalized_from = ""
        normalized_to = ""

    today = date.today()

    if quick_range and not error:
        if quick_range == "mtd":
            normalized_from = date(today.year, today.month, 1).isoformat()
            normalized_to = today.isoformat()
        elif quick_range == "ytd":
            normalized_from = date(today.year, 1, 1).isoformat()
            normalized_to = today.isoformat()
        elif quick_range == "last-12-months":
            month = today.month - 11
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            normalized_from = date(year, month, 1).isoformat()
            normalized_to = today.isoformat()
        elif quick_range == "last-year":
            normalized_from = date(today.year - 1, 1, 1).isoformat()
            normalized_to = date(today.year - 1, 12, 31).isoformat()
        else:
            error = "Unknown quick range selected."

    if not error:
        try:
            result = execute_live_view(
                source.key,
                view.key,
                int(limit),
                date_from=normalized_from or None,
                date_to=normalized_to or None,
            )
        except Exception as exc:
            error = str(exc)

    return render(
        request,
        "reports/source_view.html",
        {
            "active_section": "sources",
            "source": source,
            "view": view,
            "limit": limit,
            "date_from": date_from,
            "date_to": date_to,
            "quick_range": quick_range,
            "result": result,
            "error": error,
            "sidebar_stats": [
                {"label": "Rows", "value": result["row_count"] if result else 0},
                {"label": "Engine", "value": source.engine},
            ],
        },
    )


MONTH_ALIASES = {
    "janvier": 1,
    "january": 1,
    "fevrier": 2,
    "février": 2,
    "february": 2,
    "mars": 3,
    "march": 3,
    "avril": 4,
    "april": 4,
    "mai": 5,
    "may": 5,
    "juin": 6,
    "june": 6,
    "juillet": 7,
    "july": 7,
    "aout": 8,
    "août": 8,
    "august": 8,
    "septembre": 9,
    "september": 9,
    "octobre": 10,
    "october": 10,
    "novembre": 11,
    "november": 11,
    "decembre": 12,
    "décembre": 12,
    "december": 12,
}


def _extract_flow_rows(flow_result: dict) -> list[dict]:
    rows = flow_result.get("firstTableRows")
    if isinstance(rows, list):
        return rows
    try:
        return flow_result["results"][0]["tables"][0]["rows"]
    except Exception:
        return []


def _format_pct(value) -> str:
    if value is None or value == "":
        return "BLANK"
    try:
        return f"{float(value) * 100:.2f}%"
    except Exception:
        return str(value)


def _parse_semantic_question(question: str) -> dict:
    text = (question or "").strip()
    lowered = text.lower()
    parsed = {
        "dataset": "FPR Global DB + RLS",
        "site": "",
        "model": "",
        "year": 2026,
        "month": 5,
        "mode": "single",
        "months": 12,
    }
    site_match = re.search(r"(?:minesite|mine ?site|site)\s*[:=]\s*([a-z0-9 /_-]+)", lowered, re.I)
    if not site_match:
        site_match = re.search(r"\b(?:a|à|at|pour|sur)\s+([a-z][a-z0-9 /_-]+?)\s*(?:,| en | au | sur | pour |$)", lowered, re.I)
    if site_match:
        parsed["site"] = site_match.group(1).strip().title()

    model_match = re.search(r"(?:model|mod[eè]le)\s*[:=]?\s*([a-z0-9. -]+)", lowered, re.I)
    if not model_match:
        model_match = re.search(r"\b(6015|6020|6030|6040|6050|777 wt|777|785|789|d10|d9|992|390|395|980|988|844)\b", lowered, re.I)
    if model_match:
        parsed["model"] = model_match.group(1).strip().upper()

    if "tous les model" in lowered or "all model" in lowered or "tous les modèles" in lowered:
        parsed["mode"] = "matrix"
        parsed["model"] = ""
    if "12 derniers mois" in lowered or "douze derniers mois" in lowered or "last 12" in lowered:
        parsed["mode"] = "matrix"
        parsed["months"] = 12

    for name, number in MONTH_ALIASES.items():
        if re.search(rf"\b{re.escape(name)}\b", lowered):
            parsed["month"] = number
            break
    year_match = re.search(r"\b(20\d{2})\b", lowered)
    if year_match:
        parsed["year"] = int(year_match.group(1))

    if not parsed["site"]:
        parsed["site"] = "Fekola"
    return parsed


def _resolve_semantic_rls_role(dataset_name: str, site: str) -> str:
    candidate = str(site or "").strip()
    if not candidate:
        return ""
    resolved_roles = resolve_dataset_roles(dataset_name, [candidate])
    if resolved_roles:
        return resolved_roles[0]
    return candidate


def _sanitize_conversation(payload) -> list[dict]:
    if not isinstance(payload, list):
        return []
    cleaned = []
    for item in payload[-12:]:
        if not isinstance(item, dict):
            continue
        role = str(item.get("role") or "").strip().lower()
        content = str(item.get("content") or "").strip()
        if role in {"user", "assistant"} and content:
            cleaned.append({"role": role, "content": content})
    return cleaned


def _build_single_answer(semantic_request: dict, rows: list[dict]) -> dict:
    row = rows[0] if rows else {}
    value = row.get("[Value]")
    answer = f"{semantic_request['measure']} = {_format_pct(value)}"
    if value is None:
        interpretation = "No value was returned for this context by the semantic model."
    else:
        pct = float(value) * 100
        if pct >= 90:
            interpretation = "Availability is high for this context."
        elif pct >= 80:
            interpretation = "Availability is acceptable but requires operational monitoring."
        elif pct >= 70:
            interpretation = "Availability is low and requires analysis."
        else:
            interpretation = "Availability is critical; investigate downtime and major events."
    return {
        "answer": answer,
        "interpretation": interpretation,
        "rows": rows,
    }


def _build_matrix_answer(semantic_request: dict, rows: list[dict]) -> dict:
    grouped = {}
    for row in rows:
        model = row.get("EquipmentList_MiningProd[Model]", "")
        value = row.get("[Availability]")
        if not model or value in (None, ""):
            continue
        grouped.setdefault(model, []).append(
            {
                "month": row.get("Date[Year Month]", ""),
                "month_number": row.get("Date[Year Month Number]", ""),
                "availability": float(value),
            }
        )
    summary = []
    for model, values in grouped.items():
        avg = sum(item["availability"] for item in values) / len(values)
        worst = min(values, key=lambda item: item["availability"])
        best = max(values, key=lambda item: item["availability"])
        summary.append(
            {
                "model": model,
                "months": len(values),
                "average": avg,
                "average_display": _format_pct(avg),
                "worst_month": worst["month"],
                "worst_value": worst["availability"],
                "worst_display": _format_pct(worst["availability"]),
                "best_month": best["month"],
                "best_value": best["availability"],
                "best_display": _format_pct(best["availability"]),
            }
        )
    summary.sort(key=lambda item: item["average"])
    weak_models = [item for item in summary if item["average"] < 0.8]
    interpretation = (
        f"{len(summary)} models have values for the period. "
        f"{len(weak_models)} models average below 80%. "
    )
    if weak_models:
        interpretation += "Analysis priorities are: " + ", ".join(item["model"] for item in weak_models[:6]) + "."
    else:
        interpretation += "Average availability is generally under control."
    return {
        "answer": f"{semantic_request['measure']} by model for {semantic_request['filters'].get('MineSiteList_MiningProd[MineSite]', '')}",
        "interpretation": interpretation,
        "summary": summary,
        "rows": rows,
    }


@ensure_csrf_cookie
def ai_home(request):
    openai_enabled = is_openai_configured()
    from .agent_router_service import multi_agent_enabled
    from .models import AIAgent

    agents_enabled = multi_agent_enabled(request.user)
    return render(
        request,
        "reports/ai.html",
        {
            "active_section": "ai",
            "openai_enabled": openai_enabled,
            "is_platform_admin": _user_is_platform_admin(request.user),
            "multi_agent_enabled": agents_enabled,
            "available_agents": (
                AIAgent.objects.filter(active=True).order_by("-priority", "name")
                if agents_enabled
                else []
            ),
            "sidebar_stats": [
                {"label": "Mode", "value": "AI" if openai_enabled else "Rules"},
                {"label": "Model", "value": "OpenAI" if openai_enabled else "Semantic"},
            ],
        },
    )


def ai_semantic_test(request):
    dataset_name = request.GET.get("dataset", "FPR Global DB + RLS")
    year = int(request.GET.get("year", "2026"))
    month = int(request.GET.get("month", "5"))
    model = request.GET.get("model", "777")
    site = request.GET.get("site", "Fekola")

    semantic_request = build_availability_question(dataset_name, year, month, model, site)
    dataset_id = semantic_request["dataset_id"] or resolve_workspace_dataset_id(dataset_name)
    semantic_request["dataset_id"] = dataset_id
    semantic_request["rls_role"] = _resolve_semantic_rls_role(dataset_name, site)
    dax_query = semantic_request["dax"]

    flow_payload = {
        "datasetId": dataset_id,
        "datasetName": dataset_name,
        "query": dax_query,
        "question": semantic_request["question"],
        "metric": semantic_request["metric"],
        "measure": semantic_request["measure"],
        "filters": semantic_request["filters"],
        "period": semantic_request["period"],
        "rlsRole": semantic_request["rls_role"],
        "roles": [semantic_request["rls_role"]] if semantic_request["rls_role"] else [],
    }
    try:
        flow_result = execute_dax_via_flow(flow_payload)
        return JsonResponse(
            {
                "ok": True,
                "execution": "power_automate",
                "semantic_request": semantic_request,
                "flow_result": flow_result,
            }
        )
    except Exception as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": str(exc),
                "execution": "power_automate",
                "semantic_request": semantic_request,
                "flow_payload": flow_payload,
            },
            status=500,
        )


def ai_ask(request):
    started_at = time.monotonic()
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required."}, status=405)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Invalid JSON payload."}, status=400)

    question = (payload.get("question") or "").strip()
    section_code = (payload.get("section_code") or payload.get("section") or "").strip() or None
    conversation = _sanitize_conversation(payload.get("conversation"))
    if not question:
        return JsonResponse({"ok": False, "error": "Question is required."}, status=400)

    from .agent_router_service import multi_agent_enabled
    if multi_agent_enabled(request.user):
        try:
            from .ai_agent_execution_service import execute_agent_question

            agent_result = execute_agent_question(
                question,
                user=request.user,
                conversation_id=str(payload.get("conversation_id") or "").strip(),
                messages=conversation,
                manual_agent=str(payload.get("agent_selection") or "auto"),
                section_code=section_code,
                dataset_name=(payload.get("dataset_name") or "FPR Global DB + RLS").strip(),
                debug_mode=_user_is_platform_admin(request.user),
            )
            if agent_result.get("ok"):
                answer_text = agent_result.get("chat_message") or agent_result.get("answer") or ""
                if isinstance(answer_text, dict):
                    answer_text = answer_text.get("answer") or answer_text.get("interpretation") or ""
                return JsonResponse({
                    **agent_result,
                    "chat_message": answer_text,
                    "answer": {
                        "answer": answer_text,
                        "interpretation": answer_text,
                        "rows": agent_result.get("rows") or [],
                        "summary": agent_result.get("rows") or [],
                    },
                })
        except Exception:
            # The existing single-pipeline chatbot remains the compatibility fallback.
            pass

    from .chat_routing_service import (
        answer_without_semantic_model,
        classify_chat_question,
    )

    routing = classify_chat_question(question, section_code=section_code)
    if not routing["requires_semantic_model"]:
        chat_message = answer_without_semantic_model(question, routing)
        return JsonResponse({
            "ok": True,
            "chat_message": chat_message,
            "answer": {
                "answer": chat_message,
                "interpretation": chat_message,
                "rows": [],
                "summary": [],
            },
            "route": routing["route"],
            "routing_reason": routing["reason"],
            "semantic_model_queried": False,
        })

    has_validated_report = PowerBIReport.objects.filter(
        is_active=True,
        validation_status="Validated",
    ).exists()
    try:
        orchestrated = process_user_question(
            question,
            user_context={
                "user": request.user,
                "section_code": section_code,
                "dataset_name": (payload.get("dataset_name") or "FPR Global DB + RLS").strip(),
                "open_report": has_validated_report,
                "debug_mode": _user_is_platform_admin(request.user),
            },
            conversation_context={
                "conversation_id": str(payload.get("conversation_id") or "").strip(),
                "messages": conversation,
            },
        )
        if not orchestrated.get("ok"):
            clarification = orchestrated.get("clarification_question")
            if clarification:
                return JsonResponse({
                    **orchestrated,
                    "ok": True,
                    "chat_message": clarification,
                    "answer": {
                        "answer": clarification,
                        "interpretation": clarification,
                        "rows": [],
                        "summary": [],
                    },
                })
            return JsonResponse(orchestrated, status=400)
        return JsonResponse({
            **orchestrated,
            "chat_message": orchestrated.get("answer"),
            "route": routing["route"],
            "routing_reason": routing["reason"],
            "semantic_model_queried": True,
            "answer": {
                "answer": orchestrated.get("answer"),
                "interpretation": orchestrated.get("answer"),
                "rows": orchestrated.get("rows") or [],
                "summary": orchestrated.get("rows") or [],
            },
        })
    except Exception:
        # Keep the legacy semantic path available while the controlled
        # orchestrator is being completed for a new intent family.
        pass

    intent = extract_intent(question, section_code)
    valid, validation_errors = validate_intent(intent)
    if not valid:
        return JsonResponse(
            {
                "ok": False,
                "error": "Intent validation failed.",
                "validation": {"valid": False, "errors": validation_errors},
                "intent": intent,
            },
            status=400,
        )

    try:
        dax_payload = generate_dax_from_intent(intent)
    except IntentValidationError as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": str(exc),
                "validation": {"valid": False, "errors": [str(exc)]},
                "intent": intent,
            },
            status=400,
        )
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc), "intent": intent}, status=500)

    dataset_name = (payload.get("dataset_name") or "FPR Global DB + RLS").strip()
    dataset_id = resolve_workspace_dataset_id(dataset_name)
    section = get_section_by_code(dax_payload["section"])
    rls_role = _resolve_semantic_rls_role(dataset_name, intent.get("filters", {}).get("minesite") or intent.get("filters", {}).get("site") or "")
    flow_payload = {
        "datasetId": dataset_id,
        "datasetName": dataset_name,
        "query": dax_payload["dax"],
        "question": question,
        "metric": dax_payload["metric"],
        "measure": dax_payload["measure"],
        "filters": dax_payload["filters"],
        "section": dax_payload["section"],
        "intent": intent,
        "rlsRole": rls_role,
        "roles": [rls_role] if rls_role else [],
    }

    try:
        flow_result = execute_dax_via_flow(flow_payload)
    except Exception as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": str(exc),
                "intent": intent,
                "dax": dax_payload["dax"],
                "flow_payload": flow_payload,
            },
            status=500,
        )

    rows = _extract_flow_rows(flow_result)
    answer = {
        "answer": "",
        "interpretation": "",
        "rows": rows,
        "summary": rows,
    }
    if rows:
        measure_key_names = {dax_payload["metric_label"].lower(), dax_payload["measure"].strip("[]").lower(), "value"}
        measured_rows = []
        for row in rows:
            for key, value in row.items():
                key_norm = str(key).strip("[]").lower()
                if key_norm in measure_key_names:
                    try:
                        numeric_value = float(value)
                    except Exception:
                        numeric_value = None
                    measured_rows.append({"row": row, "value": value, "numeric_value": numeric_value})
                    break

        numeric_rows = [item for item in measured_rows if item["numeric_value"] is not None]
        if len(numeric_rows) > 1:
            avg = sum(item["numeric_value"] for item in numeric_rows) / len(numeric_rows)
            worst = min(numeric_rows, key=lambda item: item["numeric_value"])
            best = max(numeric_rows, key=lambda item: item["numeric_value"])
            answer["answer"] = f"{dax_payload['metric_label']} average = {_format_pct(avg)}"
            answer["interpretation"] = (
                f"{len(numeric_rows)} data points returned. "
                f"Moyenne: {_format_pct(avg)}. "
                f"Plus faible: {_format_pct(worst['numeric_value'])}. "
                f"Meilleur: {_format_pct(best['numeric_value'])}."
            )
            answer["summary"] = {
                "row_count": len(rows),
                "measured_points": len(numeric_rows),
                "average": avg,
                "average_display": _format_pct(avg),
                "worst": worst["row"],
                "best": best["row"],
            }
        elif measured_rows:
            measure_value = measured_rows[0]["value"]
            answer["answer"] = f"{dax_payload['metric_label']} = {_format_pct(measure_value)}"
            try:
                pct = float(measure_value) * 100
            except Exception:
                pct = None
            if pct is None:
                answer["interpretation"] = "The result was returned successfully."
            elif pct >= 90:
                answer["interpretation"] = "The value is high for this context."
            elif pct >= 80:
                answer["interpretation"] = "The value is acceptable but requires monitoring."
            elif pct >= 70:
                answer["interpretation"] = "The value is low and requires analysis."
            else:
                answer["interpretation"] = "The value is critical and requires investigation."
        else:
            answer["interpretation"] = "The model returned no usable value for this context."
    else:
        answer["interpretation"] = "The model returned no data."

    try:
        metric_code_for_target = dax_payload["metric"]
        target = AIKPITarget.objects.filter(section__code=dax_payload["section"], metric_code=metric_code_for_target, is_active=True).first()
        numeric_value = None
        if isinstance(answer.get("summary"), dict) and answer["summary"].get("average") is not None:
            numeric_value = float(answer["summary"]["average"])
        elif rows:
            for row in rows:
                for value in row.values():
                    try:
                        numeric_value = float(value)
                        break
                    except Exception:
                        continue
                if numeric_value is not None:
                    break
        if target and numeric_value is not None:
            target_payload = {
                "target": float(target.target),
                "warning_threshold": float(target.warning_threshold),
                "critical_threshold": float(target.critical_threshold),
                "unit": target.unit,
            }
            if numeric_value < float(target.critical_threshold):
                target_status = "Critical"
            elif numeric_value < float(target.warning_threshold):
                target_status = "Warning"
            elif numeric_value >= float(target.target):
                target_status = "OK"
            else:
                target_status = "Watch"
            actions = [
                item.recommendations
                for item in AIRecommendedAction.objects.filter(
                    section__code=dax_payload["section"],
                    metric_code=metric_code_for_target,
                    is_active=True,
                ).order_by("priority")
            ]
            answer["kpi_target"] = target_payload
            answer["kpi_status"] = target_status
            answer["recommended_actions"] = actions
            answer["interpretation"] = f"{answer['interpretation']} KPI status: {target_status}."
    except Exception:
        pass

    final_response = answer["interpretation"]
    try:
        final_response = generate_chat_response(question, intent, answer, conversation)
    except Exception:
        final_response = answer["interpretation"]

    execution_time_ms = int((time.monotonic() - started_at) * 1000)
    try:
        intent_prompt = get_prompt_template(dax_payload["section"], "intent_extraction") or {}
        response_prompt = get_prompt_template(dax_payload["section"], "response_generation") or {}
        prompt_sent = (
            "Intent Extraction Template:\n"
            f"{intent_prompt.get('prompt_template', '')}\n\n"
            "Response Generation Template:\n"
            f"{response_prompt.get('prompt_template', '')}"
        ).strip()
        AIDebugRun.objects.create(
            question_text=question,
            detected_section=str(intent.get("section") or ""),
            extracted_intent=intent,
            prompt_sent=prompt_sent,
            generated_json=intent,
            generated_dax=dax_payload["dax"],
            powerbi_response=flow_result if isinstance(flow_result, dict) else {"raw": str(flow_result)},
            formatted_response=final_response,
            execution_time_ms=execution_time_ms,
            token_usage={},
            errors="",
        )
        KnowledgeAILog.objects.create(
            user_question=question,
            detected_section=str(intent.get("section") or ""),
            extracted_intent=intent,
            generated_dax=dax_payload["dax"],
            powerbi_result=flow_result if isinstance(flow_result, dict) else {"raw": str(flow_result)},
            final_answer=final_response,
            status="Completed",
            execution_time_ms=execution_time_ms,
            token_usage={},
            user=request.user if request.user.is_authenticated else None,
        )
    except Exception:
        pass

    return JsonResponse(
        {
            "ok": True,
            "intent": intent,
            "validation": {"valid": True, "errors": []},
            "dax": dax_payload["dax"],
            "metric": dax_payload["metric"],
            "measure": dax_payload["measure"],
            "flow_payload": flow_payload,
            "raw_result": flow_result,
            "answer": answer,
            "chat_message": final_response,
            "debug": {
                "question": question,
                "json_intent": intent,
                "dax": dax_payload["dax"],
                "raw_powerbi": flow_result,
                "final_response": final_response,
                "execution_time_ms": execution_time_ms,
            },
        }
    )


IA_RESOURCE_TYPES = {
    "question-examples": {
        "model": AIQuestionExample,
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "question_text": item.question_text,
            "language": item.language,
            "expected_json_intent": item.expected_json_intent,
            "is_active": item.is_active,
        },
    },
    "synonyms": {
        "model": AISynonym,
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "entity_type": item.entity_type,
            "canonical_value": item.canonical_value,
            "synonym_value": item.synonym_value,
            "language": item.language,
            "is_active": item.is_active,
        },
    },
    "metrics": {
        "model": AIMetricMapping,
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "metric_code": item.metric_code,
            "metric_label": item.metric_label,
            "powerbi_measure_name": item.powerbi_measure_name,
            "description": item.description,
            "is_active": item.is_active,
        },
    },
    "filters": {
        "model": AIFilterMapping,
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "filter_code": item.filter_code,
            "filter_label": item.filter_label,
            "powerbi_table_name": item.powerbi_table_name,
            "powerbi_column_name": item.powerbi_column_name,
            "data_type": item.data_type,
            "is_required": item.is_required,
            "is_active": item.is_active,
        },
    },
    "dax-templates": {
        "model": AIDaxTemplate,
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "template_name": item.template_name,
            "template_code": item.template_code,
            "dax_template": item.dax_template,
            "description": item.description,
            "is_active": item.is_active,
        },
    },
    "semantic-tables": {
        "model": AISemanticTable,
        "search_fields": ["table_name", "display_name", "description"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "table_name": item.table_name,
            "display_name": item.display_name,
            "description": item.description,
            "is_active": item.is_active,
        },
    },
    "semantic-columns": {
        "model": AISemanticColumn,
        "search_fields": ["table_name", "column_name", "display_name", "data_type", "description"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "table_name": item.table_name,
            "column_name": item.column_name,
            "display_name": item.display_name,
            "data_type": item.data_type,
            "description": item.description,
            "is_filter": item.is_filter,
            "is_active": item.is_active,
        },
    },
    "semantic-measures": {
        "model": AISemanticMeasure,
        "search_fields": ["measure_name", "display_name", "description", "dax_name", "unit", "category"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "measure_name": item.measure_name,
            "display_name": item.display_name,
            "description": item.description,
            "dax_name": item.dax_name,
            "unit": item.unit,
            "category": item.category,
            "is_active": item.is_active,
        },
    },
    "semantic-relationships": {
        "model": AISemanticRelationship,
        "search_fields": ["parent_table", "parent_column", "child_table", "child_column", "relationship_type"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "parent_table": item.parent_table,
            "parent_column": item.parent_column,
            "child_table": item.child_table,
            "child_column": item.child_column,
            "relationship_type": item.relationship_type,
            "is_active": item.is_active,
        },
    },
    "business-vocabulary": {
        "model": AIBusinessVocabulary,
        "search_fields": ["business_term", "business_definition", "category"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "business_term": item.business_term,
            "business_definition": item.business_definition,
            "category": item.category,
            "is_active": item.is_active,
        },
    },
    "few-shot-examples": {
        "model": AIFewShotExample,
        "search_fields": ["question", "expected_dax", "expected_response", "explanation"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "question": item.question,
            "expected_json_intent": item.expected_json_intent,
            "expected_dax": item.expected_dax,
            "expected_response": item.expected_response,
            "explanation": item.explanation,
            "is_active": item.is_active,
        },
    },
    "prompt-templates": {
        "model": AIPromptTemplate,
        "search_fields": ["prompt_type", "template_name", "prompt_template", "description"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "prompt_type": item.prompt_type,
            "template_name": item.template_name,
            "prompt_template": item.prompt_template,
            "description": item.description,
            "is_active": item.is_active,
        },
    },
    "business-rules": {
        "model": AIBusinessRule,
        "search_fields": ["metric_code", "rule_name", "condition", "action", "default_value"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "metric_code": item.metric_code,
            "rule_name": item.rule_name,
            "condition": item.condition,
            "action": item.action,
            "default_value": item.default_value,
            "priority": item.priority,
            "is_active": item.is_active,
        },
    },
    "powerbi-pages": {
        "model": AIPowerBIPage,
        "search_fields": ["page_name", "report_name", "report_id", "page_display_name", "description"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "page_name": item.page_name,
            "report_name": item.report_name,
            "report_id": item.report_id,
            "page_display_name": item.page_display_name,
            "description": item.description,
            "is_default_page": item.is_default_page,
            "is_active": item.is_active,
        },
    },
    "visual-mapping": {
        "model": AIVisualMapping,
        "search_fields": ["metric_code", "recommended_visual", "description"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "metric_code": item.metric_code,
            "recommended_visual": item.recommended_visual,
            "description": item.description,
            "priority": item.priority,
            "is_active": item.is_active,
        },
    },
    "kpi-targets": {
        "model": AIKPITarget,
        "search_fields": ["metric_code", "unit"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "metric_code": item.metric_code,
            "target": str(item.target),
            "warning_threshold": str(item.warning_threshold),
            "critical_threshold": str(item.critical_threshold),
            "unit": item.unit,
            "is_active": item.is_active,
        },
    },
    "recommended-actions": {
        "model": AIRecommendedAction,
        "search_fields": ["metric_code", "condition", "recommendations"],
        "serializer": lambda item: {
            "id": item.id,
            "section": item.section.code,
            "metric_code": item.metric_code,
            "condition": item.condition,
            "recommendations": item.recommendations,
            "priority": item.priority,
            "is_active": item.is_active,
        },
    },
    "debug-runs": {
        "model": AIDebugRun,
        "admin_only": True,
        "search_fields": ["question_text", "detected_section", "generated_dax", "formatted_response", "errors"],
        "serializer": lambda item: {
            "id": item.id,
            "created_at": item.created_at.isoformat() if item.created_at else "",
            "question_text": item.question_text,
            "detected_section": item.detected_section,
            "extracted_intent": item.extracted_intent,
            "prompt_sent": item.prompt_sent,
            "generated_json": item.generated_json,
            "generated_dax": item.generated_dax,
            "powerbi_response": item.powerbi_response,
            "formatted_response": item.formatted_response,
            "execution_time_ms": item.execution_time_ms,
            "token_usage": item.token_usage,
            "errors": item.errors,
        },
    },
}


def _ia_section_payload(section: AIConfigSection) -> dict:
    return {
        "id": section.id,
        "name": section.name,
        "code": section.code,
        "description": section.description,
        "is_active": section.is_active,
    }


def _ia_get_section_or_404(section_code: str) -> AIConfigSection:
    section = get_object_or_404(AIConfigSection, code=section_code)
    return section


def _ia_resource_queryset(section: AIConfigSection | None, resource_type: str):
    model = IA_RESOURCE_TYPES[resource_type]["model"]
    if IA_RESOURCE_TYPES[resource_type].get("admin_only"):
        return model.objects.all()
    return model.objects.filter(section=section)


def _ia_normalize_bool(value, default=False) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return default
    return str(value).lower() in {"1", "true", "yes", "on"}


def _ia_payload(request) -> dict:
    return _request_payload(request)


def _ia_text(payload: dict, item, field: str, default: str = "") -> str:
    return str(payload.get(field, getattr(item, field, default) or default)).strip()


def _ia_int(payload: dict, item, field: str, default: int = 100) -> int:
    value = payload.get(field, getattr(item, field, default))
    try:
        return int(value)
    except Exception:
        return default


def _ia_decimal_text(payload: dict, item, field: str, default: str = "0"):
    value = str(payload.get(field, getattr(item, field, default) or default)).strip()
    return value or default


def _ia_json_object(payload: dict, item, field: str) -> dict:
    value = payload.get(field, getattr(item, field, {}) or {})
    if isinstance(value, str):
        try:
            value = json.loads(value) if value.strip() else {}
        except Exception as exc:
            raise ValueError(f"{field} must be valid JSON.") from exc
    if not isinstance(value, dict):
        raise ValueError(f"{field} must be a JSON object.")
    return value


def _ia_apply_resource_payload(resource_type: str, item, payload: dict, section: AIConfigSection):
    if resource_type == "question-examples":
        item.question_text = str(payload.get("question_text", item.question_text or "")).strip()
        item.language = str(payload.get("language", item.language or "fr")).strip() or "fr"
        raw_json = payload.get("expected_json_intent", item.expected_json_intent or {})
        if isinstance(raw_json, str):
            try:
                raw_json = json.loads(raw_json) if raw_json.strip() else {}
            except Exception as exc:
                raise ValueError("Expected JSON Intent must be valid JSON.") from exc
        if not isinstance(raw_json, dict):
            raise ValueError("Expected JSON Intent must be a JSON object.")
        item.expected_json_intent = raw_json
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        item.section = section
        if not item.question_text:
            raise ValueError("Question text is required.")
        return item

    if resource_type == "synonyms":
        item.entity_type = str(payload.get("entity_type", item.entity_type or "")).strip()
        item.canonical_value = str(payload.get("canonical_value", item.canonical_value or "")).strip()
        item.synonym_value = str(payload.get("synonym_value", item.synonym_value or "")).strip()
        item.language = str(payload.get("language", item.language or "fr")).strip() or "fr"
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        item.section = section
        if not item.entity_type or not item.canonical_value or not item.synonym_value:
            raise ValueError("Entity type, canonical value and synonym value are required.")
        return item

    if resource_type == "metrics":
        item.metric_code = str(payload.get("metric_code", item.metric_code or "")).strip()
        item.metric_label = str(payload.get("metric_label", item.metric_label or "")).strip()
        item.powerbi_measure_name = str(payload.get("powerbi_measure_name", item.powerbi_measure_name or "")).strip()
        item.description = str(payload.get("description", item.description or "")).strip()
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        item.section = section
        if not item.metric_code or not item.metric_label or not item.powerbi_measure_name:
            raise ValueError("Metric code, label and Power BI measure name are required.")
        return item

    if resource_type == "filters":
        item.filter_code = str(payload.get("filter_code", item.filter_code or "")).strip()
        item.filter_label = str(payload.get("filter_label", item.filter_label or "")).strip()
        item.powerbi_table_name = str(payload.get("powerbi_table_name", item.powerbi_table_name or "")).strip()
        item.powerbi_column_name = str(payload.get("powerbi_column_name", item.powerbi_column_name or "")).strip()
        item.data_type = str(payload.get("data_type", item.data_type or "Text")).strip() or "Text"
        item.is_required = _ia_normalize_bool(payload.get("is_required"), item.is_required)
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        item.section = section
        if not item.filter_code or not item.filter_label or not item.powerbi_table_name or not item.powerbi_column_name:
            raise ValueError("Filter code, label, table name and column name are required.")
        return item

    if resource_type == "dax-templates":
        item.template_name = str(payload.get("template_name", item.template_name or "")).strip()
        item.template_code = str(payload.get("template_code", item.template_code or "")).strip()
        item.dax_template = str(payload.get("dax_template", item.dax_template or "")).strip()
        item.description = str(payload.get("description", item.description or "")).strip()
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        item.section = section
        if not item.template_name or not item.template_code or not item.dax_template:
            raise ValueError("Template name, template code and DAX template are required.")
        return item

    if resource_type == "semantic-tables":
        item.section = section
        item.table_name = _ia_text(payload, item, "table_name")
        item.display_name = _ia_text(payload, item, "display_name", item.table_name)
        item.description = _ia_text(payload, item, "description")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.table_name or not item.display_name:
            raise ValueError("Table name and display name are required.")
        return item

    if resource_type == "semantic-columns":
        item.section = section
        item.table_name = _ia_text(payload, item, "table_name")
        item.column_name = _ia_text(payload, item, "column_name")
        item.display_name = _ia_text(payload, item, "display_name", item.column_name)
        item.data_type = _ia_text(payload, item, "data_type")
        item.description = _ia_text(payload, item, "description")
        item.is_filter = _ia_normalize_bool(payload.get("is_filter"), item.is_filter)
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.table_name or not item.column_name or not item.display_name:
            raise ValueError("Table, column name and display name are required.")
        return item

    if resource_type == "semantic-measures":
        item.section = section
        item.measure_name = _ia_text(payload, item, "measure_name")
        item.display_name = _ia_text(payload, item, "display_name", item.measure_name)
        item.description = _ia_text(payload, item, "description")
        item.dax_name = _ia_text(payload, item, "dax_name", f"[{item.measure_name}]")
        item.unit = _ia_text(payload, item, "unit")
        item.category = _ia_text(payload, item, "category")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.measure_name or not item.display_name or not item.dax_name:
            raise ValueError("Measure name, display name and DAX name are required.")
        return item

    if resource_type == "semantic-relationships":
        item.section = section
        item.parent_table = _ia_text(payload, item, "parent_table")
        item.parent_column = _ia_text(payload, item, "parent_column")
        item.child_table = _ia_text(payload, item, "child_table")
        item.child_column = _ia_text(payload, item, "child_column")
        item.relationship_type = _ia_text(payload, item, "relationship_type")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.parent_table or not item.parent_column or not item.child_table or not item.child_column:
            raise ValueError("Parent table/column and child table/column are required.")
        return item

    if resource_type == "business-vocabulary":
        item.section = section
        item.business_term = _ia_text(payload, item, "business_term")
        item.business_definition = _ia_text(payload, item, "business_definition")
        item.category = _ia_text(payload, item, "category")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.business_term or not item.business_definition:
            raise ValueError("Business term and definition are required.")
        return item

    if resource_type == "few-shot-examples":
        item.section = section
        item.question = _ia_text(payload, item, "question")
        item.expected_json_intent = _ia_json_object(payload, item, "expected_json_intent")
        item.expected_dax = _ia_text(payload, item, "expected_dax")
        item.expected_response = _ia_text(payload, item, "expected_response")
        item.explanation = _ia_text(payload, item, "explanation")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.question:
            raise ValueError("Question is required.")
        return item

    if resource_type == "prompt-templates":
        item.section = section
        item.prompt_type = _ia_text(payload, item, "prompt_type")
        item.template_name = _ia_text(payload, item, "template_name")
        item.prompt_template = _ia_text(payload, item, "prompt_template")
        item.description = _ia_text(payload, item, "description")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.prompt_type or not item.template_name or not item.prompt_template:
            raise ValueError("Prompt type, template name and prompt template are required.")
        return item

    if resource_type == "business-rules":
        item.section = section
        item.metric_code = _ia_text(payload, item, "metric_code")
        item.rule_name = _ia_text(payload, item, "rule_name")
        item.condition = _ia_text(payload, item, "condition")
        item.action = _ia_text(payload, item, "action")
        item.default_value = _ia_text(payload, item, "default_value")
        item.priority = _ia_int(payload, item, "priority")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.metric_code or not item.rule_name or not item.condition or not item.action:
            raise ValueError("Metric, rule name, condition and action are required.")
        return item

    if resource_type == "powerbi-pages":
        item.section = section
        item.page_name = _ia_text(payload, item, "page_name")
        item.report_name = _ia_text(payload, item, "report_name")
        item.report_id = _ia_text(payload, item, "report_id")
        item.page_display_name = _ia_text(payload, item, "page_display_name", item.page_name)
        item.description = _ia_text(payload, item, "description")
        item.is_default_page = _ia_normalize_bool(payload.get("is_default_page"), item.is_default_page)
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.page_name or not item.report_name or not item.page_display_name:
            raise ValueError("Page name, report name and page display name are required.")
        return item

    if resource_type == "visual-mapping":
        item.section = section
        item.metric_code = _ia_text(payload, item, "metric_code")
        item.recommended_visual = _ia_text(payload, item, "recommended_visual")
        item.description = _ia_text(payload, item, "description")
        item.priority = _ia_int(payload, item, "priority")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.metric_code or not item.recommended_visual:
            raise ValueError("Metric and recommended visual are required.")
        return item

    if resource_type == "kpi-targets":
        item.section = section
        item.metric_code = _ia_text(payload, item, "metric_code")
        item.target = _ia_decimal_text(payload, item, "target")
        item.warning_threshold = _ia_decimal_text(payload, item, "warning_threshold")
        item.critical_threshold = _ia_decimal_text(payload, item, "critical_threshold")
        item.unit = _ia_text(payload, item, "unit")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.metric_code:
            raise ValueError("Metric is required.")
        return item

    if resource_type == "recommended-actions":
        item.section = section
        item.metric_code = _ia_text(payload, item, "metric_code")
        item.condition = _ia_text(payload, item, "condition")
        item.recommendations = _ia_text(payload, item, "recommendations")
        item.priority = _ia_int(payload, item, "priority")
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.metric_code or not item.condition or not item.recommendations:
            raise ValueError("Metric, condition and recommendations are required.")
        return item

    raise ValueError("Unsupported IA Config resource type.")


@require_http_methods(["GET"])
def ia_config_home(request):
    sections = get_active_sections()
    return render(
        request,
        "reports/ia_config.html",
        {
            "active_section": "ia-config",
            "sections": sections,
            "is_ia_admin": request.user.is_staff or request.user.is_superuser,
            "sidebar_stats": [
                {"label": "Sections", "value": len(sections)},
                {"label": "Mode", "value": "AI Config"},
            ],
        },
    )


@require_http_methods(["GET"])
def ia_config_sections_api(request):
    return JsonResponse({"ok": True, "sections": get_active_sections()})


@require_http_methods(["GET", "POST"])
def ia_config_collection_api(request, section_code, resource_type):
    if resource_type not in IA_RESOURCE_TYPES:
        return _json_error("Unsupported IA Config resource type.", status=404)
    config = IA_RESOURCE_TYPES[resource_type]
    if config.get("admin_only") and not (request.user.is_staff or request.user.is_superuser):
        return _json_error("Administrator access required.", status=403)
    section = None if config.get("admin_only") else _ia_get_section_or_404(section_code)
    model = IA_RESOURCE_TYPES[resource_type]["model"]
    serializer = IA_RESOURCE_TYPES[resource_type]["serializer"]
    queryset = _ia_resource_queryset(section, resource_type)

    if request.method == "GET":
        query = request.GET.get("q", "").strip().lower()
        active = request.GET.get("active", "").strip().lower()
        if active in {"1", "true", "yes"} and hasattr(model, "is_active"):
            queryset = queryset.filter(is_active=True)
        elif active in {"0", "false", "no"} and hasattr(model, "is_active"):
            queryset = queryset.filter(is_active=False)
        if query:
            if resource_type == "question-examples":
                queryset = queryset.filter(question_text__icontains=query)
            elif resource_type == "synonyms":
                queryset = queryset.filter(
                    models.Q(entity_type__icontains=query)
                    | models.Q(canonical_value__icontains=query)
                    | models.Q(synonym_value__icontains=query)
                )
            elif resource_type == "metrics":
                queryset = queryset.filter(
                    models.Q(metric_code__icontains=query)
                    | models.Q(metric_label__icontains=query)
                    | models.Q(powerbi_measure_name__icontains=query)
                )
            elif resource_type == "filters":
                queryset = queryset.filter(
                    models.Q(filter_code__icontains=query)
                    | models.Q(filter_label__icontains=query)
                    | models.Q(powerbi_table_name__icontains=query)
                    | models.Q(powerbi_column_name__icontains=query)
                )
            elif resource_type == "dax-templates":
                queryset = queryset.filter(
                    models.Q(template_name__icontains=query)
                    | models.Q(template_code__icontains=query)
                    | models.Q(description__icontains=query)
                )
            else:
                search_fields = IA_RESOURCE_TYPES[resource_type].get("search_fields", [])
                if search_fields:
                    q_filter = models.Q()
                    for field in search_fields:
                        q_filter |= models.Q(**{f"{field}__icontains": query})
                    queryset = queryset.filter(q_filter)
        order_field = "-created_at" if resource_type == "debug-runs" else "-updated_at"
        items = [serializer(item) for item in queryset.order_by(order_field)]
        section_payload = _ia_section_payload(section) if section else {"code": section_code}
        return JsonResponse({"ok": True, "section": section_payload, "items": items})

    if config.get("admin_only"):
        return _json_error("This resource is read-only.", status=405)
    payload = _ia_payload(request)
    try:
        item = model(section=section)
        item = _ia_apply_resource_payload(resource_type, item, payload, section)
        item.save()
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "item": serializer(item)}, status=201)


@require_http_methods(["PUT", "DELETE"])
def ia_config_item_api(request, section_code, resource_type, item_id):
    if resource_type not in IA_RESOURCE_TYPES:
        return _json_error("Unsupported IA Config resource type.", status=404)
    config = IA_RESOURCE_TYPES[resource_type]
    if config.get("admin_only") and not (request.user.is_staff or request.user.is_superuser):
        return _json_error("Administrator access required.", status=403)
    section = None if config.get("admin_only") else _ia_get_section_or_404(section_code)
    model = config["model"]
    serializer = config["serializer"]
    item = get_object_or_404(model, id=item_id) if config.get("admin_only") else get_object_or_404(model, section=section, id=item_id)

    if request.method == "DELETE":
        if config.get("admin_only"):
            return _json_error("This resource is read-only.", status=405)
        item.delete()
        return JsonResponse({"ok": True, "deleted": True})

    if config.get("admin_only"):
        return _json_error("This resource is read-only.", status=405)
    payload = _ia_payload(request)
    try:
        item = _ia_apply_resource_payload(resource_type, item, payload, section)
        item.save()
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "item": serializer(item)})


@require_http_methods(["POST"])
def ia_config_test_intent_api(request):
    payload = _ia_payload(request)
    question_text = str(payload.get("question_text", "")).strip()
    section_code = str(payload.get("section_code", "")).strip() or None
    if not question_text:
        return _json_error("Question text is required.")
    intent = extract_intent(question_text, section_code)
    valid, validation_errors = validate_intent(intent)
    response = {
        "ok": True,
        "question_text": question_text,
        "intent": intent,
        "validation": {
            "valid": valid,
            "errors": validation_errors,
        },
    }
    if valid:
        try:
            dax_payload = generate_dax_from_intent(intent)
            response["dax"] = dax_payload["dax"]
            response["metric"] = dax_payload["metric"]
            response["measure"] = dax_payload["measure"]
            response["template_code"] = dax_payload["template_code"]
        except IntentValidationError as exc:
            response["validation"] = {"valid": False, "errors": [str(exc)]}
            return JsonResponse(response, status=400)
        except Exception as exc:
            return _json_error(str(exc))
    return JsonResponse(response)


def _ai_row_value(row: dict, *names: str):
    normalized = {
        re.sub(r"[^a-z0-9]+", "", str(key).lower()): value
        for key, value in row.items()
    }
    for name in names:
        key = re.sub(r"[^a-z0-9]+", "", str(name).lower())
        if key in normalized:
            return normalized[key]
    return ""


def _execute_powerbi_info_query(dataset_id: str, info_name: str) -> tuple[list[dict], str]:
    query = f"EVALUATE INFO.{info_name}()"
    try:
        return execute_dataset_dax(dataset_id, query), ""
    except Exception as exc:
        return [], str(exc)


@require_http_methods(["POST"])
def ia_config_import_semantic_model_api(request, section_code):
    section = _ia_get_section_or_404(section_code)
    payload = _ia_payload(request)
    dataset_name = str(payload.get("dataset_name") or "FPR Global DB + RLS").strip()
    try:
        dataset_id = str(payload.get("dataset_id") or "").strip() or resolve_workspace_dataset_id(dataset_name)
    except Exception as exc:
        return _json_error(str(exc), status=400)
    workspace_id = env_value("POWERBI_WORKSPACE_ID")
    imported_at = timezone.now()

    imported = {"tables": 0, "columns": 0, "measures": 0, "relationships": 0}
    errors = {}

    tables, error = _execute_powerbi_info_query(dataset_id, "TABLES")
    if error:
        errors["tables"] = error
    for row in tables:
        table_name = str(_ai_row_value(row, "Name", "Table", "TableName") or "").strip()
        if not table_name:
            continue
        AISemanticTable.objects.update_or_create(
            section=section,
            table_name=table_name,
            defaults={
                "display_name": table_name,
                "description": str(_ai_row_value(row, "Description") or ""),
                "source_report": dataset_name,
                "dataset_id": dataset_id,
                "workspace_id": workspace_id,
                "imported_at": imported_at,
                "business_description": str(_ai_row_value(row, "Description") or ""),
                "validation_status": "Imported",
                "is_active": True,
            },
        )
        imported["tables"] += 1

    columns, error = _execute_powerbi_info_query(dataset_id, "COLUMNS")
    if error:
        errors["columns"] = error
    for row in columns:
        table_name = str(_ai_row_value(row, "Table", "TableName", "ExplicitDataTypeTable", "SourceTable") or "").strip()
        column_name = str(_ai_row_value(row, "Name", "Column", "ColumnName") or "").strip()
        if not table_name or not column_name:
            continue
        AISemanticColumn.objects.update_or_create(
            section=section,
            table_name=table_name,
            column_name=column_name,
            defaults={
                "display_name": column_name,
                "data_type": str(_ai_row_value(row, "DataType", "Type", "ExplicitDataType") or ""),
                "description": str(_ai_row_value(row, "Description") or ""),
                "source_report": dataset_name,
                "dataset_id": dataset_id,
                "workspace_id": workspace_id,
                "imported_at": imported_at,
                "business_description": str(_ai_row_value(row, "Description") or ""),
                "validation_status": "Imported",
                "is_filter": False,
                "is_active": True,
            },
        )
        imported["columns"] += 1

    measures, error = _execute_powerbi_info_query(dataset_id, "MEASURES")
    if error:
        try:
            measures = discover_dataset_measures_rest(dataset_id)
            error = ""
        except Exception as exc:
            measures = []
            error = str(exc)
    if error:
        errors["measures"] = error
    for row in measures:
        measure_name = str(_ai_row_value(row, "Name", "Measure", "MeasureName") or "").strip()
        if not measure_name:
            continue
        AISemanticMeasure.objects.update_or_create(
            section=section,
            measure_name=measure_name,
            defaults={
                "display_name": measure_name,
                "description": str(_ai_row_value(row, "Description") or ""),
                "dax_name": f"[{measure_name}]",
                "unit": "",
                "category": str(_ai_row_value(row, "Display Folder", "DisplayFolder", "Category") or ""),
                "source_report": dataset_name,
                "dataset_id": dataset_id,
                "workspace_id": workspace_id,
                "imported_at": imported_at,
                "business_description": str(_ai_row_value(row, "Description") or ""),
                "validation_status": "Imported",
                "is_active": True,
            },
        )
        imported["measures"] += 1

    relationships, error = _execute_powerbi_info_query(dataset_id, "RELATIONSHIPS")
    if error:
        errors["relationships"] = error
    for row in relationships:
        parent_table = str(_ai_row_value(row, "FromTable", "ParentTable", "From Table") or "").strip()
        parent_column = str(_ai_row_value(row, "FromColumn", "ParentColumn", "From Column") or "").strip()
        child_table = str(_ai_row_value(row, "ToTable", "ChildTable", "To Table") or "").strip()
        child_column = str(_ai_row_value(row, "ToColumn", "ChildColumn", "To Column") or "").strip()
        if not parent_table or not parent_column or not child_table or not child_column:
            continue
        AISemanticRelationship.objects.update_or_create(
            section=section,
            parent_table=parent_table,
            parent_column=parent_column,
            child_table=child_table,
            child_column=child_column,
            defaults={
                "relationship_type": str(_ai_row_value(row, "Cardinality", "Relationship Type", "CrossFilteringBehavior") or ""),
                "source_report": dataset_name,
                "dataset_id": dataset_id,
                "workspace_id": workspace_id,
                "imported_at": imported_at,
                "validation_status": "Imported",
                "is_active": True,
            },
        )
        imported["relationships"] += 1

    return JsonResponse(
        {
            "ok": True,
            "dataset_name": dataset_name,
            "dataset_id": dataset_id,
            "imported": imported,
            "errors": errors,
        }
    )


KPI_DICTIONARY_FIELDS = [
    "kpi_code", "kpi_name", "business_definition", "formula_description",
    "powerbi_measure_name", "unit", "target", "warning_threshold",
    "critical_threshold", "aggregation_rule", "default_time_grain", "owner",
    "validation_status", "is_active", "business_purpose", "business_category",
    "business_interpretation", "higher_is_better", "lower_is_better",
    "numerator_description", "denominator_description", "calculation_type",
    "null_handling_rule", "zero_denominator_behavior", "decimal_precision",
    "display_format", "powerbi_workspace_id", "powerbi_report_id",
    "powerbi_semantic_model_id", "powerbi_measure_table",
    "powerbi_measure_full_reference", "source_report_name", "source_page_name",
    "source_page_internal_name", "primary_visual_name",
    "primary_visual_internal_name", "default_comparison_type",
    "default_comparison_period", "default_ranking_direction", "default_top_n",
    "trend_supported", "comparison_supported", "ranking_supported",
    "root_cause_supported", "forecast_supported", "supported_dimensions",
    "default_drill_down_dimension", "required_filters", "optional_filters",
    "related_kpis", "diagnostic_kpis", "parent_kpi", "child_kpis",
    "default_answer_template", "business_explanation_template",
    "clarification_message", "ai_usage_instructions", "threshold_direction",
    "target_source", "target_measure_name", "threshold_evaluation_rule",
    "minimum_data_completeness", "minimum_equipment_count",
    "freshness_requirement", "data_quality_warning_message", "business_owner",
    "technical_owner", "approved_by", "approved_at", "version",
    "effective_from", "effective_to", "review_frequency", "last_reviewed_at",
    "review_notes",
]


KB_RESOURCE_TYPES = {
    "business-glossary": {
        "model": KnowledgeBusinessGlossary,
        "columns": ["term", "category", "related_kpi", "validation_status", "is_active"],
        "search_fields": ["term", "business_definition", "category", "related_kpi", "related_powerbi_measure"],
        "fields": ["term", "business_definition", "category", "related_kpi", "related_powerbi_measure", "related_table", "related_column", "example_usage", "owner", "validation_status", "is_active"],
    },
    "kpi-dictionary": {
        "model": KnowledgeKPIDictionary,
        "columns": ["kpi_code", "kpi_name", "powerbi_measure_name", "unit", "validation_status", "is_active"],
        "search_fields": [
            "kpi_code", "kpi_name", "business_definition", "business_purpose",
            "business_category", "powerbi_measure_name", "business_owner",
        ],
        "fields": KPI_DICTIONARY_FIELDS,
    },
    "mining-terminology": {
        "model": KnowledgeMiningTerminology,
        "columns": ["term", "category", "related_process", "validation_status", "is_active"],
        "search_fields": ["term", "definition", "category", "related_process"],
        "fields": ["term", "definition", "category", "related_process", "example", "owner", "validation_status", "is_active"],
    },
    "question-library": {
        "model": KnowledgeQuestion,
        "columns": ["question_text", "intent_type", "language", "difficulty_level", "validation_status", "is_active"],
        "search_fields": ["question_text", "intent_type", "expected_answer_style"],
        "fields": ["question_text", "intent_type", "expected_json_intent", "expected_dax", "expected_answer_style", "language", "difficulty_level", "owner", "validation_status", "is_active"],
    },
    "synonym-library": {
        "model": KnowledgeSynonym,
        "columns": [
            "canonical_term", "synonym", "normalized_value", "entity_type", "language",
            "confidence", "match_type", "synonym_source", "usage_count", "is_ambiguous",
            "validation_status", "is_active", "updated_at",
        ],
        "search_fields": ["canonical_term", "synonym", "normalized_value", "ambiguity_notes"],
        "fields": [
            "canonical_term", "synonym", "normalized_value", "entity_type", "language",
            "confidence", "match_type", "resolution_priority", "is_ambiguous",
            "ambiguity_notes", "synonym_source", "owner", "validation_status", "is_active",
        ],
    },
    "business-rules": {
        "model": KnowledgeBusinessRule,
        "columns": ["rule_name", "kpi", "condition", "validation_status", "is_active"],
        "search_fields": ["rule_name", "kpi", "condition", "rule_description", "default_behavior"],
        "fields": ["rule_name", "kpi", "condition", "rule_description", "default_behavior", "required_filters", "missing_filter_behavior", "owner", "validation_status", "is_active"],
    },
    "prompt-library": {
        "model": KnowledgePrompt,
        "columns": ["prompt_name", "prompt_type", "version", "validation_status", "is_active"],
        "search_fields": ["prompt_name", "prompt_type", "prompt_content", "version"],
        "fields": ["prompt_name", "prompt_type", "prompt_content", "version", "created_by", "owner", "validation_status", "is_active"],
    },
    "recommended-actions": {
        "model": KnowledgeRecommendedAction,
        "columns": ["kpi", "condition", "priority", "validation_status", "is_active"],
        "search_fields": ["kpi", "condition", "business_context", "recommended_action"],
        "fields": ["kpi", "condition", "business_context", "recommended_action", "priority", "owner", "validation_status", "is_active"],
    },
    "ai-logs": {
        "model": KnowledgeAILog,
        "readonly": True,
        "columns": ["created_at", "user_question", "detected_section", "status", "execution_time_ms", "error_message"],
        "search_fields": ["user_question", "detected_section", "generated_dax", "final_answer", "error_message"],
        "fields": ["user_question", "detected_section", "extracted_intent", "generated_dax", "powerbi_result", "final_answer", "status", "error_message", "execution_time_ms", "token_usage"],
    },
    "user-feedback": {
        "model": KnowledgeUserFeedback,
        "columns": ["created_at", "rating", "was_answer_useful", "feedback_comment"],
        "search_fields": ["feedback_comment", "corrected_answer"],
        "fields": ["rating", "feedback_comment", "was_answer_useful", "corrected_intent", "corrected_answer"],
    },
}


def _kb_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if hasattr(value, "quantize"):
        return str(value)
    if isinstance(value, User):
        return value.get_full_name() or value.username
    return value


def _kb_item_payload(item, fields: list[str] | None = None) -> dict:
    fields = fields or [field.name for field in item._meta.fields]
    payload = {"id": item.id}
    for field in fields:
        if hasattr(item, field):
            payload[field] = _kb_value(getattr(item, field))
    if hasattr(item, "section_id"):
        payload["section"] = item.section.code if item.section_id else ""
        payload["section_name"] = item.section.name if item.section_id else ""
    if hasattr(item, "user_id"):
        payload["user"] = item.user.username if item.user_id else ""
    return payload


def _kb_filter_queryset(queryset, model, request, search_fields):
    section_code = request.GET.get("section", "").strip()
    validation_status = request.GET.get("status", "").strip()
    active = request.GET.get("active", "").strip().lower()
    query = request.GET.get("q", "").strip()

    if section_code and hasattr(model, "section"):
        queryset = queryset.filter(section__code=section_code)
    if validation_status and hasattr(model, "validation_status"):
        queryset = queryset.filter(validation_status=validation_status)
    if active in {"1", "true", "yes"} and hasattr(model, "is_active"):
        queryset = queryset.filter(is_active=True)
    elif active in {"0", "false", "no"} and hasattr(model, "is_active"):
        queryset = queryset.filter(is_active=False)
    if query and search_fields:
        q_filter = models.Q()
        for field in search_fields:
            q_filter |= models.Q(**{f"{field}__icontains": query})
        queryset = queryset.filter(q_filter)
    if model is KnowledgeSynonym:
        mapping = {
            "entity_type": "entity_type",
            "language": "language",
            "source": "synonym_source",
            "match_type": "match_type",
            "owner": "owner",
        }
        for parameter, field in mapping.items():
            value = request.GET.get(parameter, "").strip()
            if value:
                queryset = queryset.filter(**{field: value})
        ambiguous = request.GET.get("ambiguous", "").strip().lower()
        if ambiguous in {"1", "true", "yes"}:
            queryset = queryset.filter(is_ambiguous=True)
        elif ambiguous in {"0", "false", "no"}:
            queryset = queryset.filter(is_ambiguous=False)
        quick = request.GET.get("quick", "").strip().lower()
        if quick == "unused":
            queryset = queryset.filter(usage_count=0)
        elif quick == "most-used":
            queryset = queryset.order_by("-usage_count", "-last_used_at")
        elif quick == "ai-generated":
            queryset = queryset.filter(synonym_source="AI Generated")
        elif quick == "validated":
            queryset = queryset.filter(validation_status="Validated")
        elif quick == "draft":
            queryset = queryset.filter(validation_status="Draft")
        elif quick == "to-review":
            queryset = queryset.filter(validation_status="To Review")
        elif quick == "ambiguous":
            queryset = queryset.filter(is_ambiguous=True)
        elif quick == "inactive":
            queryset = queryset.filter(is_active=False)
        min_usage = request.GET.get("min_usage", "").strip()
        max_usage = request.GET.get("max_usage", "").strip()
        min_confidence = request.GET.get("min_confidence", "").strip()
        max_confidence = request.GET.get("max_confidence", "").strip()
        if min_usage.isdigit():
            queryset = queryset.filter(usage_count__gte=int(min_usage))
        if max_usage.isdigit():
            queryset = queryset.filter(usage_count__lte=int(max_usage))
        if min_confidence:
            queryset = queryset.filter(confidence__gte=min_confidence)
        if max_confidence:
            queryset = queryset.filter(confidence__lte=max_confidence)
    return queryset


def _validation_error_message(exc: ValidationError) -> str:
    if hasattr(exc, "message_dict"):
        return "; ".join(
            f"{field}: {', '.join(messages)}"
            for field, messages in exc.message_dict.items()
        )
    return "; ".join(exc.messages)


def _kpi_dictionary_form_metadata(section_code: str = "") -> dict:
    filters = AIFilterMapping.objects.filter(is_active=True).select_related("section")
    kpis = KnowledgeKPIDictionary.objects.all().select_related("section")
    if section_code:
        filters = filters.filter(section__code=section_code)
        kpis = kpis.filter(section__code=section_code)
    return {
        "filters": [
            {
                "value": item.filter_code,
                "label": item.filter_label,
                "section": item.section.code,
            }
            for item in filters.order_by("filter_label")
        ],
        "kpis": [
            {
                "value": item.kpi_code,
                "label": item.kpi_name,
                "section": item.section.code,
            }
            for item in kpis.order_by("kpi_name")
        ],
    }


def _kb_apply_payload(item, payload: dict, fields: list[str], section=None):
    model_field_names = {field.name for field in item._meta.fields}
    if section is not None and "section" in model_field_names:
        item.section = section
    for field in fields:
        if field in {"created_at", "updated_at"}:
            continue
        try:
            model_field = item._meta.get_field(field)
        except Exception:
            model_field = None
        if isinstance(model_field, models.BooleanField):
            setattr(
                item, field,
                _ia_normalize_bool(payload.get(field), getattr(item, field, False)),
            )
            continue
        if isinstance(model_field, models.JSONField):
            raw_value = payload.get(field, getattr(item, field, []))
            if isinstance(raw_value, str):
                try:
                    raw_value = json.loads(raw_value) if raw_value.strip() else []
                except json.JSONDecodeError as exc:
                    raise ValueError(f"{model_field.verbose_name} must be valid JSON.") from exc
            setattr(item, field, raw_value)
            continue
        value = payload.get(field, getattr(item, field, ""))
        if isinstance(model_field, models.DecimalField):
            value = None if value in ("", None) else value
        if isinstance(model_field, models.IntegerField) and not isinstance(model_field, models.BooleanField):
            if value in ("", None) and getattr(model_field, "null", False):
                value = None
            else:
                value = int(value or 0)
        if isinstance(model_field, models.DateTimeField):
            value = None if value in ("", None) else parse_datetime(str(value))
            if payload.get(field) not in ("", None) and value is None:
                raise ValueError(f"{model_field.verbose_name} must be a valid datetime.")
        elif isinstance(model_field, models.DateField):
            value = None if value in ("", None) else parse_date(str(value))
            if payload.get(field) not in ("", None) and value is None:
                raise ValueError(f"{model_field.verbose_name} must be a valid date.")
        if field in {"priority", "rating", "execution_time_ms"} and model_field is None:
            value = _ia_int(payload, item, field, 0)
        setattr(item, field, value)
    return item


def _save_knowledge_synonym(item, user, *, previous=None):
    critical_changed = False
    if previous and previous.validation_status == "Validated":
        critical_changed = any(
            getattr(previous, field) != getattr(item, field)
            for field in KnowledgeSynonym.CRITICAL_FIELDS
        )
    if item.pk:
        item.updated_by = user if getattr(user, "is_authenticated", False) else None
    else:
        item.created_by = user if getattr(user, "is_authenticated", False) else None
        item.updated_by = item.created_by
    if critical_changed:
        item.validation_status = "To Review"
        item.validated_at = None
        item.validated_by = None
    elif item.validation_status == "Validated":
        if not previous or previous.validation_status != "Validated":
            item.validated_at = timezone.now()
            item.validated_by = user if getattr(user, "is_authenticated", False) else None
    elif item.validation_status != "Validated":
        item.validated_at = None
        item.validated_by = None
    item.save()
    return item


def _kb_powerbi_metadata_items(request) -> list[dict]:
    rows = []
    metadata_models = [
        ("Table", AISemanticTable, ["table_name", "display_name", "description", "business_description", "source_report", "dataset_id", "workspace_id", "validation_status", "is_active"]),
        ("Column", AISemanticColumn, ["table_name", "column_name", "display_name", "data_type", "business_description", "source_report", "dataset_id", "workspace_id", "validation_status", "is_active"]),
        ("Measure", AISemanticMeasure, ["measure_name", "display_name", "dax_name", "unit", "category", "business_description", "source_report", "dataset_id", "workspace_id", "validation_status", "is_active"]),
        ("Relationship", AISemanticRelationship, ["parent_table", "parent_column", "child_table", "child_column", "relationship_type", "business_description", "source_report", "dataset_id", "workspace_id", "validation_status", "is_active"]),
        ("Page", AIPowerBIPage, ["page_name", "report_name", "report_id", "page_display_name", "description", "is_default_page", "is_active"]),
    ]
    for item_type, model, fields in metadata_models:
        queryset = _kb_filter_queryset(model.objects.all(), model, request, ["display_name", "description", "business_description"] if hasattr(model, "business_description") else ["page_name", "report_name", "page_display_name"])
        for item in queryset.order_by("-updated_at")[:500]:
            payload = _kb_item_payload(item, fields)
            payload["item_type"] = item_type
            rows.append(payload)
    return rows


def _kb_overview_payload() -> dict:
    sections = get_active_section_objects()
    total_items = (
        AISemanticTable.objects.count()
        + AISemanticColumn.objects.count()
        + AISemanticMeasure.objects.count()
        + AISemanticRelationship.objects.count()
        + KnowledgeBusinessGlossary.objects.count()
        + KnowledgeKPIDictionary.objects.count()
        + KnowledgeMiningTerminology.objects.count()
        + KnowledgeQuestion.objects.count()
        + KnowledgeSynonym.objects.count()
        + KnowledgeBusinessRule.objects.count()
        + KnowledgePrompt.objects.count()
        + KnowledgeRecommendedAction.objects.count()
    )
    semantic_import_dates = [
        value
        for value in [
            AISemanticTable.objects.order_by("-imported_at").values_list("imported_at", flat=True).first(),
            AISemanticColumn.objects.order_by("-imported_at").values_list("imported_at", flat=True).first(),
            AISemanticMeasure.objects.order_by("-imported_at").values_list("imported_at", flat=True).first(),
        ]
        if value
    ]
    coverage = []
    coverage_scores = []
    for section in sections:
        metrics = AIMetricMapping.objects.filter(section=section, is_active=True).count()
        semantic = AISemanticMeasure.objects.filter(section=section, is_active=True).count() + AISemanticTable.objects.filter(section=section, is_active=True).count()
        synonyms = KnowledgeSynonym.objects.filter(section=section, is_active=True).count() + AISynonym.objects.filter(section=section, is_active=True).count()
        questions = KnowledgeQuestion.objects.filter(section=section, is_active=True).count() + AIQuestionExample.objects.filter(section=section, is_active=True).count()
        kpis = KnowledgeKPIDictionary.objects.filter(section=section, is_active=True).count() + AIKPITarget.objects.filter(section=section, is_active=True).count()
        rules = KnowledgeBusinessRule.objects.filter(section=section, is_active=True).count() + AIBusinessRule.objects.filter(section=section, is_active=True).count()
        score = round(sum(1 for value in [semantic, synonyms, questions, kpis, rules] if value > 0) / 5 * 100)
        coverage_scores.append(score)
        coverage.append(
            {
                "section": section.name,
                "code": section.code,
                "metadata": semantic,
                "synonyms": synonyms,
                "questions": questions,
                "kpis": kpis,
                "rules": rules,
                "score": score,
                "metrics": metrics,
            }
        )
    return {
        "total_items": total_items,
        "semantic_tables": AISemanticTable.objects.count(),
        "semantic_measures": AISemanticMeasure.objects.count(),
        "synonyms": KnowledgeSynonym.objects.count() + AISynonym.objects.count(),
        "question_examples": KnowledgeQuestion.objects.count() + AIQuestionExample.objects.count() + AIFewShotExample.objects.count(),
        "business_rules": KnowledgeBusinessRule.objects.count() + AIBusinessRule.objects.count(),
        "prompts": KnowledgePrompt.objects.count() + AIPromptTemplate.objects.count(),
        "recommended_actions": KnowledgeRecommendedAction.objects.count() + AIRecommendedAction.objects.count(),
        "last_imported_at": max(semantic_import_dates).isoformat() if semantic_import_dates else "",
        "coverage_score": round(sum(coverage_scores) / len(coverage_scores)) if coverage_scores else 0,
        "coverage": coverage,
    }


@require_http_methods(["GET"])
def knowledge_base_home(request):
    sections = [_ia_section_payload(section) for section in get_active_section_objects()]
    return render(
        request,
        "reports/knowledge_base.html",
        {
            "active_section": "knowledge-base",
            "sections": sections,
            "sidebar_stats": [
                {"label": "Knowledge", "value": "Base"},
                {"label": "Sections", "value": len(sections)},
            ],
        },
    )


@require_http_methods(["GET"])
def knowledge_base_overview_api(request):
    return JsonResponse({"ok": True, "overview": _kb_overview_payload()})


@require_http_methods(["GET"])
def synonym_analytics_api(request):
    if not is_platform_admin(request.user):
        return _json_error("Administrator access required.", status=403)
    queryset = _kb_filter_queryset(
        KnowledgeSynonym.objects.select_related("section"),
        KnowledgeSynonym,
        request,
        KB_RESOURCE_TYPES["synonym-library"]["search_fields"],
    )
    total = queryset.count()
    by = lambda field: list(
        queryset.values(field).annotate(count=models.Count("id")).order_by("-count", field)
    )
    serialize = lambda item: {
        "id": item.id,
        "synonym": item.synonym,
        "canonical_term": item.canonical_term,
        "usage_count": item.usage_count,
        "last_used_at": _kb_value(item.last_used_at),
        "created_at": _kb_value(item.created_at),
        "entity_type": item.entity_type,
        "source": item.synonym_source,
        "validation_status": item.validation_status,
        "confidence": float(item.confidence),
        "ambiguity_notes": item.ambiguity_notes,
    }
    return JsonResponse({"ok": True, "analytics": {
        "summary": {
            "total": total,
            "active": queryset.filter(is_active=True).count(),
            "validated": queryset.filter(validation_status="Validated").count(),
            "draft": queryset.filter(validation_status="Draft").count(),
            "ambiguous": queryset.filter(is_ambiguous=True).count(),
            "unused": queryset.filter(usage_count=0).count(),
            "ai_generated": queryset.filter(synonym_source="AI Generated").count(),
            "total_usages": queryset.aggregate(value=models.Sum("usage_count"))["value"] or 0,
        },
        "by_entity_type": by("entity_type"),
        "by_language": by("language"),
        "by_source": by("synonym_source"),
        "by_status": by("validation_status"),
        "most_used": [
            serialize(item) for item in queryset.order_by("-usage_count", "-last_used_at")[:20]
        ],
        "never_used": [
            serialize(item) for item in queryset.filter(usage_count=0).order_by("-created_at")[:20]
        ],
        "ambiguous": [
            serialize(item) for item in queryset.filter(is_ambiguous=True).order_by("synonym")[:100]
        ],
        "usage_trend": list(
            queryset.filter(last_used_at__isnull=False)
            .annotate(day=TruncDate("last_used_at"))
            .values("day")
            .annotate(count=models.Sum("usage_count"))
            .order_by("day")
        ),
    }})


@require_http_methods(["POST"])
def synonym_resolution_test_api(request):
    if not is_platform_admin(request.user):
        return _json_error("Administrator access required.", status=403)
    payload = _ia_payload(request)
    question = str(payload.get("question") or "").strip()
    if not question:
        return _json_error("Question is required.")
    mode = str(payload.get("mode") or "Production").title()
    if mode not in {"Production", "Debug"}:
        return _json_error("Execution mode must be Production or Debug.")
    result = resolve_synonyms(
        question,
        section_code=str(payload.get("section") or "").strip() or None,
        mode=mode,
        count_usage=_ia_normalize_bool(payload.get("count_usage"), False),
    )
    return JsonResponse({"ok": True, "result": result})


@require_http_methods(["GET", "PUT"])
def synonym_resolution_settings_api(request, section_code):
    if not is_platform_admin(request.user):
        return _json_error("Administrator access required.", status=403)
    section = get_object_or_404(AIConfigSection, code=section_code)
    if request.method == "PUT":
        payload = _ia_payload(request)
        try:
            threshold = int(payload.get("ambiguity_threshold"))
        except (TypeError, ValueError):
            return _json_error("Ambiguity threshold must be an integer between 1 and 100.")
        if not 1 <= threshold <= 100:
            return _json_error("Ambiguity threshold must be between 1 and 100.")
        section.synonym_ambiguity_threshold = threshold
        section.save(update_fields=["synonym_ambiguity_threshold", "updated_at"])
    return JsonResponse({"ok": True, "settings": {
        "section": section.code,
        "ambiguity_threshold": section.synonym_ambiguity_threshold,
    }})


@require_http_methods(["POST"])
def knowledge_resolution_api(request):
    if not is_platform_admin(request.user):
        return _json_error("Administrator access required.", status=403)
    payload = _ia_payload(request)
    question = str(payload.get("question") or "").strip()
    mode = str(payload.get("mode") or "Production").strip().title()
    if not question:
        return _json_error("User Question is required.")
    if mode not in {"Production", "Debug"}:
        return _json_error("Execution mode must be Production or Debug.")
    try:
        trace = resolve_knowledge_question(question, mode, request.user)
    except Exception as exc:
        return _json_error(str(exc), status=500)
    return JsonResponse({"ok": True, "trace": trace})


@require_http_methods(["GET"])
def knowledge_resolution_export(request, trace_id, file_type):
    if not is_platform_admin(request.user):
        return _json_error("Administrator access required.", status=403)
    trace = get_cached_trace(trace_id, request.user)
    if not trace:
        return _json_error("Resolution trace expired or was not found.", status=404)
    file_type = file_type.lower()
    if file_type == "json":
        content = json.dumps(trace, ensure_ascii=False, indent=2).encode("utf-8")
        content_type = "application/json"
    elif file_type in {"md", "markdown"}:
        content = trace_as_markdown(trace).encode("utf-8")
        content_type = "text/markdown; charset=utf-8"
        file_type = "md"
    elif file_type == "pdf":
        content = trace_as_basic_pdf(trace)
        content_type = "application/pdf"
    else:
        return _json_error("Export format must be JSON, Markdown or PDF.")
    response = HttpResponse(content, content_type=content_type)
    response["Content-Disposition"] = (
        f'attachment; filename="Mining360_Knowledge_Resolution_{trace_id[:8]}.{file_type}"'
    )
    return response


@require_http_methods(["GET"])
def knowledge_base_synonyms_export(request, file_type="xlsx"):
    queryset = _kb_filter_queryset(
        KnowledgeSynonym.objects.select_related("section").all(),
        KnowledgeSynonym,
        request,
        KB_RESOURCE_TYPES["synonym-library"]["search_fields"],
    ).order_by("section__name", "canonical_term", "synonym")
    export_rows = [{
        "section": item.section.code,
        "canonical_term": item.canonical_term,
        "synonym": item.synonym,
        "normalized_value": item.normalized_value,
        "entity_type": item.entity_type,
        "language": item.language,
        "confidence": float(item.confidence),
        "match_type": item.match_type,
        "synonym_source": item.synonym_source,
        "resolution_priority": item.resolution_priority,
        "is_ambiguous": item.is_ambiguous,
        "ambiguity_notes": item.ambiguity_notes,
        "usage_count": item.usage_count,
        "owner": item.owner,
        "validation_status": item.validation_status,
        "active": item.is_active,
        "updated_at": item.updated_at.isoformat(),
    } for item in queryset]
    file_type = str(file_type or "xlsx").lower()
    if file_type == "json":
        response = JsonResponse(export_rows, safe=False, json_dumps_params={"ensure_ascii": False, "indent": 2})
        response["Content-Disposition"] = 'attachment; filename="Mining360_Synonyms.json"'
        return response
    if file_type == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="Mining360_Synonyms.csv"'
        response.write("\ufeff")
        writer = csv.DictWriter(response, fieldnames=list(export_rows[0].keys()) if export_rows else [
            "section", "canonical_term", "synonym", "normalized_value", "entity_type",
            "language", "confidence", "match_type", "synonym_source",
        ])
        writer.writeheader()
        writer.writerows(export_rows)
        return response
    if file_type != "xlsx":
        return _json_error("Export format must be xlsx, csv or json.")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return _json_error("Excel export requires openpyxl.", status=500)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Synonyms"
    headers = [
        "Section", "Canonical Term", "Synonym", "Normalized Value", "Entity Type",
        "Language", "Confidence", "Match Type", "Synonym Source", "Resolution Priority",
        "Is Ambiguous", "Ambiguity Notes", "Usage Count", "Owner",
        "Validation Status", "Active", "Created At", "Updated At",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill("solid", fgColor="FFCD11")
    for item in queryset:
        worksheet.append([
            item.section.name,
            item.canonical_term,
            item.synonym,
            item.normalized_value,
            item.entity_type,
            item.language,
            float(item.confidence),
            item.match_type,
            item.synonym_source,
            item.resolution_priority,
            "Yes" if item.is_ambiguous else "No",
            item.ambiguity_notes,
            item.usage_count,
            item.owner,
            item.validation_status,
            "Yes" if item.is_active else "No",
            item.created_at.astimezone(timezone.get_current_timezone()).replace(tzinfo=None),
            item.updated_at.astimezone(timezone.get_current_timezone()).replace(tzinfo=None),
        ])
    widths = [20, 26, 30, 26, 18, 10, 12, 16, 18, 16, 14, 34, 12, 20, 16, 10, 20, 20]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    from io import BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Mining360_AI_Knowledge_Synonyms.xlsx"'
    return response


@require_http_methods(["GET"])
def knowledge_base_synonyms_template(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return _json_error("Excel template requires openpyxl.", status=500)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Synonyms Import"
    headers = [
        "Section Code", "Canonical Term", "Synonym", "Normalized Value", "Entity Type",
        "Language", "Confidence", "Synonym Source", "Match Type",
        "Resolution Priority", "Is Ambiguous", "Ambiguity Notes", "Owner",
        "Validation Status", "Active",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill("solid", fgColor="FFCD11")
    worksheet.append([
        "performance", "availability", "physical availability", "availability", "KPI",
        "en", 100, "Business", "Phrase", 80, "No", "",
        "Reliability Team", "To Review", "Yes",
    ])
    widths = [20, 28, 32, 28, 20, 12, 14, 20, 18, 18, 14, 36, 24, 20, 12]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:O5000"

    entity_values = ",".join(value for value, _ in KnowledgeSynonym.ENTITY_TYPES)
    status_values = ",".join(value for value, _ in KnowledgeSynonym.VALIDATION_STATUSES)
    source_values = ",".join(value for value, _ in KnowledgeSynonym.SOURCES)
    match_values = ",".join(value for value, _ in KnowledgeSynonym.MATCH_TYPES)
    validations = [
        (DataValidation(type="list", formula1=f'"{entity_values}"', allow_blank=False), "E2:E5000"),
        (DataValidation(type="list", formula1=f'"{source_values}"', allow_blank=False), "H2:H5000"),
        (DataValidation(type="list", formula1=f'"{match_values}"', allow_blank=False), "I2:I5000"),
        (DataValidation(type="list", formula1=f'"{status_values}"', allow_blank=False), "N2:N5000"),
        (DataValidation(type="list", formula1='"Yes,No"', allow_blank=False), "K2:K5000"),
        (DataValidation(type="list", formula1='"Yes,No"', allow_blank=False), "O2:O5000"),
        (DataValidation(type="decimal", operator="between", formula1="0", formula2="100"), "G2:G5000"),
        (DataValidation(type="whole", operator="between", formula1="1", formula2="100"), "J2:J5000"),
    ]
    for validation, cells in validations:
        worksheet.add_data_validation(validation)
        validation.add(cells)

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Field", "Requirement"])
    instructions.append(["Section Code", "Required. Must match an active IA Config section code, for example performance."])
    instructions.append(["Canonical Term", "Required. The normalized business term used by Mining360 AI."])
    instructions.append(["Synonym", "Required. A real language variation of the canonical term."])
    instructions.append(["Normalized Value", "Optional. Defaults to Canonical Term."])
    instructions.append(["Entity Type", "Required. Select a value from the dropdown."])
    instructions.append(["Language", "ISO language code such as en or fr. Default: en."])
    instructions.append(["Confidence", "Number between 0 and 100. Default: 100."])
    instructions.append(["Synonym Source", "Manual, Business, Imported, System Generated or AI Generated."])
    instructions.append(["Match Type", "Exact, Phrase, Contains, Abbreviation, Fuzzy or Semantic."])
    instructions.append(["Resolution Priority", "Integer between 1 and 100."])
    instructions.append(["Is Ambiguous", "Yes or No."])
    instructions.append(["Validation Status", "Draft, To Review, Validated, Rejected or Deprecated."])
    instructions.append(["Active", "Yes or No."])
    for cell in instructions[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFCD11")
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100

    from io import BytesIO
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Mining360_Synonyms_Import_Template.xlsx"'
    return response


def _import_boolean(value, default=True):
    if value is None or str(value).strip() == "":
        return default
    normalized = str(value).strip().lower()
    if normalized in {"yes", "y", "true", "1", "active", "oui"}:
        return True
    if normalized in {"no", "n", "false", "0", "inactive", "non"}:
        return False
    raise ValueError("Active must be Yes or No.")


@require_http_methods(["POST"])
def knowledge_base_synonyms_import(request):
    upload = request.FILES.get("file")
    if not upload:
        return _json_error("Select an Excel file to import.")
    extension = upload.name.lower().rsplit(".", 1)[-1]
    if extension not in {"xlsx", "csv"}:
        return _json_error("The synonym import file must use the .xlsx or .csv format.")
    try:
        if extension == "xlsx":
            from openpyxl import load_workbook
            workbook = load_workbook(upload, read_only=True, data_only=True)
            worksheet = workbook["Synonyms Import"] if "Synonyms Import" in workbook.sheetnames else workbook.active
            rows = worksheet.iter_rows(values_only=True)
        else:
            content = upload.read().decode("utf-8-sig")
            rows = iter(csv.reader(content.splitlines()))
        headers = next(rows, None)
    except Exception as exc:
        return _json_error(f"Unable to read the import file: {exc}")
    expected = [
        "Section Code", "Canonical Term", "Synonym", "Normalized Value", "Entity Type",
        "Language", "Confidence", "Synonym Source", "Match Type",
        "Resolution Priority", "Is Ambiguous", "Ambiguity Notes", "Owner",
        "Validation Status", "Active",
    ]
    normalized_headers = [str(value or "").strip() for value in (headers or [])]
    if normalized_headers[:len(expected)] != expected:
        return _json_error(
            "Invalid template. Download and use the current Mining360 synonym import template."
        )

    section_map = {
        item.code.lower(): item
        for item in AIConfigSection.objects.filter(is_active=True)
    }
    entity_types = {value for value, _ in KnowledgeSynonym.ENTITY_TYPES}
    statuses = {value for value, _ in KnowledgeSynonym.VALIDATION_STATUSES}
    sources = {value for value, _ in KnowledgeSynonym.SOURCES}
    match_types = {value for value, _ in KnowledgeSynonym.MATCH_TYPES}
    summary = {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0, "errors": []}
    imported_keys = set()
    for excel_row, values in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in values):
            continue
        summary["total_rows"] += 1
        try:
            padded = list(values) + [None] * (len(expected) - len(values))
            (
                section_code, canonical, synonym, normalized_value, entity_type,
                language, confidence, source, match_type, priority, ambiguous,
                ambiguity_notes, owner, status, active,
            ) = padded[:15]
            section = section_map.get(str(section_code or "").strip().lower())
            if not section:
                raise ValueError(f"Unknown or inactive section code: {section_code}")
            canonical = str(canonical or "").strip()
            synonym = str(synonym or "").strip()
            entity_type = str(entity_type or "").strip()
            if not canonical or not synonym:
                raise ValueError("Canonical Term and Synonym are required.")
            if entity_type not in entity_types:
                raise ValueError(f"Invalid Entity Type: {entity_type}")
            source = str(source or "Imported").strip()
            if source not in sources:
                raise ValueError(f"Invalid Synonym Source: {source}")
            match_type = str(match_type or default_match_type(synonym)).strip()
            if match_type not in match_types:
                raise ValueError(f"Invalid Match Type: {match_type}")
            status = str(status or "To Review").strip()
            if status not in statuses:
                raise ValueError(f"Invalid Validation Status: {status}")
            if source == "AI Generated":
                status = "Draft"
            try:
                confidence = float(confidence if confidence not in (None, "") else 100)
            except (TypeError, ValueError):
                raise ValueError("Confidence must be a number between 0 and 100.")
            if not 0 <= confidence <= 100:
                raise ValueError("Confidence must be between 0 and 100.")
            priority = int(priority or 50)
            if not 1 <= priority <= 100:
                raise ValueError("Resolution Priority must be between 1 and 100.")
            normalized_key = normalize_synonym_key(synonym)
            duplicate_key = (
                section.id,
                entity_type.casefold(),
                str(language or "en").strip().lower(),
                normalized_key,
            )
            if duplicate_key in imported_keys:
                raise ValueError("Duplicate synonym in the import file.")
            imported_keys.add(duplicate_key)
            item = KnowledgeSynonym.objects.filter(
                section=section,
                entity_type=entity_type,
                language=str(language or "en").strip().lower(),
                normalized_synonym_key=normalized_key,
            ).first()
            defaults = {
                "canonical_term": canonical,
                "synonym": synonym,
                "normalized_value": str(normalized_value or canonical).strip(),
                "language": str(language or "en").strip().lower(),
                "confidence": confidence,
                "synonym_source": source,
                "match_type": match_type,
                "resolution_priority": priority,
                "is_ambiguous": _import_boolean(ambiguous, False),
                "ambiguity_notes": str(ambiguity_notes or "").strip(),
                "owner": str(owner or "").strip(),
                "validation_status": status,
                "is_active": _import_boolean(active),
            }
            if item:
                raise ValueError(
                    f"Duplicate synonym already exists (record ID {item.id})."
                )
            else:
                imported = KnowledgeSynonym(
                    section=section,
                    entity_type=entity_type,
                    created_by=request.user if request.user.is_authenticated else None,
                    updated_by=request.user if request.user.is_authenticated else None,
                    **defaults,
                )
                imported.full_clean()
                _save_knowledge_synonym(imported, request.user)
                summary["created"] += 1
        except Exception as exc:
            if len(summary["errors"]) < 20:
                summary["errors"].append({"row": excel_row, "message": str(exc)})
    summary["error_count"] = summary["total_rows"] - summary["created"] - summary["updated"] - summary["skipped"]
    return JsonResponse({"ok": True, "summary": summary})


@require_http_methods(["GET", "POST"])
def knowledge_base_collection_api(request, resource_type):
    if resource_type == "powerbi-metadata":
        return JsonResponse({"ok": True, "items": _kb_powerbi_metadata_items(request)})
    if resource_type not in KB_RESOURCE_TYPES:
        return _json_error("Unsupported Knowledge Base resource type.", status=404)
    config = KB_RESOURCE_TYPES[resource_type]
    model = config["model"]
    if request.method == "GET":
        queryset = _kb_filter_queryset(model.objects.all(), model, request, config.get("search_fields", []))
        order_field = "-created_at" if resource_type in {"ai-logs", "user-feedback"} else "-updated_at"
        if resource_type == "synonym-library" and request.GET.get("quick") == "most-used":
            queryset = queryset.order_by("-usage_count", "-last_used_at", "synonym")
        else:
            queryset = queryset.order_by(order_field)
        limit = 2000 if resource_type == "synonym-library" else 500
        items = [_kb_item_payload(item) for item in queryset[:limit]]
        response = {"ok": True, "items": items}
        if resource_type == "kpi-dictionary":
            response["form_metadata"] = _kpi_dictionary_form_metadata(
                request.GET.get("section", "").strip()
            )
        return JsonResponse(response)

    if config.get("readonly"):
        return _json_error("This Knowledge Base resource is read-only.", status=405)
    payload = _ia_payload(request)
    section_value = payload.get("section") or payload.get("section_code") or request.GET.get("section")
    if not section_value and payload.get("section_id"):
        section = get_object_or_404(AIConfigSection, id=payload["section_id"])
    else:
        section = _ia_get_section_or_404(str(section_value or "performance"))
    try:
        item = _kb_apply_payload(model(), payload, config["fields"], section=section)
        if isinstance(item, (KnowledgeKPIDictionary, KnowledgeSynonym)):
            item.full_clean()
        if isinstance(item, KnowledgeSynonym):
            _save_knowledge_synonym(item, request.user)
        else:
            item.save()
    except ValidationError as exc:
        return _json_error(_validation_error_message(exc))
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "item": _kb_item_payload(item)}, status=201)


@require_http_methods(["PUT", "DELETE"])
def knowledge_base_item_api(request, resource_type, item_id):
    if resource_type not in KB_RESOURCE_TYPES:
        return _json_error("Unsupported Knowledge Base resource type.", status=404)
    config = KB_RESOURCE_TYPES[resource_type]
    if config.get("readonly"):
        return _json_error("This Knowledge Base resource is read-only.", status=405)
    model = config["model"]
    item = get_object_or_404(model, id=item_id)
    if request.method == "DELETE":
        hard_delete = (
            resource_type in {"business-glossary", "kpi-dictionary"}
            and request.GET.get("hard", "").strip().lower() in {"1", "true", "yes"}
        )
        if hard_delete:
            if not is_platform_admin(request.user):
                return _json_error(
                    "Administrator access is required to permanently delete this item.",
                    status=403,
                )
            try:
                item.delete()
            except (ProtectedError, RestrictedError):
                return _json_error(
                    "This item cannot be deleted because it is referenced by another configuration.",
                    status=409,
                )
            return JsonResponse({"ok": True, "deleted": True})
        if hasattr(item, "is_active"):
            previous = KnowledgeSynonym.objects.get(pk=item.pk) if isinstance(item, KnowledgeSynonym) else None
            item.is_active = False
            if isinstance(item, KnowledgeSynonym):
                _save_knowledge_synonym(item, request.user, previous=previous)
            else:
                item.save(update_fields=["is_active", "updated_at"] if hasattr(item, "updated_at") else ["is_active"])
            return JsonResponse({"ok": True, "deactivated": True})
        item.delete()
        return JsonResponse({"ok": True, "deleted": True})
    payload = _ia_payload(request)
    if any(field.name == "section" for field in item._meta.fields):
        section_value = payload.get("section") or payload.get("section_code")
        if not section_value and payload.get("section_id"):
            section = get_object_or_404(AIConfigSection, id=payload["section_id"])
        else:
            section = _ia_get_section_or_404(str(section_value or item.section.code))
    else:
        section = None
    previous = None
    if isinstance(item, KnowledgeSynonym):
        previous = KnowledgeSynonym.objects.get(pk=item.pk)
    try:
        item = _kb_apply_payload(item, payload, config["fields"], section=section)
        if isinstance(item, (KnowledgeKPIDictionary, KnowledgeSynonym)):
            item.full_clean()
        if isinstance(item, KnowledgeSynonym):
            _save_knowledge_synonym(item, request.user, previous=previous)
        else:
            item.save()
    except ValidationError as exc:
        return _json_error(_validation_error_message(exc))
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "item": _kb_item_payload(item)})


@require_http_methods(["POST"])
def knowledge_base_kpi_duplicate_api(request, item_id):
    source = get_object_or_404(KnowledgeKPIDictionary, id=item_id)
    base_code = f"{source.kpi_code}_copy"
    code = base_code
    suffix = 2
    while KnowledgeKPIDictionary.objects.filter(section=source.section, kpi_code=code).exists():
        code = f"{base_code}_{suffix}"
        suffix += 1
    values = {}
    for field in source._meta.fields:
        if field.primary_key or field.name in {"created_at", "updated_at"}:
            continue
        values[field.attname] = getattr(source, field.attname)
    values.update({
        "kpi_code": code,
        "kpi_name": f"{source.kpi_name} Copy",
        "validation_status": "Draft",
        "approved_by": "",
        "approved_at": None,
    })
    item = KnowledgeKPIDictionary(**values)
    try:
        item.full_clean()
        item.save()
    except ValidationError as exc:
        return _json_error(_validation_error_message(exc))
    return JsonResponse(
        {"ok": True, "item": _kb_item_payload(item, KPI_DICTIONARY_FIELDS)},
        status=201,
    )


def _kpi_threshold_interpretation(item: KnowledgeKPIDictionary, value) -> str:
    if value in (None, ""):
        return "No Power BI value is available for threshold evaluation."
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return "The Power BI result is not numeric; thresholds were not evaluated."
    target = float(item.target) if item.target is not None else None
    warning = float(item.warning_threshold) if item.warning_threshold is not None else None
    critical = float(item.critical_threshold) if item.critical_threshold is not None else None
    lower_is_better = item.lower_is_better or item.threshold_direction == "Lower Is Better"
    if lower_is_better:
        if critical is not None and numeric_value >= critical:
            return "Critical"
        if warning is not None and numeric_value >= warning:
            return "Warning"
        if target is not None and numeric_value <= target:
            return "On Target"
    else:
        if critical is not None and numeric_value <= critical:
            return "Critical"
        if warning is not None and numeric_value <= warning:
            return "Warning"
        if target is not None and numeric_value >= target:
            return "On Target"
    return "Within configured range"


def _format_kpi_display_value(item: KnowledgeKPIDictionary, value) -> str:
    if value in (None, ""):
        return "<value>"
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return str(value)
    precision = max(0, int(item.decimal_precision or 0))
    is_percentage = item.unit.strip() == "%" or "%" in (item.display_format or "")
    if is_percentage:
        numeric_value *= 100
    return f"{numeric_value:,.{precision}f}"


@require_http_methods(["POST"])
def knowledge_base_kpi_test_api(request, item_id):
    item = get_object_or_404(
        KnowledgeKPIDictionary.objects.select_related("section"), id=item_id
    )
    payload = _request_payload(request)
    warnings = []
    validation_errors = []
    try:
        item.full_clean()
    except ValidationError as exc:
        validation_errors = exc.messages

    required_configuration = [
        "kpi_code", "kpi_name", "business_definition", "formula_description",
        "powerbi_measure_name", "unit", "aggregation_rule", "default_time_grain",
        "business_purpose", "calculation_type",
    ]
    configured = [
        field for field in required_configuration
        if getattr(item, field, None) not in (None, "", [])
    ]
    completeness = round(len(configured) / len(required_configuration) * 100)
    filters = list(
        AIFilterMapping.objects.filter(section=item.section, is_active=True)
        .values("filter_code", "filter_label", "powerbi_table_name", "powerbi_column_name")
    )
    dax_template = (
        AIDaxTemplate.objects.filter(section=item.section, is_active=True)
        .order_by("template_code").first()
    )
    measure_reference = (
        item.powerbi_measure_full_reference
        or (f"[{item.powerbi_measure_name.strip('[]')}]" if item.powerbi_measure_name else "")
    )
    generated_dax = (
        f'EVALUATE\nROW("{item.kpi_name.replace(chr(34), chr(34) * 2)}", '
        f"{measure_reference})"
        if measure_reference else ""
    )
    generated_intent = {
        "section": item.section.code,
        "intent_type": "single_kpi",
        "metric": item.kpi_code,
        "filters": {code: None for code in (item.required_filters or [])},
    }
    powerbi_result = None
    powerbi_error = ""
    if _ia_normalize_bool(payload.get("execute_powerbi"), False):
        if not item.powerbi_semantic_model_id:
            powerbi_error = "Power BI Semantic Model ID is not configured."
        elif not generated_dax:
            powerbi_error = "Power BI Measure Name is not configured."
        else:
            try:
                if get_flow_url():
                    flow_result = execute_dax_via_flow({
                        "datasetId": item.powerbi_semantic_model_id,
                        "datasetName": item.source_report_name or item.section.name,
                        "query": generated_dax,
                        "question": f"Test KPI {item.kpi_name}",
                        "metric": item.kpi_code,
                        "measure": item.powerbi_measure_name,
                        "filters": {},
                        "period": {},
                        "rlsRole": "",
                        "roles": [],
                    })
                    powerbi_result = _extract_flow_rows(flow_result)
                else:
                    powerbi_result = execute_dataset_dax(
                        item.powerbi_semantic_model_id, generated_dax
                    )
            except Exception as exc:
                powerbi_error = str(exc)
        if powerbi_error:
            warnings.append(powerbi_error)

    result_value = None
    if isinstance(powerbi_result, list) and powerbi_result:
        first_row = powerbi_result[0]
        if isinstance(first_row, dict) and first_row:
            result_value = next(iter(first_row.values()))
    interpretation = _kpi_threshold_interpretation(item, result_value)
    formatted_value = _format_kpi_display_value(item, result_value)
    formatted_target = _format_kpi_display_value(item, item.target)
    answer_template = item.default_answer_template or (
        f"The {item.kpi_name} value is {{value}} {item.unit}."
    )
    example_answer = (
        answer_template
        .replace("{value}", formatted_value)
        .replace("{target}", formatted_target if item.target is not None else "<target>")
        .replace("{minesite}", "<minesite>")
        .replace("{period}", "<period>")
        .replace("{status}", interpretation)
    )
    return JsonResponse({
        "ok": True,
        "test": {
            "configuration_completeness": {
                "score": completeness,
                "configured": configured,
                "missing": [
                    field for field in required_configuration if field not in configured
                ],
            },
            "powerbi_measure_mapping": {
                "measure_name": item.powerbi_measure_name,
                "measure_table": item.powerbi_measure_table,
                "full_reference": measure_reference,
                "semantic_model_id": item.powerbi_semantic_model_id,
            },
            "supported_filters": filters,
            "required_filters": item.required_filters,
            "generated_json_intent": generated_intent,
            "selected_dax_template": {
                "code": dax_template.template_code if dax_template else "",
                "name": dax_template.template_name if dax_template else "",
            },
            "generated_dax": generated_dax,
            "powerbi_result": powerbi_result,
            "formatted_value": (
                f"{formatted_value}{item.unit}"
                if result_value is not None and item.unit else formatted_value
            ),
            "threshold_interpretation": interpretation,
            "example_ai_answer": example_answer,
            "validation_status": "Valid" if not validation_errors else "Invalid",
            "validation_warnings": validation_errors + warnings,
        },
    })


@require_http_methods(["POST"])
def knowledge_base_generate_api(request):
    payload = _ia_payload(request)
    section_code = str(payload.get("section") or "").strip()
    sections = AIConfigSection.objects.filter(is_active=True)
    if section_code:
        sections = sections.filter(code=section_code)
    created = {
        "kpis": 0,
        "glossary": 0,
        "terminology": 0,
        "questions": 0,
        "synonyms": 0,
        "rules": 0,
        "prompts": 0,
        "actions": 0,
        "metadata_marked": 0,
    }
    for section in sections:
        for metric in AIMetricMapping.objects.filter(section=section, is_active=True):
            item, was_created = KnowledgeKPIDictionary.objects.update_or_create(
                section=section,
                kpi_code=metric.metric_code,
                defaults={
                    "kpi_name": metric.metric_label,
                    "business_definition": metric.description or f"Business KPI for {metric.metric_label}.",
                    "powerbi_measure_name": metric.powerbi_measure_name,
                    "unit": "",
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["kpis"] += int(was_created)
            _, glossary_created = KnowledgeBusinessGlossary.objects.update_or_create(
                section=section,
                term=metric.metric_label,
                defaults={
                    "business_definition": metric.description or f"Mining 360 KPI mapped to {metric.powerbi_measure_name}.",
                    "category": "KPI",
                    "related_kpi": metric.metric_code,
                    "related_powerbi_measure": metric.powerbi_measure_name,
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["glossary"] += int(glossary_created)
        for synonym in AISynonym.objects.filter(section=section, is_active=True):
            _, was_created = KnowledgeSynonym.objects.update_or_create(
                section=section,
                canonical_term=synonym.canonical_value,
                synonym=synonym.synonym_value,
                entity_type="KPI" if synonym.entity_type in {"metric", "measure"} else "Business Term",
                defaults={
                    "language": synonym.language,
                    "confidence": 1,
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["synonyms"] += int(was_created)
        for question in AIQuestionExample.objects.filter(section=section, is_active=True):
            _, was_created = KnowledgeQuestion.objects.update_or_create(
                section=section,
                question_text=question.question_text,
                defaults={
                    "intent_type": "Single KPI",
                    "expected_json_intent": question.expected_json_intent,
                    "language": question.language,
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["questions"] += int(was_created)
        for rule in AIBusinessRule.objects.filter(section=section, is_active=True):
            _, was_created = KnowledgeBusinessRule.objects.update_or_create(
                section=section,
                rule_name=rule.rule_name,
                defaults={
                    "kpi": rule.metric_code,
                    "condition": rule.condition,
                    "rule_description": rule.action,
                    "default_behavior": rule.default_value,
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["rules"] += int(was_created)
        for prompt in AIPromptTemplate.objects.filter(section=section, is_active=True):
            _, was_created = KnowledgePrompt.objects.update_or_create(
                section=section,
                prompt_name=prompt.template_name,
                prompt_type=prompt.get_prompt_type_display(),
                defaults={
                    "prompt_content": prompt.prompt_template,
                    "version": "1.0",
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["prompts"] += int(was_created)
        for action in AIRecommendedAction.objects.filter(section=section, is_active=True):
            _, was_created = KnowledgeRecommendedAction.objects.update_or_create(
                section=section,
                kpi=action.metric_code,
                condition=action.condition,
                defaults={
                    "recommended_action": action.recommendations,
                    "priority": action.priority,
                    "validation_status": "To Review",
                    "is_active": True,
                },
            )
            created["actions"] += int(was_created)
        metadata_updates = (
            AISemanticTable.objects.filter(section=section, validation_status="Imported").update(validation_status="To Review")
            + AISemanticColumn.objects.filter(section=section, validation_status="Imported").update(validation_status="To Review")
            + AISemanticMeasure.objects.filter(section=section, validation_status="Imported").update(validation_status="To Review")
            + AISemanticRelationship.objects.filter(section=section, validation_status="Imported").update(validation_status="To Review")
        )
        created["metadata_marked"] += metadata_updates
    return JsonResponse({"ok": True, "generated": created, "overview": _kb_overview_payload()})


@require_http_methods(["POST"])
def knowledge_base_coverage_test_api(request):
    payload = _ia_payload(request)
    mode = str(payload.get("mode") or "Production").strip().title()
    if mode not in {"Production", "Debug"}:
        return _json_error("Execution mode must be Production or Debug.")
    if mode == "Debug" and not is_platform_admin(request.user):
        return _json_error(
            "Debug mode is restricted to administrators.",
            status=403,
        )
    allowed_statuses = (
        ["Validated"]
        if mode == "Production"
        else ["Validated", "Draft", "To Review"]
    )
    section_code = str(payload.get("section") or "").strip().lower() or "performance"
    kpi = str(payload.get("kpi") or "").strip().lower()
    question = str(payload.get("question") or "").strip()
    intent = {}
    if question:
        try:
            intent = extract_intent(question, section_code)
            kpi = kpi or str(intent.get("metric") or "").strip().lower()
        except Exception:
            intent = {}
    section = get_section_by_code(section_code) or _ia_get_section_or_404(section_code)
    kpi_candidate = (
        KnowledgeKPIDictionary.objects.filter(
            section__code=section_code,
            kpi_code=kpi,
        )
        .order_by(
            models.Case(
                models.When(validation_status="Validated", then=0),
                models.When(validation_status="To Review", then=1),
                models.When(validation_status="Draft", then=2),
                default=3,
                output_field=models.IntegerField(),
            ),
            "id",
        )
        .first()
        if kpi else None
    )
    matched_kpi = (
        kpi_candidate
        if (
            kpi_candidate
            and kpi_candidate.is_active
            and kpi_candidate.validation_status in allowed_statuses
        )
        else None
    )
    rejection_reason = ""
    if not kpi:
        rejection_reason = "No KPI code was extracted or provided."
    elif not kpi_candidate:
        rejection_reason = (
            f"No KPI Dictionary entry matches section '{section_code}' "
            f"and KPI code '{kpi}'."
        )
    elif not kpi_candidate.is_active:
        rejection_reason = f"KPI Dictionary entry {kpi_candidate.id} is inactive."
    elif kpi_candidate.validation_status not in allowed_statuses:
        rejection_reason = (
            f"KPI Dictionary entry {kpi_candidate.id} has validation status "
            f"'{kpi_candidate.validation_status}'; allowed statuses in {mode} "
            f"mode are {', '.join(allowed_statuses)}."
        )

    measure_mapped = bool(
        matched_kpi
        and str(matched_kpi.powerbi_measure_name or "").strip()
        and str(matched_kpi.powerbi_semantic_model_id or "").strip()
    )

    evidence = []

    def add_evidence(repository, item, status, used, item_id=None, reason=""):
        evidence.append({
            "repository": repository,
            "item": item or "Not configured",
            "status": status or "Not Found",
            "used": bool(used),
            "mode": mode,
            "item_id": item_id,
            "reason": reason,
        })

    add_evidence(
        "KPI Dictionary",
        kpi_candidate.kpi_name if kpi_candidate else kpi,
        kpi_candidate.validation_status if kpi_candidate else "Not Found",
        bool(matched_kpi),
        kpi_candidate.id if kpi_candidate else None,
        rejection_reason,
    )
    add_evidence(
        "Power BI Mapping",
        (
            matched_kpi.powerbi_measure_full_reference
            or matched_kpi.powerbi_measure_name
            if matched_kpi else kpi
        ),
        matched_kpi.validation_status if matched_kpi else (
            kpi_candidate.validation_status if kpi_candidate else "Not Found"
        ),
        measure_mapped,
        matched_kpi.id if matched_kpi else None,
        "" if measure_mapped else (
            rejection_reason
            or "Power BI Measure Name and Semantic Model ID are required."
        ),
    )

    repository_queries = [
        (
            "Synonym Library",
            KnowledgeSynonym.objects.filter(
                section=section,
                canonical_term__iexact=kpi,
            ),
            lambda item: f"{item.canonical_term} = {item.synonym}",
        ),
        (
            "Business Rules",
            KnowledgeBusinessRule.objects.filter(section=section).filter(
                models.Q(kpi__iexact=kpi) | models.Q(kpi="")
            ),
            lambda item: item.rule_name,
        ),
        (
            "Prompt Library",
            KnowledgePrompt.objects.filter(section=section),
            lambda item: item.prompt_name,
        ),
        (
            "Question Library",
            KnowledgeQuestion.objects.filter(section=section),
            lambda item: item.question_text[:100],
        ),
        (
            "Recommended Actions",
            KnowledgeRecommendedAction.objects.filter(
                section=section,
                kpi__iexact=kpi,
            ),
            lambda item: item.recommended_action[:100],
        ),
    ]
    repository_usage = {}
    for repository, queryset, label_getter in repository_queries:
        candidates = list(queryset.order_by("-updated_at")[:100])
        usable = [
            item for item in candidates
            if item.is_active and item.validation_status in allowed_statuses
        ]
        repository_usage[repository] = bool(usable)
        if usable:
            for item in usable:
                add_evidence(
                    repository,
                    label_getter(item),
                    item.validation_status,
                    True,
                    item.id,
                )
        elif candidates:
            candidate = candidates[0]
            reason = (
                "Item is inactive."
                if not candidate.is_active
                else (
                    f"Status '{candidate.validation_status}' is not allowed "
                    f"in {mode} mode."
                )
            )
            add_evidence(
                repository,
                label_getter(candidate),
                candidate.validation_status,
                False,
                candidate.id,
                reason,
            )
        else:
            add_evidence(
                repository,
                kpi,
                "Not Found",
                False,
                reason=f"No matching {repository} item was found.",
            )

    interaction_candidates = list(
        KPIPageMapping.objects.filter(
            section=section,
            metric_code__iexact=kpi,
        )
        .select_related("report", "page")
        .order_by("priority", "id")
    )
    usable_interactions = [
        item for item in interaction_candidates
        if (
            item.is_active
            and item.report.is_active
            and item.page.is_active
            and item.report.validation_status in allowed_statuses
            and item.page.validation_status in allowed_statuses
        )
    ]
    repository_usage["Power BI Interaction"] = bool(usable_interactions)
    if usable_interactions:
        for item in usable_interactions:
            add_evidence(
                "Power BI Interaction",
                f"{item.report.display_name} / {item.page.page_display_name}",
                item.page.validation_status,
                True,
                item.id,
            )
    elif interaction_candidates:
        item = interaction_candidates[0]
        add_evidence(
            "Power BI Interaction",
            f"{item.report.display_name} / {item.page.page_display_name}",
            item.page.validation_status,
            False,
            item.id,
            "Mapping, report and page must be active with an allowed validation status.",
        )
    else:
        add_evidence(
            "Power BI Interaction",
            kpi,
            "Not Found",
            False,
            reason="No KPI-to-page mapping was found.",
        )

    checks = {
        "kpi_defined": bool(matched_kpi),
        "measure_mapped": measure_mapped,
        "filters_mapped": AIFilterMapping.objects.filter(section=section, is_active=True).exists(),
        "synonyms_available": repository_usage["Synonym Library"],
        "dax_template": AIDaxTemplate.objects.filter(section=section, is_active=True).exists(),
        "business_rules": repository_usage["Business Rules"],
        "prompt_library": repository_usage["Prompt Library"],
        "question_library": repository_usage["Question Library"],
        "powerbi_interaction": repository_usage["Power BI Interaction"],
        "recommended_actions": repository_usage["Recommended Actions"],
    }
    score = round(sum(1 for value in checks.values() if value) / len(checks) * 100)
    debug_items_used = [
        item for item in evidence
        if item["used"] and item["status"] in {"Draft", "To Review"}
    ]
    warnings = []
    if mode == "Debug" and debug_items_used:
        warnings.append(
            "This test uses non-validated knowledge. Results may differ from Production."
        )
    return JsonResponse({
        "ok": True,
        "mode": mode,
        "test_type": "Knowledge Readiness",
        "business_result_executed": False,
        "message": (
            "This score measures Knowledge Base readiness. "
            "It is not the KPI value and does not execute Power BI."
        ),
        "intent": intent,
        "checks": checks,
        "repositories": evidence,
        "coverage_score": score,
        "warnings": warnings,
        "debug": {
            "searched_section_code": section_code,
            "searched_kpi_code": kpi,
            "allowed_validation_statuses": allowed_statuses,
            "candidate_kpi_id": kpi_candidate.id if kpi_candidate else None,
            "matched_kpi_id": matched_kpi.id if matched_kpi else None,
            "rejection_reason": rejection_reason,
            "candidate_active": (
                kpi_candidate.is_active if kpi_candidate else None
            ),
            "candidate_validation_status": (
                kpi_candidate.validation_status if kpi_candidate else None
            ),
            "measure_name_configured": bool(
                matched_kpi
                and str(matched_kpi.powerbi_measure_name or "").strip()
            ),
            "semantic_model_id_configured": bool(
                matched_kpi
                and str(matched_kpi.powerbi_semantic_model_id or "").strip()
            ),
        },
    })


def _mask_secret(value: str) -> str:
    return "********" if value else ""


def _system_db_payload(item: SystemDatabaseConfig, reveal: bool = False) -> dict:
    return {
        "id": item.id,
        "name": item.name,
        "engine": item.engine,
        "purpose": item.purpose,
        "host": item.host,
        "port": item.port or "",
        "database_name": item.database_name,
        "schema_name": item.schema_name,
        "username": item.username,
        "password": item.password if reveal else _mask_secret(item.password),
        "driver": item.driver,
        "connection_options": item.connection_options,
        "is_default": item.is_default,
        "is_active": item.is_active,
        "last_verified_at": item.last_verified_at.isoformat() if item.last_verified_at else "",
        "last_status": item.last_status,
        "last_message": item.last_message,
        "updated_at": item.updated_at.isoformat() if item.updated_at else "",
    }


def _system_table_payload(item: SystemManagedTable) -> dict:
    return {
        "id": item.id,
        "database_config": item.database_config_id,
        "database_config_name": item.database_config.name if item.database_config_id else "",
        "schema_name": item.schema_name,
        "table_name": item.table_name,
        "category": item.category,
        "model_name": item.model_name,
        "description": item.description,
        "row_count": item.row_count,
        "last_synced_at": item.last_synced_at.isoformat() if item.last_synced_at else "",
        "is_active": item.is_active,
    }


def _ensure_default_system_config() -> SystemDatabaseConfig:
    ensure_portable_configuration()
    config = SystemDatabaseConfig.objects.filter(is_default=True).first()
    if config:
        return config
    integration = SystemIntegrationConfig.objects.filter(
        integration_type="Database", is_default=True, is_active=True,
    ).first()
    values = integration.settings_json if integration else {}
    config = SystemDatabaseConfig.objects.create(
        name="Mining360 Database",
        engine=str(values.get("engine") or "SQL Server"),
        purpose="Primary Mining360 configuration database",
        host=str(values.get("host") or "not-configured"),
        port=values.get("port") or None,
        database_name=str(values.get("database") or ""),
        schema_name=str(values.get("schema") or "dbo"),
        username=str(values.get("username") or ""),
        driver=str(values.get("driver") or ""),
        is_default=True,
        is_active=bool(values.get("host")),
    )
    return config


def _system_config_table_category(model_name: str, table_name: str) -> str:
    if table_name.startswith("bp_"):
        return "Business Performance"
    if table_name.startswith("ai_"):
        return "IA Config"
    if table_name.startswith("kb_"):
        return "Knowledge Base"
    if table_name in {"BrowserList", "reports_databrowsercolumn", "reports_databrowsersynclog"}:
        return "Data Browser"
    if table_name in {"LiveSourceConfig", "LiveSourceCustomView"}:
        return "Sources"
    if table_name in {"PowerBIReportAlias"}:
        return "Power BI"
    if table_name in {"ResourceFile"}:
        return "Resources"
    if "log" in table_name.lower() or "run" in table_name.lower():
        return "Logs"
    return "Django Config"


def _refresh_system_table_registry() -> int:
    from .mining360_repository import CONFIG_MODEL_NAMES

    config = _ensure_default_system_config()
    updated = 0
    for model_name in CONFIG_MODEL_NAMES:
        try:
            model = apps.get_model("reports", model_name)
        except Exception:
            continue
        table_name = model._meta.db_table
        row_count = model.objects.count()
        SystemManagedTable.objects.update_or_create(
            schema_name="dbo",
            table_name=table_name,
            defaults={
                "database_config": config,
                "category": _system_config_table_category(model_name, table_name),
                "model_name": model_name,
                "description": f"Django model mirror table for {model_name}.",
                "row_count": row_count,
                "last_synced_at": timezone.now(),
                "is_active": True,
            },
        )
        updated += 1
    for table_name, category in [
        ("PowerBIReportAlias", "Power BI"),
        ("ResourceFile", "Resources"),
        ("LiveSourceConfig", "Sources"),
        ("LiveSourceCustomView", "Sources"),
    ]:
        SystemManagedTable.objects.update_or_create(
            schema_name="dbo",
            table_name=table_name,
            defaults={
                "database_config": config,
                "category": category,
                "description": f"SQL Server managed table for {category}.",
                "last_synced_at": timezone.now(),
                "is_active": True,
            },
        )
        updated += 1
    return updated


@login_required
@require_http_methods(["GET"])
def system_config_home(request):
    if not _user_is_platform_admin(request.user):
        return redirect("dashboard")
    ensure_portable_configuration()
    return render(
        request,
        "reports/system_config.html",
        {
            "active_section": "system-config",
            "sidebar_stats": [
                {"label": "Config", "value": SystemDatabaseConfig.objects.count()},
                {"label": "Connections", "value": SystemIntegrationConfig.objects.count()},
                {"label": "Parameters", "value": SystemParameter.objects.count()},
                {"label": "Tables", "value": SystemManagedTable.objects.count()},
            ],
        },
    )


@require_http_methods(["GET"])
def system_configuration_overview_api(request):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    ensure_portable_configuration()
    integrations = SystemIntegrationConfig.objects.all()
    parameters = SystemParameter.objects.filter(is_active=True)
    return JsonResponse({
        "ok": True,
        "summary": {
            "connections": integrations.count(),
            "connected": integrations.filter(status="Connected").count(),
            "configured": integrations.filter(status__in=["Configured", "Connected"]).count(),
            "failed": integrations.filter(status="Failed").count(),
            "parameters": parameters.count(),
            "database_servers": SystemDatabaseConfig.objects.count(),
            "managed_tables": SystemManagedTable.objects.count(),
        },
        "connections": [integration_payload(item, include_schema=False) for item in integrations],
        "categories": list(parameters.values_list("category", flat=True).distinct()),
    })


@require_http_methods(["GET"])
def system_integration_schemas_api(request):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    return JsonResponse({"ok": True, "items": schema_payload()})


@require_http_methods(["GET", "POST"])
def system_integrations_api(request):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    ensure_portable_configuration()
    if request.method == "GET":
        queryset = SystemIntegrationConfig.objects.all()
        integration_type = request.GET.get("type", "").strip()
        query = request.GET.get("q", "").strip()
        if integration_type:
            queryset = queryset.filter(integration_type=integration_type)
        if query:
            queryset = queryset.filter(
                models.Q(name__icontains=query)
                | models.Q(code__icontains=query)
                | models.Q(provider__icontains=query)
                | models.Q(description__icontains=query)
            )
        return JsonResponse({"ok": True, "items": [integration_payload(item) for item in queryset]})
    try:
        item = save_integration(SystemIntegrationConfig(), _request_payload(request), request.user)
        return JsonResponse({"ok": True, "item": integration_payload(item)}, status=201)
    except Exception as exc:
        return _json_error(str(exc))


@require_http_methods(["PUT", "DELETE"])
def system_integration_item_api(request, integration_id):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    item = get_object_or_404(SystemIntegrationConfig, pk=integration_id)
    if request.method == "DELETE":
        item.is_active = False
        item.status = "Disabled"
        item.updated_by = request.user if request.user.is_authenticated else None
        item.save(update_fields=["is_active", "status", "updated_by", "updated_at"])
        return JsonResponse({"ok": True, "deactivated": True})
    try:
        item = save_integration(item, _request_payload(request), request.user)
        return JsonResponse({"ok": True, "item": integration_payload(item)})
    except Exception as exc:
        return _json_error(str(exc))


@require_http_methods(["POST"])
def system_integration_verify_api(request, integration_id):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    item = get_object_or_404(SystemIntegrationConfig, pk=integration_id)
    connected, message = test_integration(item)
    return JsonResponse({
        "ok": connected,
        "status": item.status,
        "message": message,
    }, status=200 if connected else 400)


def _parameter_payload(item):
    return {
        "id": item.pk,
        "key": item.key,
        "category": item.category,
        "label": item.label,
        "description": item.description,
        "value_type": item.value_type,
        "value": item.value_json,
        "default_value": item.default_value_json,
        "options": item.options_json or [],
        "is_required": item.is_required,
        "is_runtime_editable": item.is_runtime_editable,
        "is_active": item.is_active,
        "updated_at": item.updated_at.isoformat() if item.updated_at else "",
    }


def _coerce_system_parameter_value(value_type, value):
    if value in (None, ""):
        return None
    if value_type in {"Integer", "Duration"}:
        return int(value)
    if value_type == "Decimal":
        return float(value)
    if value_type == "Boolean":
        return _ia_normalize_bool(value, False)
    if value_type == "JSON" and isinstance(value, str):
        return json.loads(value)
    return value


@require_http_methods(["GET", "POST"])
def system_parameters_api(request):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    ensure_portable_configuration()
    if request.method == "POST":
        try:
            payload = _request_payload(request)
            key = str(payload.get("key") or "").strip().lower()
            label = str(payload.get("label") or "").strip()
            category = str(payload.get("category") or "General").strip()
            value_type = str(payload.get("value_type") or "Text").strip()
            if not key or not label:
                raise ValueError("Key and label are required.")
            if value_type not in dict(SystemParameter.VALUE_TYPES):
                raise ValueError("Unsupported parameter value type.")
            item = SystemParameter.objects.create(
                key=key,
                label=label,
                category=category,
                description=str(payload.get("description") or "").strip(),
                value_type=value_type,
                value_json=_coerce_system_parameter_value(value_type, payload.get("value")),
                default_value_json=_coerce_system_parameter_value(value_type, payload.get("default_value")),
                options_json=payload.get("options") if isinstance(payload.get("options"), list) else [],
                is_required=_ia_normalize_bool(payload.get("is_required"), False),
                is_runtime_editable=_ia_normalize_bool(payload.get("is_runtime_editable"), True),
                is_active=_ia_normalize_bool(payload.get("is_active"), True),
                created_by=request.user,
                updated_by=request.user,
            )
            return JsonResponse({"ok": True, "item": _parameter_payload(item)}, status=201)
        except Exception as exc:
            return _json_error(str(exc))
    queryset = SystemParameter.objects.all()
    category = request.GET.get("category", "").strip()
    query = request.GET.get("q", "").strip()
    if category:
        queryset = queryset.filter(category=category)
    if query:
        queryset = queryset.filter(
            models.Q(key__icontains=query)
            | models.Q(label__icontains=query)
            | models.Q(description__icontains=query)
        )
    return JsonResponse({"ok": True, "items": [_parameter_payload(item) for item in queryset]})


@require_http_methods(["PUT", "DELETE"])
def system_parameter_item_api(request, parameter_id):
    if not _user_is_platform_admin(request.user):
        return _json_error("Administrator access is required.", status=403)
    item = get_object_or_404(SystemParameter, pk=parameter_id, is_runtime_editable=True)
    if request.method == "DELETE":
        item.is_active = False
        item.updated_by = request.user
        item.save(update_fields=["is_active", "updated_by", "updated_at"])
        return JsonResponse({"ok": True, "deactivated": True})
    payload = _request_payload(request)
    value = payload.get("value")
    try:
        item.value_json = _coerce_system_parameter_value(item.value_type, value)
        item.updated_by = request.user if request.user.is_authenticated else None
        item.save(update_fields=["value_json", "updated_by", "updated_at"])
        return JsonResponse({"ok": True, "item": _parameter_payload(item)})
    except Exception as exc:
        return _json_error(str(exc))


@require_http_methods(["GET", "POST"])
def system_database_configs_api(request):
    _ensure_default_system_config()
    if request.method == "GET":
        query = request.GET.get("q", "").strip()
        queryset = SystemDatabaseConfig.objects.all()
        if query:
            queryset = queryset.filter(
                models.Q(name__icontains=query)
                | models.Q(engine__icontains=query)
                | models.Q(host__icontains=query)
                | models.Q(database_name__icontains=query)
                | models.Q(username__icontains=query)
            )
        return JsonResponse({"ok": True, "items": [_system_db_payload(item) for item in queryset]})

    payload = _request_payload(request)
    item = SystemDatabaseConfig()
    return _save_system_database_config(item, payload, status=201)


def _save_system_database_config(item: SystemDatabaseConfig, payload: dict, status: int = 200):
    try:
        item.name = str(payload.get("name", item.name or "")).strip()
        item.engine = str(payload.get("engine", item.engine or "SQL Server")).strip()
        item.purpose = str(payload.get("purpose", item.purpose or "")).strip()
        item.host = str(payload.get("host", item.host or "")).strip()
        port = payload.get("port", item.port)
        item.port = int(port) if str(port or "").strip() else None
        item.database_name = str(payload.get("database_name", item.database_name or "")).strip()
        item.schema_name = str(payload.get("schema_name", item.schema_name or "")).strip()
        item.username = str(payload.get("username", item.username or "")).strip()
        password = str(payload.get("password", "") or "")
        if password and password != "********":
            item.password = password
        item.driver = str(payload.get("driver", item.driver or "")).strip()
        options = payload.get("connection_options", item.connection_options or {})
        if isinstance(options, str):
            options = json.loads(options) if options.strip() else {}
        item.connection_options = options if isinstance(options, dict) else {}
        item.is_default = _ia_normalize_bool(payload.get("is_default"), item.is_default)
        item.is_active = _ia_normalize_bool(payload.get("is_active"), item.is_active)
        if not item.name or not item.host:
            return _json_error("Name and host are required.")
        if item.is_default:
            SystemDatabaseConfig.objects.exclude(id=item.id).update(is_default=False)
        item.save()
    except Exception as exc:
        return _json_error(str(exc))
    return JsonResponse({"ok": True, "item": _system_db_payload(item)}, status=status)


@require_http_methods(["PUT", "DELETE"])
def system_database_config_item_api(request, config_id):
    item = get_object_or_404(SystemDatabaseConfig, id=config_id)
    if request.method == "DELETE":
        item.is_active = False
        item.save(update_fields=["is_active", "updated_at"])
        return JsonResponse({"ok": True, "deactivated": True})
    return _save_system_database_config(item, _request_payload(request))


@require_http_methods(["POST"])
def system_database_config_verify_api(request, config_id):
    item = get_object_or_404(SystemDatabaseConfig, id=config_id)
    try:
        with connect(
            server=item.host,
            database=item.database_name or "master",
            user=item.username or None,
            password=item.password or None,
            port=item.port or None,
        ) as connection:
            row = connection.cursor().execute("SELECT @@SERVERNAME, DB_NAME()").fetchone()
        item.last_status = "Active"
        item.last_message = f"Connected to {row[0]} / {row[1]}"
        item.last_verified_at = timezone.now()
        item.save(update_fields=["last_status", "last_message", "last_verified_at", "updated_at"])
        return JsonResponse({"ok": True, "status": item.last_status, "message": item.last_message})
    except Exception as exc:
        item.last_status = "Failed"
        item.last_message = str(exc)
        item.last_verified_at = timezone.now()
        item.save(update_fields=["last_status", "last_message", "last_verified_at", "updated_at"])
        return _json_error(str(exc), status=400)


@require_http_methods(["GET", "POST"])
def system_managed_tables_api(request):
    if request.method == "POST":
        count = _refresh_system_table_registry()
        return JsonResponse({"ok": True, "refreshed": count})
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    queryset = SystemManagedTable.objects.select_related("database_config").all()
    if category:
        queryset = queryset.filter(category=category)
    if query:
        queryset = queryset.filter(
            models.Q(table_name__icontains=query)
            | models.Q(schema_name__icontains=query)
            | models.Q(category__icontains=query)
            | models.Q(model_name__icontains=query)
            | models.Q(description__icontains=query)
        )
    return JsonResponse({"ok": True, "items": [_system_table_payload(item) for item in queryset]})


def resources(request):
    query = request.GET.get("q", "")
    selected_section = request.GET.get("section", "")
    selected_category = request.GET.get("category", "")
    selected_level = request.GET.get("level", "")
    resource_source = "Local files"
    resource_items = list_resources(
        query,
        section=selected_section,
        category=selected_category,
        level=selected_level,
    )
    facets = list_resource_facets()
    return render(
        request,
        "reports/resources.html",
        {
            "active_section": "resources",
            "resources": resource_items,
            "resource_count": len(resource_items),
            "query": query,
            "selected_section": selected_section,
            "selected_category": selected_category,
            "selected_level": selected_level,
            "sections": facets["sections"],
            "categories": facets["categories"],
            "levels": facets["levels"],
            "section_cards": facets["section_cards"],
            "resource_source": resource_source,
            "is_platform_admin": _user_is_platform_admin(request.user),
            "sidebar_stats": [
                {"label": "Files", "value": len(resource_items)},
                {"label": "Source", "value": resource_source},
            ],
        },
    )


@require_http_methods(["POST"])
def resource_upload(request):
    uploaded_file = request.FILES.get("file")
    title = request.POST.get("title", "").strip()
    section = request.POST.get("section", "").strip()
    category = request.POST.get("category", "").strip()
    level = request.POST.get("level", "").strip()
    if not uploaded_file:
        messages.error(request, "Document file is required.")
        return redirect("resources")
    if not title:
        title = uploaded_file.name.rsplit(".", 1)[0]
    if not section or not category:
        messages.error(request, "Section and category are required.")
        return redirect("resources")

    try:
        resource = save_uploaded_resource(
            uploaded_file,
            title=title,
            section=section,
            category=category,
            level=level or "General",
        )
        try:
            from .mining360_repository import create_tables, sync_resources

            with connect(database="Mining360") as connection:
                create_tables(connection)
                sync_resources(connection)
        except Exception as exc:
            messages.warning(request, f"Document uploaded, but SQL Server sync failed: {exc}")
        else:
            messages.success(request, f"Document uploaded: {resource.title}.")
        try:
            from .resource_knowledge_index_service import start_index_job

            run = start_index_job(
                user=request.user,
                resource_id=resource.id,
                with_ai=True,
                with_embeddings=True,
            )
            messages.success(
                request,
                f"Knowledge indexing started automatically (run {run.id}).",
            )
        except Exception as exc:
            messages.warning(
                request,
                f"Document uploaded, but knowledge indexing could not start: {exc}",
            )
    except Exception as exc:
        messages.error(request, str(exc))
    return redirect("resources")


def resource_detail(request, resource_id):
    try:
        resource = get_resource(resource_id)
        text_content = ""
        if resource.is_text:
            text_content = read_text_resource(get_resource_path(resource_id))
    except (ValueError, FileNotFoundError):
        raise Http404("Resource not found")

    return render(
        request,
        "reports/resource_detail.html",
        {
            "active_section": "resources",
            "resource": resource,
            "text_content": text_content,
            "sidebar_stats": [
                {"label": "File", "value": resource.extension},
                {"label": "Class", "value": resource.section},
            ],
        },
    )


@xframe_options_sameorigin
def resource_file(request, resource_id):
    try:
        resource = get_resource(resource_id)
        path = get_resource_path(resource_id)
    except (ValueError, FileNotFoundError):
        raise Http404("Resource not found")

    return FileResponse(
        path.open("rb"),
        as_attachment=False,
        filename=resource.filename,
        content_type=resource.mime_type,
    )


BP_PAGES = {
    "overview": ("Overview", "Executive view of fleet and commercial performance."),
    "customers": ("Customers", "Customer portfolio, segmentation and contribution."),
    "parts-sales": ("Parts Sales", "Parts revenue analysis and transactions."),
    "machine-sales": ("Machine Sales", "Prime revenue and machine sales analysis."),
    "forecast": ("Forecast & Opportunities", "Reserved for predictive scenarios."),
}

BUSINESS_PERFORMANCE_ENABLED = False


def _business_performance_access(user) -> bool:
    if _user_is_platform_admin(user):
        return True
    platform_user = getattr(user, "platformuser", None)
    return bool(
        platform_user
        and platform_user.is_active
        and platform_user.can_access_reporting
        and platform_user.business_performance_role
    )


def _bp_filters(request) -> dict:
    filters = {}
    for key in BusinessPerformanceService.FILTER_KEYS:
        values = []
        for raw in request.GET.getlist(key):
            values.extend(item.strip() for item in raw.split(",") if item.strip())
        if values:
            filters[key] = values
    return filters


def _bp_json_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _business_performance_disabled_response(request):
    if wants_json(request):
        return JsonResponse({"ok": False, "error": "Business Performance is currently disabled."}, status=404)
    raise Http404("Business Performance is currently disabled.")


@login_required
def business_performance_home(request, page="overview", customer=None):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    if not _business_performance_access(request.user):
        return JsonResponse({"ok": False, "error": "Business Performance access required."}, status=403)
    page = page if page in BP_PAGES else "overview"
    config = BusinessPerformanceConfig.objects.filter(is_active=True).first()
    return render(
        request,
        "reports/business_performance.html",
        {
            "active_section": "business-performance",
            "page_code": page,
            "page_title": "Customer Details" if customer else BP_PAGES[page][0],
            "page_description": BP_PAGES[page][1],
            "customer": customer or "",
            "config": config,
            "is_bp_admin": _user_is_platform_admin(request.user),
            "bp_pages": BP_PAGES,
        },
    )


@login_required
def business_performance_customer(request, customer):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    return business_performance_home(request, "customers", unquote(customer))


@login_required
def business_performance_api(request, section):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    if not _business_performance_access(request.user):
        return JsonResponse({"ok": False, "error": "Business Performance access required."}, status=403)
    filters = _bp_filters(request)
    try:
        service = BusinessPerformanceService(request.user)
        if section == "overview":
            payload = service.overview(filters, request.GET.get("top_n"))
        elif section == "customers":
            payload = {"customers": service.customers(filters, int(request.GET.get("limit", 500)))}
        elif section in {"parts-sales", "machine-sales"}:
            category = {"parts-sales": "parts", "machine-sales": "prime"}[section]
            payload = {"rows": service.detail_rows(category, filters, int(request.GET.get("limit", 1000)))}
        elif section == "customer":
            customer = (request.GET.get("customer") or "").strip()
            if not customer:
                return JsonResponse({"ok": False, "error": "Customer is required."}, status=400)
            payload = service.customer_details(customer, filters)
        elif section == "filter-options":
            logical_name = (request.GET.get("filter") or "").strip()
            payload = {"filter": logical_name, "items": service.filter_options(logical_name, filters)}
        else:
            return JsonResponse({"ok": False, "error": "Unknown Business Performance endpoint."}, status=404)
        return JsonResponse({"ok": True, **payload})
    except MappingNotConfigured as exc:
        return JsonResponse({"ok": False, "code": "mapping_missing", "error": str(exc)}, status=422)
    except BusinessPerformanceError as exc:
        return JsonResponse({"ok": False, "code": "semantic_model_unavailable", "error": str(exc)}, status=503)
    except (TypeError, ValueError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


@login_required
def business_performance_export(request, category, file_type):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    if not _business_performance_access(request.user):
        return JsonResponse({"ok": False, "error": "Business Performance access required."}, status=403)
    category_map = {"parts": "parts", "prime": "prime", "fleet": "fleet", "customers": "customers"}
    if category not in category_map or file_type not in {"csv", "xlsx"}:
        raise Http404
    try:
        service = BusinessPerformanceService(request.user)
        filters = _bp_filters(request)
        rows = service.customers(filters, 10000) if category == "customers" else service.detail_rows(category, filters, 10000)
    except BusinessPerformanceError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=503)
    columns = list(dict.fromkeys(key for row in rows for key in row.keys()))
    stamp = timezone.localdate().strftime("%Y")
    filename = f"Business_Performance_{category}_{stamp}.{file_type}"
    if file_type == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.write("\ufeff")
        writer = csv.DictWriter(response, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
        return response
    try:
        from openpyxl import Workbook
    except ImportError:
        return JsonResponse({"ok": False, "error": "Excel export requires openpyxl."}, status=500)
    from io import BytesIO
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Business Performance")
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def business_performance_config(request):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    config, _ = BusinessPerformanceConfig.objects.get_or_create(name="Business Performance")
    fleet_mapping_names = {
        "active_fleet", "fleet_share", "parts_revenue_per_fleet", "prime_revenue_per_fleet",
        "total_revenue_per_fleet", "minesite", "equipment_type", "model", "fleet_status",
        "serial_number", "equipment_number",
    }
    return render(request, "reports/business_performance_config.html", {
        "active_section": "business-performance-config",
        "config": config,
        "mappings": BusinessPerformanceMapping.objects.exclude(logical_name__in=fleet_mapping_names),
    })


@login_required
@require_http_methods(["POST"])
def business_performance_config_api(request):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Invalid JSON payload."}, status=400)
    config, _ = BusinessPerformanceConfig.objects.get_or_create(name="Business Performance")
    editable = {
        "workspace_id", "semantic_model_name", "semantic_model_id", "report_id", "tenant_id",
        "authentication_mode", "api_endpoint", "xmla_endpoint", "default_currency",
        "default_date_range", "default_lob", "default_division", "cache_duration_seconds",
        "query_timeout_seconds", "top_n_default", "active_fleet_status_value",
        "opportunity_threshold_mode", "opportunity_fleet_threshold", "opportunity_revenue_threshold",
    }
    for field in editable:
        if field in payload:
            setattr(config, field, payload[field] if payload[field] != "" else None if field.endswith("_threshold") else "")
    config.save()
    for item in payload.get("mappings", []):
        mapping = BusinessPerformanceMapping.objects.filter(id=item.get("id")).first()
        if not mapping:
            continue
        for field in ("display_name", "table_name", "object_name", "data_type", "format_string", "description"):
            if field in item:
                setattr(mapping, field, str(item[field]).strip())
        for field in ("is_active", "is_visible", "is_required"):
            if field in item:
                setattr(mapping, field, bool(item[field]))
        mapping.save()
    cache.clear()
    return JsonResponse({"ok": True, "message": "Business Performance configuration saved."})


@login_required
@require_http_methods(["POST"])
def business_performance_import_model_api(request):
    if not BUSINESS_PERFORMANCE_ENABLED:
        return _business_performance_disabled_response(request)
    if not _user_is_platform_admin(request.user):
        return JsonResponse({"ok": False, "error": "Admin access required."}, status=403)
    config, _ = BusinessPerformanceConfig.objects.get_or_create(name="Business Performance")
    try:
        dataset_id = config.semantic_model_id or resolve_workspace_dataset_id(config.semantic_model_name)
        measures, measure_error = _execute_powerbi_info_query(dataset_id, "MEASURES")
        columns, column_error = _execute_powerbi_info_query(dataset_id, "COLUMNS")
        if measure_error and column_error:
            raise RuntimeError(f"Measures: {measure_error}; Columns: {column_error}")
        config.semantic_model_id = dataset_id
        config.save(update_fields=["semantic_model_id", "updated_at"])
        normalized = lambda value: re.sub(r"[^a-z0-9]+", "", str(value or "").lower())
        measure_catalog = {}
        for row in measures:
            name = str(_ai_row_value(row, "Name", "Measure", "MeasureName") or "").strip()
            if name:
                measure_catalog[normalized(name)] = name
        column_catalog = []
        for row in columns:
            table_name = str(_ai_row_value(row, "Table", "TableName", "SourceTable") or "").strip()
            column_name = str(_ai_row_value(row, "Name", "Column", "ColumnName") or "").strip()
            if table_name and column_name:
                column_catalog.append((table_name, column_name))
        matched = 0
        for mapping in BusinessPerformanceMapping.objects.all():
            candidates = {normalized(mapping.object_name), normalized(mapping.display_name), normalized(mapping.logical_name)}
            if mapping.object_type == "measure":
                match = next((measure_catalog[key] for key in candidates if key in measure_catalog), "")
                if match:
                    mapping.object_name = match
                    mapping.save(update_fields=["object_name", "updated_at"])
                    matched += 1
            else:
                match = next(((table, column) for table, column in column_catalog if normalized(column) in candidates), None)
                if match:
                    mapping.table_name, mapping.object_name = match
                    mapping.save(update_fields=["table_name", "object_name", "updated_at"])
                    matched += 1
        return JsonResponse({
            "ok": True, "dataset_id": dataset_id, "measures_found": len(measure_catalog),
            "columns_found": len(column_catalog), "mappings_matched": matched,
            "message": f"Semantic model imported. {matched} mappings matched; review before use.",
        })
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=503)


def report_detail(request, report_id):
    report = None
    reports = []
    embed_token = None
    error = None
    selected_role = request.GET.get("role") or "Global"

    try:
        reports = list_workspace_reports()
        report = get_workspace_report(str(report_id), reports)
        embed_token = generate_report_embed_token(report, [selected_role])
    except Exception as exc:
        error = str(exc)

    return render(
        request,
        "reports/detail.html",
        {
            "report": report,
            "reports": reports,
            "embed_token": embed_token,
            "error": error,
            "workspace_name": "Efficience Mine Workspace",
            "active_section": "reporting",
            "role_options": RLS_ROLE_OPTIONS,
            "selected_role": selected_role,
            "sidebar_stats": [
                {"label": "Reports", "value": len(reports)},
                {"label": "Role", "value": selected_role},
            ],
        },
    )


def data_quality_run(request):
    if request.method != "POST" or not _is_ajax_request(request):
        raise Http404("Not found")

    source_key = (request.POST.get("source_key") or request.GET.get("source_key") or "").strip()
    preview_url = (request.POST.get("preview_url") or request.GET.get("preview_url") or "").strip()
    control_key = (request.POST.get("control_key") or request.GET.get("control_key") or "").strip() or None
    payload_text = request.POST.get("payload") or request.GET.get("payload") or "{}"

    try:
        payload = json.loads(payload_text) if payload_text else {}
    except Exception:
        payload = {}

    if not source_key:
        return JsonResponse({"ok": False, "error": "Source key is required."}, status=400)
    if not preview_url:
        return JsonResponse({"ok": False, "error": "Preview target is required."}, status=400)

    try:
        source = get_live_source(source_key)
    except KeyError:
        return JsonResponse({"ok": False, "error": "Source not found."}, status=404)

    try:
        run, context, preview_payload, results, summary, score = _run_data_quality(source, preview_url, control_key, payload)
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    return JsonResponse(
        {
            "ok": True,
            "run": {
                "id": run.id,
                "score": float(run.score),
                "status": run.status,
                "source_key": run.source_key,
                "source_name": run.source_name,
                "object_kind": run.object_kind,
                "object_name": run.object_name,
                "created_at": run.created_at.isoformat() if run.created_at else "",
                "finished_at": run.finished_at.isoformat() if run.finished_at else "",
                "total_rows": run.total_rows,
                "controls_count": run.controls_count,
                "summary": run.summary,
                "preview_url": run.request_payload.get("preview_url", ""),
            },
            "results": [serialize_result(result) for result in results],
            "summary": summary,
            "preview": preview_payload.get("preview", {}),
        },
    )


def data_quality_records(request, run_id: int, control_key: str):
    if request.method != "GET" or not _is_ajax_request(request):
        raise Http404("Not found")

    try:
        run = DataQualityRun.objects.get(id=run_id)
    except DataQualityRun.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Run not found."}, status=404)

    results = run.results if isinstance(run.results, list) else []
    result = next((item for item in results if isinstance(item, dict) and item.get("key") == control_key), None)
    if not result:
        return JsonResponse({"ok": False, "error": "Control not found in run."}, status=404)

    records = result.get("records") or []
    columns = list(records[0].keys()) if records else []
    return JsonResponse(
        {
            "ok": True,
            "run": {
                "id": run.id,
                "source_name": run.source_name,
                "object_name": run.object_name,
                "created_at": run.created_at.isoformat() if run.created_at else "",
            },
            "control": result,
            "columns": columns,
            "records": records,
        }
    )


def data_quality_export(request, run_id: int, control_key: str):
    try:
        run = DataQualityRun.objects.get(id=run_id)
    except DataQualityRun.DoesNotExist:
        raise Http404("Run not found")

    results = run.results if isinstance(run.results, list) else []
    result = next((item for item in results if isinstance(item, dict) and item.get("key") == control_key), None)
    if not result:
        raise Http404("Control not found")

    records = result.get("records") or []
    try:
        from io import BytesIO
        from openpyxl import Workbook
    except Exception as exc:
        return JsonResponse({"ok": False, "error": f"Excel export unavailable: {exc}"}, status=500)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (result.get("name") or "Data Quality")[:31]
    columns = list(records[0].keys()) if records else ["message"]
    sheet.append(columns)
    if records:
        for record in records:
            sheet.append([_json_safe(record.get(column)) for column in columns])
    else:
        sheet.append(["No impacted records"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"dq_{run.source_key}_{control_key}_{run.id}.xlsx".replace(" ", "_")
    response = FileResponse(buffer, as_attachment=True, filename=filename)
    response["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response
