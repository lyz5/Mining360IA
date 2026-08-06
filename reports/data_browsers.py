import re
import csv
import json
import tempfile
import threading
from datetime import date, datetime
from io import StringIO
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from django.utils import timezone
from django.db import close_old_connections

from .models import DataBrowser, DataBrowserColumn, DataBrowserSyncLog
from .sqlserver import connect


IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
DATA_TYPE_SQL = {
    "Integer": "INT",
    "Decimal": "DECIMAL(18,2)",
    "Date": "DATE",
    "DateTime": "DATETIME2",
    "Boolean": "BIT",
}
BROWSER_DATABASE = "Mining360"
SYSTEM_SQL_COLUMNS = {
    "browserrecordid", "eventchainid", "createdat", "createdby", "updatedat", "updatedby",
}
IMPORT_SESSION_DIR = Path(tempfile.gettempdir()) / "mining360_import_sessions"
IMPORT_JOB_DIR = Path(tempfile.gettempdir()) / "mining360_import_jobs"
IMPORT_JOB_LOCKS: dict[str, threading.RLock] = {}
IMPORT_JOB_LOCKS_GUARD = threading.Lock()


class DataBrowserValidationError(ValueError):
    pass


def validate_identifier(value: str, label: str = "SQL name") -> str:
    value = (value or "").strip()
    if not IDENTIFIER_RE.match(value):
        raise DataBrowserValidationError(
            f"{label} must start with a letter or underscore and contain only letters, numbers and underscores."
        )
    return value


def parse_sql_object_name(value: str, label: str = "SQL object") -> tuple[str, str]:
    raw = (value or "").strip()
    parts = raw.split(".")
    if len(parts) == 1:
        schema, name = "dbo", parts[0]
    elif len(parts) == 2:
        schema, name = parts
    else:
        raise DataBrowserValidationError(f"{label} must use schema.name or name format.")
    return validate_identifier(schema, f"{label} schema"), validate_identifier(name, f"{label} name")


def quote_identifier(value: str) -> str:
    return f"[{validate_identifier(value)}]"


def quote_object_name(value: str, label: str = "SQL object") -> str:
    schema, name = parse_sql_object_name(value, label)
    return f"{quote_identifier(schema)}.{quote_identifier(name)}"


def audit_user_sql_literal(value: str | None) -> str:
    text = str(value or "System").strip() or "System"
    return "N'" + text[:150].replace("'", "''") + "'"


def sql_type_for_column(column: DataBrowserColumn) -> str:
    if column.data_type == "Text":
        length = int(column.length or 255)
        if length <= 0:
            raise DataBrowserValidationError("Text length must be greater than zero.")
        if length > 4000:
            return "NVARCHAR(MAX)"
        return f"NVARCHAR({length})"
    try:
        return DATA_TYPE_SQL[column.data_type]
    except KeyError as exc:
        raise DataBrowserValidationError(f"Unsupported data type: {column.data_type}") from exc


def default_sql_literal(column: DataBrowserColumn) -> str:
    value = (column.default_value or "").strip()
    if not value:
        return ""
    if column.data_type == "Text":
        return " DEFAULT N'" + value.replace("'", "''") + "'"
    if column.data_type == "Integer":
        int(value)
        return f" DEFAULT {value}"
    if column.data_type == "Decimal":
        Decimal(value)
        return f" DEFAULT {value}"
    if column.data_type == "Boolean":
        normalized = value.lower()
        if normalized in {"-1", "1", "true", "yes", "y"}:
            return " DEFAULT 1"
        if normalized in {"0", "false", "no", "n"}:
            return " DEFAULT 0"
        raise DataBrowserValidationError("Boolean default value must be true/false or 1/0.")
    if column.data_type in {"Date", "DateTime"}:
        return " DEFAULT '" + value.replace("'", "''") + "'"
    return ""


def value_sql_literal(column: DataBrowserColumn, value) -> str:
    if value is None or value == "":
        if column.default_value:
            value = column.default_value
        elif column.is_required:
            raise DataBrowserValidationError(f"{column.display_name} is required.")
        else:
            return "NULL"
    value = str(value).strip()
    if column.data_type == "Text":
        return "N'" + value.replace("'", "''") + "'"
    if column.data_type == "Integer":
        return str(int(value))
    if column.data_type == "Decimal":
        return str(Decimal(value))
    if column.data_type == "Boolean":
        normalized = value.lower()
        if normalized in {"-1", "1", "true", "yes", "y", "on"}:
            return "1"
        if normalized in {"0", "false", "no", "n", "off"}:
            return "0"
        raise DataBrowserValidationError(f"{column.display_name} must be boolean.")
    if column.data_type in {"Date", "DateTime"}:
        temporal_value = _parse_temporal_value(column, value)
        return "'" + temporal_value.isoformat(sep=" ") + "'" if isinstance(temporal_value, datetime) else "'" + temporal_value.isoformat() + "'"
    raise DataBrowserValidationError(f"Unsupported data type: {column.data_type}")


def _parse_temporal_value(column: DataBrowserColumn, value):
    if isinstance(value, datetime):
        return value if column.data_type == "DateTime" else value.date()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()) if column.data_type == "DateTime" else value
    text = str(value or "").strip()
    if not text:
        return None
    if column.data_type == "Date" and re.fullmatch(r"\d{4}-\d{2}", text):
        text += "-01"
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
        return parsed if column.data_type == "DateTime" else parsed.date()
    except ValueError:
        pass
    formats = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
    )
    for date_format in formats:
        try:
            parsed = datetime.strptime(text, date_format)
            return parsed if column.data_type == "DateTime" else parsed.date()
        except ValueError:
            continue
    raise DataBrowserValidationError(
        f"{column.display_name}: '{value}' is not a valid {column.data_type} value."
    )


def import_parameter_value(column: DataBrowserColumn, value):
    """Validate an imported value and return a DB-API parameter value."""
    if value is None or value == "":
        if column.default_value:
            value = column.default_value
        elif column.is_required:
            raise DataBrowserValidationError(f"{column.display_name} is required.")
        else:
            return None
    value = str(value).strip()
    if column.data_type == "Text":
        return value
    if column.data_type == "Integer":
        return int(value)
    if column.data_type == "Decimal":
        return Decimal(value)
    if column.data_type == "Boolean":
        normalized = value.lower()
        if normalized in {"-1", "1", "true", "yes", "y", "on"}:
            return 1
        if normalized in {"0", "false", "no", "n", "off"}:
            return 0
        raise DataBrowserValidationError(f"{column.display_name} must be boolean.")
    if column.data_type in {"Date", "DateTime"}:
        return _parse_temporal_value(column, value)
    raise DataBrowserValidationError(f"Unsupported data type: {column.data_type}")


def column_definition(column: DataBrowserColumn) -> str:
    sql_name = quote_identifier(column.sql_name)
    nullability = "NOT NULL" if column.is_required else "NULL"
    return f"{sql_name} {sql_type_for_column(column)} {nullability}{default_sql_literal(column)}"


def _effective_browser_columns(browser: DataBrowser) -> list[DataBrowserColumn]:
    columns = []
    seen: set[str] = set()
    for column in browser.columns.all():
        sql_name = str(column.sql_name or "").strip().lower()
        if not sql_name or sql_name in seen:
            continue
        seen.add(sql_name)
        columns.append(column)
    return columns


def log_sync(browser: DataBrowser, action: str, status: str, message: str = "", sql_statement: str = "") -> None:
    DataBrowserSyncLog.objects.create(
        browser=browser,
        action=action,
        status=status,
        message=message,
        sql_statement=sql_statement,
    )


def validate_browser_definition(browser: DataBrowser) -> None:
    if browser.source_mode != "managed_table":
        from .external_data_browsers import validate_external_browser

        try:
            validate_external_browser(browser)
        except ValueError as exc:
            raise DataBrowserValidationError(str(exc)) from exc
        return
    parse_sql_object_name(browser.table_name, "Table name")
    parse_sql_object_name(browser.source_view_name, "Source view name")
    for column in browser.columns.all():
        validate_identifier(column.sql_name, "Column SQL name")
        sql_type_for_column(column)
        if column.is_lookup:
            if not column.lookup_source_name:
                raise DataBrowserValidationError(f"Lookup source is required for {column.sql_name}.")
            if not column.lookup_value_column:
                raise DataBrowserValidationError(f"Lookup value column is required for {column.sql_name}.")
            parse_sql_object_name(column.lookup_source_name, "Lookup source")
            validate_identifier(column.lookup_value_column, "Lookup value column")
            if column.lookup_label_column:
                validate_identifier(column.lookup_label_column, "Lookup label column")
            lookup_browser = DataBrowser.objects.filter(table_name=column.lookup_source_name).first()
            if not lookup_browser:
                raise DataBrowserValidationError("Lookup source must be an existing Browser.")
            lookup_columns = {item.sql_name for item in lookup_browser.columns.all()}
            if column.lookup_value_column not in lookup_columns:
                raise DataBrowserValidationError("Lookup value column must belong to the selected Browser.")
            if column.lookup_label_column and column.lookup_label_column not in lookup_columns:
                raise DataBrowserValidationError("Lookup label column must belong to the selected Browser.")
            if column.lookup_filter and column.lookup_filter not in lookup_columns:
                raise DataBrowserValidationError("Lookup filter column must belong to the selected Browser.")
        try:
            default_sql_literal(column)
        except (ValueError, InvalidOperation) as exc:
            raise DataBrowserValidationError(f"Invalid default value for {column.sql_name}.") from exc


def table_exists_sql(browser: DataBrowser) -> str:
    schema, table = parse_sql_object_name(browser.table_name, "Table name")
    return (
        "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES "
        f"WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table}'"
    )


def existing_columns_sql(browser: DataBrowser) -> str:
    schema, table = parse_sql_object_name(browser.table_name, "Table name")
    return (
        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
        f"WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table}'"
    )


def existing_column_nullability_sql(browser: DataBrowser) -> str:
    schema, table = parse_sql_object_name(browser.table_name, "Table name")
    return (
        "SELECT COLUMN_NAME, IS_NULLABLE FROM INFORMATION_SCHEMA.COLUMNS "
        f"WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table}'"
    )


def browser_table_exists(browser: DataBrowser) -> bool:
    if browser.source_mode != "managed_table":
        return True
    validate_browser_definition(browser)
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        cursor.execute(table_exists_sql(browser))
        return bool((cursor.fetchone() or [0])[0])


def create_table_statement(browser: DataBrowser) -> str:
    table = quote_object_name(browser.table_name, "Table name")
    configured_columns = [
        column_definition(column)
        for column in _effective_browser_columns(browser)
        if str(column.sql_name or "").strip().lower() not in SYSTEM_SQL_COLUMNS
    ]
    base_columns = [
        "[BrowserRecordId] BIGINT IDENTITY(1,1) NOT NULL PRIMARY KEY",
        "[EventChainID] INT NOT NULL",
        "[CreatedAt] DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()",
        "[CreatedBy] NVARCHAR(150) NOT NULL DEFAULT N'System'",
        "[UpdatedAt] DATETIME2 NULL",
        "[UpdatedBy] NVARCHAR(150) NULL",
    ]
    all_columns = base_columns + configured_columns
    return f"CREATE TABLE {table} (\n    " + ",\n    ".join(all_columns) + "\n)"


def add_column_statement(browser: DataBrowser, column: DataBrowserColumn) -> str:
    if str(column.sql_name or "").strip().lower() in SYSTEM_SQL_COLUMNS:
        raise DataBrowserValidationError(f"{column.sql_name} is a system column and is managed automatically.")
    return f"ALTER TABLE {quote_object_name(browser.table_name, 'Table name')} ADD {column_definition(column)}"


def add_eventchain_column_statement(browser: DataBrowser) -> str:
    return f"ALTER TABLE {quote_object_name(browser.table_name, 'Table name')} ADD [EventChainID] INT NULL"


def unique_index_name(browser: DataBrowser, column: DataBrowserColumn) -> str:
    _, table = parse_sql_object_name(browser.table_name, "Table name")
    base = f"UX_{table}_{column.sql_name}"
    safe = re.sub(r"[^A-Za-z0-9_]", "_", base)
    return safe[:128]


def unique_index_exists_sql(index_name: str) -> str:
    safe_name = index_name.replace("'", "''")
    return f"SELECT COUNT(*) FROM sys.indexes WHERE name = '{safe_name}'"


def create_unique_index_statement(browser: DataBrowser, column: DataBrowserColumn) -> str:
    index_name = quote_identifier(unique_index_name(browser, column))
    column_name = quote_identifier(column.sql_name)
    return (
        f"CREATE UNIQUE INDEX {index_name} "
        f"ON {quote_object_name(browser.table_name, 'Table name')} ({column_name}) "
        f"WHERE {column_name} IS NOT NULL"
    )


def sync_browser_sql(browser: DataBrowser) -> dict:
    validate_browser_definition(browser)
    if browser.source_mode != "managed_table":
        browser.last_synced_at = timezone.now()
        browser.last_sync_status = "Success"
        browser.last_sync_message = "External browser configuration validated; no table was created."
        browser.save(
            update_fields=[
                "last_synced_at",
                "last_sync_status",
                "last_sync_message",
                "updated_at",
            ]
        )
        log_sync(browser, "validate_external_source", "Success", browser.last_sync_message)
        return {"executed": [], "message": browser.last_sync_message}
    executed = []
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        sql = table_exists_sql(browser)
        cursor.execute(sql)
        exists = bool((cursor.fetchone() or [0])[0])
        if not exists:
            statement = create_table_statement(browser)
            cursor.execute(statement)
            executed.append({"action": "create_table", "sql": statement})
            log_sync(browser, "create_table", "Success", "Table created.", statement)
        else:
            sql = existing_column_nullability_sql(browser)
            cursor.execute(sql)
            existing_nullability = {
                str(row[0]).lower(): str(row[1]).upper() == "YES"
                for row in cursor.fetchall()
            }
            existing = set(existing_nullability)
            if "eventchainid" not in existing:
                statement = add_eventchain_column_statement(browser)
                cursor.execute(statement)
                executed.append({"action": "add_eventchainid", "sql": statement})
                log_sync(browser, "add_eventchainid", "Success", "EventChainID column added.", statement)
            system_columns = {
                "createdat": "[CreatedAt] DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()",
                "createdby": "[CreatedBy] NVARCHAR(150) NOT NULL DEFAULT N'System'",
                "updatedat": "[UpdatedAt] DATETIME2 NULL",
                "updatedby": "[UpdatedBy] NVARCHAR(150) NULL",
            }
            for column_key, definition in system_columns.items():
                if column_key in existing:
                    continue
                statement = (
                    f"ALTER TABLE {quote_object_name(browser.table_name, 'Table name')} "
                    f"ADD {definition}"
                )
                cursor.execute(statement)
                existing.add(column_key)
                executed.append({"action": "add_audit_column", "column": column_key, "sql": statement})
                log_sync(browser, "add_audit_column", "Success", f"Audit column {column_key} added.", statement)
            for column in _effective_browser_columns(browser):
                if column.sql_name.lower() in SYSTEM_SQL_COLUMNS:
                    continue
                if column.sql_name.lower() in existing:
                    continue
                statement = add_column_statement(browser, column)
                cursor.execute(statement)
                executed.append({"action": "add_column", "column": column.sql_name, "sql": statement})
                log_sync(browser, "add_column", "Success", f"Column {column.sql_name} added.", statement)
            for column in _effective_browser_columns(browser):
                column_key = column.sql_name.lower()
                if column_key in SYSTEM_SQL_COLUMNS or column_key not in existing_nullability:
                    continue
                should_allow_null = not column.is_required
                if existing_nullability[column_key] == should_allow_null:
                    continue
                nullability = "NULL" if should_allow_null else "NOT NULL"
                statement = (
                    f"ALTER TABLE {quote_object_name(browser.table_name, 'Table name')} "
                    f"ALTER COLUMN {quote_identifier(column.sql_name)} "
                    f"{sql_type_for_column(column)} {nullability}"
                )
                cursor.execute(statement)
                executed.append({"action": "alter_nullability", "column": column.sql_name, "sql": statement})
                log_sync(
                    browser,
                    "alter_nullability",
                    "Success",
                    f"Column {column.sql_name} nullability updated.",
                    statement,
                )
        for column in _effective_browser_columns(browser):
            if column.sql_name.lower() in SYSTEM_SQL_COLUMNS:
                continue
            if not column.is_unique:
                continue
            index_name = unique_index_name(browser, column)
            cursor.execute(unique_index_exists_sql(index_name))
            if bool((cursor.fetchone() or [0])[0]):
                continue
            statement = create_unique_index_statement(browser, column)
            cursor.execute(statement)
            executed.append({"action": "add_unique_index", "column": column.sql_name, "sql": statement})
            log_sync(browser, "add_unique_index", "Success", f"Unique index {index_name} created.", statement)

    browser.last_synced_at = timezone.now()
    browser.last_sync_status = "Success"
    browser.last_sync_message = f"{len(executed)} SQL action(s) executed."
    browser.save(update_fields=["last_synced_at", "last_sync_status", "last_sync_message", "updated_at"])
    if not executed:
        log_sync(browser, "sync_sql", "Success", "No SQL changes required.")
    return {"executed": executed, "message": browser.last_sync_message}


def _normalize_preview_limit(limit) -> int | None:
    raw = str(limit or "100").strip().lower()
    if raw in {"all", "0", "-1"}:
        return None
    try:
        return max(1, min(int(raw), 1000))
    except Exception as exc:
        raise DataBrowserValidationError("Invalid preview limit.") from exc


def _normalize_preview_operator(value) -> str:
    return "like" if str(value or "").strip().lower() == "like" else "equal"


def _normalize_preview_joiner(value) -> str:
    return "OR" if str(value or "").strip().upper() == "OR" else "AND"


def _normalize_preview_type(value) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in {"integer", "decimal", "number", "int", "float", "double"}:
        return "number"
    if normalized in {"date", "datetime", "datetime2"}:
        return "date"
    return "text"


def _preview_column_sql_name(browser: DataBrowser, visible_columns: list[DataBrowserColumn], index: int) -> str:
    if index == 0:
        return "BrowserRecordId"
    if index == 1:
        return "EventChainID"
    visible_index = index - 2
    if visible_index < 0 or visible_index >= len(visible_columns):
        raise DataBrowserValidationError("Invalid preview column index.")
    return visible_columns[visible_index].sql_name


def _preview_column_type(browser: DataBrowser, visible_columns: list[DataBrowserColumn], index: int) -> str:
    if index in {0, 1}:
        return "number"
    visible_index = index - 2
    if visible_index < 0 or visible_index >= len(visible_columns):
        return "text"
    return _normalize_preview_type(visible_columns[visible_index].data_type)


def _parse_preview_filters(filters) -> list[dict]:
    if not filters:
        return []
    if isinstance(filters, list):
        raw_filters = filters
    else:
        try:
            raw_filters = json.loads(str(filters))
        except Exception as exc:
            raise DataBrowserValidationError("Invalid preview filters.") from exc
    if not isinstance(raw_filters, list):
        raise DataBrowserValidationError("Invalid preview filters.")
    normalized = []
    for item in raw_filters:
        if not isinstance(item, dict):
            continue
        normalized.append({
            "columnIndex": int(item.get("columnIndex") or 0),
            "operator": _normalize_preview_operator(item.get("operator")),
            "value": str(item.get("value") or ""),
            "joiner": _normalize_preview_joiner(item.get("joiner")),
        })
    return normalized


def _preview_filter_clause(browser: DataBrowser, visible_columns: list[DataBrowserColumn], filter_data: dict) -> str:
    value = str(filter_data.get("value") or "").strip()
    if not value:
        return "1 = 1"
    column_index = int(filter_data.get("columnIndex") or 0)
    column_name = _preview_column_sql_name(browser, visible_columns, column_index)
    column_sql = quote_identifier(column_name)
    column_type = _preview_column_type(browser, visible_columns, column_index)
    operator = _normalize_preview_operator(filter_data.get("operator"))
    if column_type == "date":
        quoted_value = value[:10].replace("'", "''")
        return f"CONVERT(date, {column_sql}) = '{quoted_value}'"
    if operator == "like":
        like_value = value.replace("'", "''")
        return f"{column_sql} LIKE N'%{like_value}%'"
    if column_type == "number":
        try:
            numeric_value = str(Decimal(value))
        except Exception as exc:
            raise DataBrowserValidationError("Numeric filter value is invalid.") from exc
        return f"{column_sql} = {numeric_value}"
    quoted_value = value.replace("'", "''")
    return f"{column_sql} = N'{quoted_value}'"


def _build_preview_where_clause(browser: DataBrowser, visible_columns: list[DataBrowserColumn], filters: list[dict]) -> str:
    active_filters = [item for item in filters if str(item.get("value") or "").strip()]
    if not active_filters:
        return ""
    clauses = []
    for index, filter_data in enumerate(active_filters):
        clause = _preview_filter_clause(browser, visible_columns, filter_data)
        if index == 0:
            clauses.append(clause)
        else:
            clauses.append(f"{_normalize_preview_joiner(filter_data.get('joiner'))} {clause}")
    return " WHERE " + " ".join(clauses)


def preview_browser_data(
    browser: DataBrowser,
    limit: int | str = 100,
    filters=None,
    page: int | str = 1,
    sort=None,
) -> dict:
    if browser.source_mode != "managed_table":
        from .external_data_browsers import preview_external_browser_data

        try:
            return preview_external_browser_data(
                browser,
                limit=limit,
                filters=filters,
                page=page,
                sort=sort,
            )
        except ValueError as exc:
            raise DataBrowserValidationError(str(exc)) from exc
    validate_browser_definition(browser)
    limit_value = _normalize_preview_limit(limit)
    visible_columns = [column for column in _effective_browser_columns(browser) if column.is_visible]
    column_definitions = [
        {"id": None, "sql_name": "BrowserRecordId", "display_name": "BrowserRecordId", "data_type": "Integer", "is_lookup": False},
        {"id": None, "sql_name": "EventChainID", "display_name": "EventChainID", "data_type": "Integer", "is_lookup": False},
    ] + [
        {
            "id": column.id,
            "sql_name": column.sql_name,
            "display_name": column.display_name,
            "data_type": column.data_type,
            "is_lookup": column.is_lookup,
        }
        for column in visible_columns
    ] + [
        {"id": None, "sql_name": "CreatedAt", "display_name": "Created On", "data_type": "DateTime", "is_lookup": False},
        {"id": None, "sql_name": "CreatedBy", "display_name": "Created By", "data_type": "Text", "is_lookup": False},
        {"id": None, "sql_name": "UpdatedAt", "display_name": "Modified On", "data_type": "DateTime", "is_lookup": False},
        {"id": None, "sql_name": "UpdatedBy", "display_name": "Modified By", "data_type": "Text", "is_lookup": False},
    ]
    preview_column_types = (
        ["number", "number"]
        + [_normalize_preview_type(column.data_type) for column in visible_columns]
        + ["date", "text", "date", "text"]
    )
    preview_filters = _parse_preview_filters(filters)
    if not browser_table_exists(browser):
        return {
            "columns": [item["display_name"] for item in column_definitions],
            "column_types": preview_column_types,
            "column_definitions": column_definitions,
            "rows": [],
            "row_values": [],
            "row_count": 0,
            "sql": "",
            "limit": "all" if limit_value is None else limit_value,
            "needs_sync": True,
            "message": "SQL table does not exist yet. Click Sync with SQL Server first.",
        }
    if visible_columns:
        data_select = ", ".join(quote_identifier(column.sql_name) for column in visible_columns)
        select_list = "[BrowserRecordId], [EventChainID]"
        if data_select:
            select_list += ", " + data_select
        select_list += ", [CreatedAt], [CreatedBy], [UpdatedAt], [UpdatedBy]"
        display_columns = [item["display_name"] for item in column_definitions]
    else:
        select_list = "*"
        display_columns = []
    top_clause = "" if limit_value is None else f"TOP ({limit_value}) "
    where_clause = _build_preview_where_clause(browser, visible_columns, preview_filters)
    sql = f"SELECT {top_clause}{select_list} FROM {quote_object_name(browser.table_name, 'Table name')}{where_clause} ORDER BY [BrowserRecordId] DESC"
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        sql_columns = [column[0] for column in cursor.description or []]
    columns = display_columns if display_columns and len(display_columns) == len(sql_columns) else sql_columns
    values = [[_json_safe(value) for value in row] for row in rows]
    records = [dict(zip(columns, row)) for row in values]
    return {
        "columns": columns,
        "column_types": preview_column_types,
        "column_definitions": column_definitions,
        "rows": records,
        "row_values": values,
        "row_count": len(values),
        "sql": sql,
        "limit": "all" if limit_value is None else limit_value,
        "filters": preview_filters,
    }


def update_browser_record(browser: DataBrowser, record_id: int, values: dict, actor: str = "System") -> dict:
    if browser.source_mode != "managed_table":
        from .external_data_browsers import ensure_external_write_allowed

        try:
            ensure_external_write_allowed(browser, "edit")
        except ValueError as exc:
            raise DataBrowserValidationError(str(exc)) from exc
    validate_browser_definition(browser)
    if not browser_table_exists(browser):
        raise DataBrowserValidationError("SQL table does not exist yet. Click Sync with SQL Server first.")
    record_id = int(record_id)
    assignments = []
    columns = _effective_browser_columns(browser)
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        _validate_lookup_values(values, columns, cursor)
        _validate_protected_update(browser, record_id, values, cursor)
        for column in columns:
            if column.sql_name.lower() in {"browserrecordid", "createdat", "updatedat"}:
                continue
            if column.sql_name in values or column.display_name in values:
                value = values.get(column.sql_name, values.get(column.display_name))
                assignments.append(f"{quote_identifier(column.sql_name)} = {value_sql_literal(column, value)}")
        if not assignments:
            raise DataBrowserValidationError("No editable value was provided.")
        assignments.append("[UpdatedAt] = SYSUTCDATETIME()")
        assignments.append(f"[UpdatedBy] = {audit_user_sql_literal(actor)}")
        statement = (
            f"UPDATE {quote_object_name(browser.table_name, 'Table name')} "
            f"SET {', '.join(assignments)} "
            f"WHERE [BrowserRecordId] = {record_id}"
        )
        cursor.execute(statement)
    return {"record_id": record_id, "message": "Record updated."}


def delete_browser_records(browser: DataBrowser, record_ids: list[int]) -> dict:
    if browser.source_mode != "managed_table":
        from .external_data_browsers import ensure_external_write_allowed

        try:
            ensure_external_write_allowed(browser, "delete")
        except ValueError as exc:
            raise DataBrowserValidationError(str(exc)) from exc
    validate_browser_definition(browser)
    if not browser_table_exists(browser):
        raise DataBrowserValidationError("SQL table does not exist yet.")
    ids = sorted({int(item) for item in record_ids if str(item).strip()})
    if not ids:
        raise DataBrowserValidationError("No record selected.")
    id_list = ", ".join(str(item) for item in ids)
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        dependencies = _record_dependencies(browser, ids, cursor)
        if dependencies:
            details = "; ".join(
                f"{item['browser']}.{item['column']} uses '{item['value']}' "
                f"in {item['count']} record(s)"
                for item in dependencies[:10]
            )
            raise DataBrowserValidationError(
                f"Deletion blocked because the selected record(s) are referenced: {details}. "
                "Update or remove the dependent records first."
            )
        cursor.execute(
            f"SELECT EventChainID FROM {quote_object_name(browser.table_name, 'Table name')} "
            f"WHERE [BrowserRecordId] IN ({id_list})"
        )
        eventchain_ids = [int(row[0]) for row in cursor.fetchall() if row[0] is not None]
        cursor.execute(
            f"DELETE FROM {quote_object_name(browser.table_name, 'Table name')} "
            f"WHERE [BrowserRecordId] IN ({id_list})"
        )
        if eventchain_ids:
            eventchain_list = ", ".join(str(item) for item in eventchain_ids)
            cursor.execute(f"DELETE FROM dbo.EventChain WHERE EventChainID IN ({eventchain_list})")
    return {"deleted": len(ids), "record_ids": ids}


def lookup_options(column: DataBrowserColumn, limit: int | None = 500) -> dict:
    if not column.is_lookup:
        raise DataBrowserValidationError("This column is not configured as a lookup.")
    parse_sql_object_name(column.lookup_source_name, "Lookup source")
    value_column = validate_identifier(column.lookup_value_column, "Lookup value column")
    label_column = validate_identifier(column.lookup_label_column or column.lookup_value_column, "Lookup label column")
    limit = None if limit is None else max(1, int(limit or 500))
    predicates = [
        f"{quote_identifier(value_column)} IS NOT NULL",
        f"{quote_identifier(label_column)} IS NOT NULL",
    ]
    if column.lookup_filter.strip():
        filter_column = validate_identifier(column.lookup_filter.strip(), "Lookup filter column")
        predicates.append(f"{quote_identifier(filter_column)} = 1")
    top_clause = "" if limit is None else f"TOP ({limit}) "
    where_sql = f" WHERE {' AND '.join(predicates)}"
    sql = (
        f"SELECT DISTINCT {top_clause}"
        f"{quote_identifier(value_column)} AS [value], "
        f"{quote_identifier(label_column)} AS [label] "
        f"FROM {quote_object_name(column.lookup_source_name, 'Lookup source')}"
        f"{where_sql} "
        f"ORDER BY {quote_identifier(label_column)}"
    )
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
    return {
        "options": [
            {"value": _json_safe(row[0]), "label": _json_safe(row[1])}
            for row in rows
        ],
        "sql": sql,
        "limit": "all" if limit is None else limit,
    }


def _configured_column(browser: DataBrowser, sql_name: str) -> DataBrowserColumn | None:
    expected = str(sql_name or "").strip().lower()
    return next(
        (column for column in browser.columns.all() if str(column.sql_name).strip().lower() == expected),
        None,
    )


def _same_sql_object(left: str, right: str) -> bool:
    try:
        return tuple(part.lower() for part in parse_sql_object_name(left)) == tuple(
            part.lower() for part in parse_sql_object_name(right)
        )
    except DataBrowserValidationError:
        return False


def _incoming_lookup_columns(browser: DataBrowser) -> list[DataBrowserColumn]:
    return [
        column
        for column in DataBrowserColumn.objects.select_related("browser").filter(is_lookup=True)
        if _same_sql_object(column.lookup_source_name, browser.table_name)
    ]


def _lookup_validation_cache(
    columns: list[DataBrowserColumn], cursor
) -> dict[int, set[str]]:
    cache: dict[int, set[str]] = {}
    for column in columns:
        if not column.is_lookup:
            continue
        source_browser = DataBrowser.objects.prefetch_related("columns").filter(
            table_name=column.lookup_source_name
        ).first()
        if not source_browser:
            raise DataBrowserValidationError(
                f"Lookup source {column.lookup_source_name} for {column.display_name} was not found."
            )
        target_column = _configured_column(source_browser, column.lookup_value_column)
        if not target_column:
            raise DataBrowserValidationError(
                f"Lookup column {column.lookup_value_column} for {column.display_name} was not found."
            )
        predicates = [f"{quote_identifier(target_column.sql_name)} IS NOT NULL"]
        if column.lookup_filter.strip():
            predicates.append(f"{quote_identifier(column.lookup_filter.strip())} = 1")
        cursor.execute(
            f"SELECT DISTINCT {quote_identifier(target_column.sql_name)} "
            f"FROM {quote_object_name(source_browser.table_name, 'Lookup source')} "
            f"WHERE {' AND '.join(predicates)}"
        )
        cache[column.id] = {
            normalized
            for row in cursor.fetchall()
            if (normalized := _normalize_import_key(row[0]))
        }
    return cache


def _validate_lookup_values(
    values: dict,
    columns: list[DataBrowserColumn],
    cursor,
    cache: dict[int, set[str]] | None = None,
) -> None:
    lookup_columns = [column for column in columns if column.is_lookup]
    if not lookup_columns:
        return
    cache = cache if cache is not None else _lookup_validation_cache(lookup_columns, cursor)
    for column in lookup_columns:
        value = values.get(column.sql_name, values.get(column.display_name))
        if value in (None, ""):
            continue
        if _normalize_import_key(value) not in cache.get(column.id, set()):
            source_browser = DataBrowser.objects.filter(table_name=column.lookup_source_name).first()
            source_label = source_browser.name if source_browser else column.lookup_source_name
            raise DataBrowserValidationError(
                f"{column.display_name}: value '{value}' does not exist in "
                f"{source_label}.{column.lookup_value_column}. Select an existing value."
            )


def _dependent_record_count(
    relation: DataBrowserColumn,
    value,
    cursor,
    excluded_record_ids: list[int] | None = None,
) -> int:
    predicate = (
        f"{quote_identifier(relation.sql_name)} = {value_sql_literal(relation, value)}"
    )
    if excluded_record_ids:
        ids = ", ".join(str(int(item)) for item in excluded_record_ids)
        predicate += f" AND [BrowserRecordId] NOT IN ({ids})"
    cursor.execute(
        f"SELECT COUNT(*) FROM {quote_object_name(relation.browser.table_name, 'Dependent Browser')} "
        f"WHERE {predicate}"
    )
    return int((cursor.fetchone() or [0])[0])


def _remaining_source_value_count(
    browser: DataBrowser,
    target_column: DataBrowserColumn,
    value,
    excluded_record_ids: list[int],
    cursor,
) -> int:
    ids = ", ".join(str(int(item)) for item in excluded_record_ids)
    cursor.execute(
        f"SELECT COUNT(*) FROM {quote_object_name(browser.table_name, 'Browser table')} "
        f"WHERE {quote_identifier(target_column.sql_name)} = "
        f"{value_sql_literal(target_column, value)} AND [BrowserRecordId] NOT IN ({ids})"
    )
    return int((cursor.fetchone() or [0])[0])


def _record_dependencies(
    browser: DataBrowser, record_ids: list[int], cursor
) -> list[dict]:
    dependencies = []
    for relation in _incoming_lookup_columns(browser):
        target_column = _configured_column(browser, relation.lookup_value_column)
        if not target_column:
            continue
        ids = ", ".join(str(int(item)) for item in record_ids)
        cursor.execute(
            f"SELECT DISTINCT {quote_identifier(target_column.sql_name)} "
            f"FROM {quote_object_name(browser.table_name, 'Browser table')} "
            f"WHERE [BrowserRecordId] IN ({ids}) AND "
            f"{quote_identifier(target_column.sql_name)} IS NOT NULL"
        )
        for row in cursor.fetchall():
            value = row[0]
            if _remaining_source_value_count(browser, target_column, value, record_ids, cursor):
                continue
            excluded_ids = record_ids if _same_sql_object(relation.browser.table_name, browser.table_name) else None
            count = _dependent_record_count(relation, value, cursor, excluded_ids)
            if count:
                dependencies.append({
                    "browser": relation.browser.name,
                    "column": relation.display_name,
                    "value": _json_safe(value),
                    "count": count,
                })
    return dependencies


def _validate_protected_update(
    browser: DataBrowser,
    record_id: int,
    values: dict,
    cursor,
) -> None:
    relations = _incoming_lookup_columns(browser)
    if not relations:
        return
    for relation in relations:
        target_column = _configured_column(browser, relation.lookup_value_column)
        if not target_column:
            continue
        supplied = target_column.sql_name in values or target_column.display_name in values
        if not supplied:
            continue
        new_value = values.get(target_column.sql_name, values.get(target_column.display_name))
        cursor.execute(
            f"SELECT {quote_identifier(target_column.sql_name)} "
            f"FROM {quote_object_name(browser.table_name, 'Browser table')} "
            f"WHERE [BrowserRecordId] = {int(record_id)}"
        )
        row = cursor.fetchone()
        if not row or _normalize_import_key(row[0]) == _normalize_import_key(new_value):
            continue
        old_value = row[0]
        if old_value in (None, ""):
            continue
        if _remaining_source_value_count(browser, target_column, old_value, [record_id], cursor):
            continue
        excluded_ids = [record_id] if _same_sql_object(relation.browser.table_name, browser.table_name) else None
        count = _dependent_record_count(relation, old_value, cursor, excluded_ids)
        if count:
            raise DataBrowserValidationError(
                f"Modification blocked: value '{old_value}' is used by {count} record(s) in "
                f"{relation.browser.name}.{relation.display_name}. Update those records first."
            )


def _json_safe(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return value.isoformat()
    return str(value)


def _read_import_rows(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        text = uploaded_file.read().decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))
        headers = list(rows[0].keys()) if rows else []
        return headers, rows
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DataBrowserValidationError("Excel import requires openpyxl.") from exc
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in raw_rows[0]] if raw_rows else []
        rows = [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in raw_rows[1:]
        ]
        return headers, rows
    raise DataBrowserValidationError("Import file must be CSV or XLSX.")


def inspect_import_file(uploaded_file, sample_size: int = 5) -> dict:
    headers, rows = _read_import_rows(uploaded_file)
    IMPORT_SESSION_DIR.mkdir(parents=True, exist_ok=True)
    token = uuid4().hex
    session_path = IMPORT_SESSION_DIR / f"{token}.json"
    session_path.write_text(
        json.dumps({"headers": headers, "rows": rows, "file_name": uploaded_file.name}, ensure_ascii=False, default=_json_safe),
        encoding="utf-8",
    )
    sample_size = max(1, min(int(sample_size or 5), 20))
    sample_rows = rows[:sample_size]
    return {
        "token": token,
        "headers": headers,
        "sample_rows": [{key: _json_safe(value) for key, value in row.items()} for row in sample_rows],
        "total_rows": len(rows),
        "file_name": uploaded_file.name,
    }


def _load_import_session(token: str) -> dict:
    session_path = IMPORT_SESSION_DIR / f"{str(token or '').strip()}.json"
    if not session_path.exists():
        raise DataBrowserValidationError("Import session expired. Please select the file again.")
    payload = json.loads(session_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise DataBrowserValidationError("Import session is invalid.")
    return payload


def _job_path(job_token: str) -> Path:
    return IMPORT_JOB_DIR / f"{str(job_token or '').strip()}.json"


def _job_lock(job_token: str) -> threading.RLock:
    token = str(job_token or "").strip()
    with IMPORT_JOB_LOCKS_GUARD:
        return IMPORT_JOB_LOCKS.setdefault(token, threading.RLock())


def _write_import_job(job_token: str, payload: dict) -> None:
    IMPORT_JOB_DIR.mkdir(parents=True, exist_ok=True)
    path = _job_path(job_token)
    temporary_path = path.with_name(f"{path.stem}.{uuid4().hex}.tmp")
    serialized = json.dumps(payload, ensure_ascii=False, default=_json_safe)
    with _job_lock(job_token):
        try:
            temporary_path.write_text(serialized, encoding="utf-8")
            last_error = None
            for _ in range(10):
                try:
                    temporary_path.replace(path)
                    return
                except PermissionError as exc:
                    last_error = exc
                    threading.Event().wait(0.05)

            # Windows can briefly prevent atomic replacement when antivirus or
            # another process opens the target. The in-process lock still keeps
            # Django readers away during this fallback write.
            try:
                path.write_text(serialized, encoding="utf-8")
                return
            except OSError as exc:
                raise DataBrowserValidationError(
                    f"Unable to publish import progress after retries: {exc or last_error}"
                ) from exc
        finally:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass


def _load_import_job(job_token: str) -> dict:
    path = _job_path(job_token)
    if not path.exists():
        raise DataBrowserValidationError("Import job was not found.")
    payload = None
    last_error = None
    with _job_lock(job_token):
        for _ in range(3):
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
                break
            except (OSError, json.JSONDecodeError) as exc:
                last_error = exc
                threading.Event().wait(0.05)
    if payload is None:
        raise DataBrowserValidationError(f"Import status is temporarily unavailable: {last_error}")
    if not isinstance(payload, dict):
        raise DataBrowserValidationError("Import job is invalid.")
    return payload


def _build_import_row_values(browser: DataBrowser, row: dict, column_map: dict) -> dict:
    values = {}
    browser_columns = list(browser.columns.all())
    for column in browser_columns:
        column_mapping = column_map.get(column.sql_name, {})
        source_column = str(column_mapping.get("source_column") or "").strip()
        default_value = column_mapping.get("default_value")
        if source_column and source_column in row and str(row.get(source_column, "")).strip() != "":
            values[column.sql_name] = row.get(source_column)
        elif default_value not in (None, ""):
            values[column.sql_name] = default_value
        elif column.sql_name in row:
            values[column.sql_name] = row.get(column.sql_name)
        elif column.display_name in row:
            values[column.sql_name] = row.get(column.display_name)
        else:
            values[column.sql_name] = None
    return values


def _next_eventchain_id(cursor) -> int:
    cursor.execute("SELECT ISNULL(MAX(EventChainID), 0) + 1 FROM dbo.EventChain WITH (TABLOCKX)")
    return int((cursor.fetchone() or [1])[0])


def _normalize_import_key(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return value.isoformat()
    return str(value).strip()


def _build_existing_record_index(browser: DataBrowser, columns: list[DataBrowserColumn], cursor) -> dict[str, dict[str, int]]:
    index: dict[str, dict[str, int]] = {}
    unique_columns = [
        column for column in columns
        if column.is_unique and column.sql_name.lower() not in SYSTEM_SQL_COLUMNS
    ]
    for column in unique_columns:
        key_map: dict[str, int] = {}
        cursor.execute(
            f"SELECT [BrowserRecordId], {quote_identifier(column.sql_name)} "
            f"FROM {quote_object_name(browser.table_name, 'Table name')} "
            f"WHERE {quote_identifier(column.sql_name)} IS NOT NULL"
        )
        for record_id, value in cursor.fetchall():
            normalized = _normalize_import_key(value)
            if normalized:
                key_map[normalized] = int(record_id)
        index[column.sql_name] = key_map
    return index


def insert_browser_record(browser: DataBrowser, values: dict, actor: str = "System") -> dict:
    if browser.source_mode != "managed_table":
        from .external_data_browsers import ensure_external_write_allowed

        try:
            ensure_external_write_allowed(browser, "create")
        except ValueError as exc:
            raise DataBrowserValidationError(str(exc)) from exc
    validate_browser_definition(browser)
    if not browser_table_exists(browser):
        sync_browser_sql(browser)
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        eventchain_id = _insert_browser_record_with_cursor(browser, values, cursor, actor=actor)
    return {"eventchain_id": eventchain_id, "message": "Record inserted."}


def _insert_browser_record_with_cursor(
    browser: DataBrowser,
    values: dict,
    cursor,
    columns: list[DataBrowserColumn] | None = None,
    next_eventchain_id: int | None = None,
    actor: str = "System",
    lookup_cache: dict[int, set[str]] | None = None,
) -> int:
    columns = columns or _effective_browser_columns(browser)
    _validate_lookup_values(values, columns, cursor, cache=lookup_cache)
    eventchain_id_value = values.get("EventChainID", values.get("eventchainid"))
    explicit_eventchain = eventchain_id_value not in (None, "")
    if explicit_eventchain:
        eventchain_id = int(eventchain_id_value)
    elif next_eventchain_id is not None:
        eventchain_id = int(next_eventchain_id)
    else:
        eventchain_id = _next_eventchain_id(cursor)
    if explicit_eventchain:
        cursor.execute(f"SELECT COUNT(*) FROM dbo.EventChain WHERE EventChainID = {eventchain_id}")
        exists = bool((cursor.fetchone() or [0])[0])
    else:
        exists = False
    if not exists:
        cursor.execute(
            "INSERT INTO dbo.EventChain (EventChainID, EventChainTypeID, Created_By, User_ID) "
            f"VALUES ({eventchain_id}, 1, 0, 0)"
        )
    sql_columns = ["EventChainID", "CreatedBy"]
    sql_values = [str(eventchain_id), audit_user_sql_literal(actor)]
    for column in columns:
        if column.sql_name.lower() in {"browserrecordid", "createdat", "updatedat", "eventchainid"}:
            continue
        sql_columns.append(quote_identifier(column.sql_name))
        sql_values.append(value_sql_literal(column, values.get(column.sql_name, values.get(column.display_name))))
    statement = (
        f"INSERT INTO {quote_object_name(browser.table_name, 'Table name')} "
        f"({', '.join(sql_columns)}) VALUES ({', '.join(sql_values)})"
    )
    cursor.execute(statement)
    return eventchain_id


def _existing_import_record_id(browser: DataBrowser, values: dict, cursor, columns: list[DataBrowserColumn]) -> int | None:
    predicates = []
    for column in columns:
        if not column.is_unique:
            continue
        value = values.get(column.sql_name, values.get(column.display_name))
        if value in (None, ""):
            continue
        predicates.append(
            f"{quote_identifier(column.sql_name)} = {value_sql_literal(column, value)}"
        )
    eventchain_id = values.get("EventChainID", values.get("eventchainid"))
    if eventchain_id not in (None, ""):
        predicates.append(f"[EventChainID] = {int(eventchain_id)}")
    if not predicates:
        return None

    cursor.execute(
        f"SELECT TOP (2) [BrowserRecordId] "
        f"FROM {quote_object_name(browser.table_name, 'Table name')} "
        f"WHERE {' OR '.join(predicates)}"
    )
    matches = [int(row[0]) for row in cursor.fetchall()]
    if len(matches) > 1:
        raise DataBrowserValidationError(
            "Unique values match multiple existing records; the row cannot be updated safely."
        )
    return matches[0] if matches else None


def _existing_import_record_id_from_index(
    values: dict,
    columns: list[DataBrowserColumn],
    existing_index: dict[str, dict[str, int]],
) -> int | None:
    matches = []
    for column in columns:
        if not column.is_unique:
            continue
        value = values.get(column.sql_name, values.get(column.display_name))
        normalized = _normalize_import_key(value)
        if not normalized:
            continue
        record_id = (existing_index.get(column.sql_name) or {}).get(normalized)
        if record_id is not None:
            matches.append(int(record_id))
    if not matches:
        return None
    unique_matches = sorted(set(matches))
    if len(unique_matches) > 1:
        raise DataBrowserValidationError(
            "Unique values match multiple existing records; the row cannot be updated safely."
        )
    return unique_matches[0]


def _update_import_record_with_cursor(
    browser: DataBrowser,
    record_id: int,
    values: dict,
    cursor,
    columns: list[DataBrowserColumn],
    actor: str = "System",
    lookup_cache: dict[int, set[str]] | None = None,
) -> None:
    _validate_lookup_values(values, columns, cursor, cache=lookup_cache)
    _validate_protected_update(browser, record_id, values, cursor)
    assignments = []
    for column in columns:
        if column.sql_name.lower() in {"browserrecordid", "createdat", "updatedat", "eventchainid"}:
            continue
        value = values.get(column.sql_name, values.get(column.display_name))
        assignments.append(
            f"{quote_identifier(column.sql_name)} = {value_sql_literal(column, value)}"
        )
    assignments.append("[UpdatedAt] = SYSUTCDATETIME()")
    assignments.append(f"[UpdatedBy] = {audit_user_sql_literal(actor)}")
    cursor.execute(
        f"UPDATE {quote_object_name(browser.table_name, 'Table name')} "
        f"SET {', '.join(assignments)} WHERE [BrowserRecordId] = {int(record_id)}"
    )


def _execute_import_rows(cursor, insert_sql: str, rows: list[tuple], placeholder: str) -> None:
    if not rows:
        return
    if not cursor.__class__.__module__.startswith("pytds"):
        cursor.executemany(
            f"{insert_sql} VALUES ({', '.join([placeholder] * len(rows[0]))})",
            rows,
        )
        return

    # python-tds implements executemany as one network round-trip per row.
    # Multi-row VALUES keeps each statement below SQL Server's 2,100 parameter limit.
    column_count = len(rows[0])
    batch_size = max(1, 2000 // column_count)
    row_placeholder = f"({', '.join([placeholder] * column_count)})"
    for offset in range(0, len(rows), batch_size):
        batch = rows[offset:offset + batch_size]
        parameters = tuple(value for row in batch for value in row)
        cursor.execute(
            f"{insert_sql} VALUES {', '.join([row_placeholder] * len(batch))}",
            parameters,
        )


def _bulk_insert_import_rows(browser, prepared_rows, connection, columns) -> None:
    if not prepared_rows:
        return
    data_columns = [
        column for column in columns
        if column.sql_name.lower() not in {"browserrecordid", "createdat", "updatedat", "eventchainid"}
    ]
    eventchain_rows = [(row["eventchain_id"], 1, 0, 0) for row in prepared_rows]
    value_rows = [
        tuple([row["eventchain_id"], str(row.get("actor") or "System")[:150]] + [
            import_parameter_value(
                column,
                row["values"].get(column.sql_name, row["values"].get(column.display_name)),
            )
            for column in data_columns
        ])
        for row in prepared_rows
    ]
    batch_cursor = connection.cursor()
    parameter_placeholder = (
        "%s" if batch_cursor.__class__.__module__.startswith("pytds") else "?"
    )
    if hasattr(batch_cursor, "fast_executemany"):
        batch_cursor.fast_executemany = True
    batch_cursor.execute(
        "CREATE TABLE #Mining360ImportEventChain ("
        "EventChainID INT NOT NULL, EventChainTypeID INT NOT NULL, "
        "Created_By INT NOT NULL, User_ID INT NOT NULL)"
    )
    _execute_import_rows(
        batch_cursor,
        "INSERT INTO #Mining360ImportEventChain "
        "(EventChainID, EventChainTypeID, Created_By, User_ID)",
        eventchain_rows,
        parameter_placeholder,
    )
    batch_cursor.execute(
        "INSERT INTO dbo.EventChain (EventChainID, EventChainTypeID, Created_By, User_ID) "
        "SELECT source.EventChainID, MAX(source.EventChainTypeID), "
        "MAX(source.Created_By), MAX(source.User_ID) "
        "FROM #Mining360ImportEventChain source "
        "LEFT JOIN dbo.EventChain target ON target.EventChainID = source.EventChainID "
        "WHERE target.EventChainID IS NULL GROUP BY source.EventChainID"
    )
    batch_cursor.execute("DROP TABLE #Mining360ImportEventChain")
    sql_columns = ["[EventChainID]", "[CreatedBy]"] + [quote_identifier(column.sql_name) for column in data_columns]
    _execute_import_rows(
        batch_cursor,
        f"INSERT INTO {quote_object_name(browser.table_name, 'Table name')} "
        f"({', '.join(sql_columns)})",
        value_rows,
        parameter_placeholder,
    )


def import_browser_records(browser: DataBrowser, uploaded_file, mapping: dict | None = None) -> dict:
    headers, rows = _read_import_rows(uploaded_file)
    mapping = mapping if isinstance(mapping, dict) else {}
    column_map = mapping.get("column_map") if isinstance(mapping.get("column_map"), dict) else {}
    actor = str(mapping.get("_audit_user") or "System")[:150]

    inserted = 0
    updated = 0
    skipped = 0
    errors = []
    error_count = 0
    columns = _effective_browser_columns(browser)
    duplicate_mode = str(mapping.get("duplicate_mode") or "skip").strip().lower()
    if duplicate_mode not in {"skip", "replace"}:
        duplicate_mode = "skip"
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        lookup_cache = _lookup_validation_cache(columns, cursor)
        existing_index = _build_existing_record_index(browser, columns, cursor)
        next_eventchain_id = _next_eventchain_id(cursor)
        for index, row in enumerate(rows, start=2):
            try:
                values = _build_import_row_values(browser, row, column_map)
                existing_record_id = _existing_import_record_id_from_index(values, columns, existing_index)
                if existing_record_id is not None and duplicate_mode == "skip":
                    skipped += 1
                elif existing_record_id is not None:
                    _update_import_record_with_cursor(
                        browser, existing_record_id, values, cursor, columns,
                        actor=actor, lookup_cache=lookup_cache,
                    )
                    updated += 1
                else:
                    inserted_eventchain_id = _insert_browser_record_with_cursor(
                        browser,
                        values,
                        cursor,
                        columns=columns,
                        next_eventchain_id=next_eventchain_id,
                        actor=actor,
                        lookup_cache=lookup_cache,
                    )
                    next_eventchain_id = int(inserted_eventchain_id) + 1
                    inserted += 1
            except Exception as exc:
                error_count += 1
                if len(errors) < 50:
                    errors.append({"row": index, "error": str(exc)})
        connection.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "error_count": error_count,
        "total_rows": len(rows),
        "headers": headers,
    }


def import_browser_records_batch(browser: DataBrowser, token: str, mapping: dict | None = None, start: int = 0, batch_size: int = 50) -> dict:
    session = _load_import_session(token)
    rows = session.get("rows") if isinstance(session.get("rows"), list) else []
    headers = session.get("headers") if isinstance(session.get("headers"), list) else []
    mapping = mapping if isinstance(mapping, dict) else {}
    column_map = mapping.get("column_map") if isinstance(mapping.get("column_map"), dict) else {}
    actor = str(mapping.get("_audit_user") or "System")[:150]
    start = max(0, int(start or 0))
    batch_size = max(1, min(int(batch_size or 50), 250))
    end = min(len(rows), start + batch_size)

    inserted = 0
    updated = 0
    skipped = 0
    errors = []
    error_count = 0
    columns = _effective_browser_columns(browser)
    with connect(database=BROWSER_DATABASE) as connection:
        cursor = connection.cursor()
        lookup_cache = _lookup_validation_cache(columns, cursor)
        existing_index = _build_existing_record_index(browser, columns, cursor)
        next_eventchain_id = _next_eventchain_id(cursor)
        for index, row in enumerate(rows[start:end], start=start + 2):
            try:
                values = _build_import_row_values(browser, row, column_map)
                existing_record_id = _existing_import_record_id_from_index(values, columns, existing_index)
                if existing_record_id is not None and str(mapping.get("duplicate_mode") or "skip").strip().lower() == "skip":
                    skipped += 1
                elif existing_record_id is not None:
                    _update_import_record_with_cursor(
                        browser, existing_record_id, values, cursor, columns,
                        actor=actor, lookup_cache=lookup_cache,
                    )
                    updated += 1
                else:
                    inserted_eventchain_id = _insert_browser_record_with_cursor(
                        browser,
                        values,
                        cursor,
                        columns=columns,
                        next_eventchain_id=next_eventchain_id,
                        actor=actor,
                        lookup_cache=lookup_cache,
                    )
                    next_eventchain_id = int(inserted_eventchain_id) + 1
                    inserted += 1
            except Exception as exc:
                error_count += 1
                if len(errors) < 50:
                    errors.append({"row": index, "error": str(exc)})
        connection.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "error_count": error_count,
        "total_rows": len(rows),
        "headers": headers,
        "next_start": end,
        "done": end >= len(rows),
        "processed": end,
    }


def _run_import_job(job_token: str, browser_id: int, import_token: str, mapping: dict) -> None:
    close_old_connections()
    status = {
        "job_token": job_token,
        "browser_id": browser_id,
        "token": import_token,
        "total_rows": 0,
        "processed": 0,
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "errors": [],
        "error_count": 0,
        "done": False,
        "failed": False,
        "message": "Import started.",
    }
    try:
        browser = DataBrowser.objects.prefetch_related("columns").get(id=browser_id)
        validate_browser_definition(browser)
        if not browser_table_exists(browser):
            sync_browser_sql(browser)
        session = _load_import_session(import_token)
        rows = session.get("rows") if isinstance(session.get("rows"), list) else []
        total_rows = len(rows)
        column_map = mapping.get("column_map") if isinstance(mapping.get("column_map"), dict) else {}
        duplicate_mode = str(mapping.get("duplicate_mode") or "skip").strip().lower()
        actor = str(mapping.get("_audit_user") or "System")[:150]
        if duplicate_mode not in {"skip", "replace"}:
            duplicate_mode = "skip"
        columns = _effective_browser_columns(browser)
        status["total_rows"] = total_rows
        _write_import_job(job_token, status)
        with connect(database=BROWSER_DATABASE) as connection:
            cursor = connection.cursor()
            lookup_cache = _lookup_validation_cache(columns, cursor)
            existing_index = _build_existing_record_index(browser, columns, cursor)
            next_eventchain_id = _next_eventchain_id(cursor)
            pending_inserts = []

            def flush_pending_inserts():
                nonlocal pending_inserts, cursor
                if not pending_inserts:
                    return
                try:
                    _bulk_insert_import_rows(browser, pending_inserts, connection, columns)
                    connection.commit()
                    cursor = connection.cursor()
                    status["inserted"] = int(status["inserted"]) + len(pending_inserts)
                except Exception:
                    connection.rollback()
                    cursor = connection.cursor()
                    for prepared in pending_inserts:
                        try:
                            _bulk_insert_import_rows(browser, [prepared], connection, columns)
                            connection.commit()
                            cursor = connection.cursor()
                            status["inserted"] = int(status["inserted"]) + 1
                        except Exception as row_exc:
                            connection.rollback()
                            cursor = connection.cursor()
                            error_list = status.get("errors") if isinstance(status.get("errors"), list) else []
                            error_list.append({"row": prepared["row"], "error": str(row_exc)})
                            status["errors"] = error_list[:50]
                            status["error_count"] = int(status["error_count"]) + 1
                pending_inserts = []

            for index, row in enumerate(rows, start=1):
                try:
                    values = _build_import_row_values(browser, row, column_map)
                    existing_record_id = _existing_import_record_id_from_index(values, columns, existing_index)
                    if existing_record_id is not None and duplicate_mode == "skip":
                        status["skipped"] = int(status["skipped"]) + 1
                    elif existing_record_id is not None:
                        _update_import_record_with_cursor(
                            browser,
                            existing_record_id,
                            values,
                            cursor,
                            columns,
                            actor=actor,
                            lookup_cache=lookup_cache,
                        )
                        status["updated"] = int(status["updated"]) + 1
                    else:
                        # Validate before queuing so malformed rows do not poison a batch.
                        _validate_lookup_values(values, columns, cursor, cache=lookup_cache)
                        for column in columns:
                            if column.sql_name.lower() not in SYSTEM_SQL_COLUMNS:
                                import_parameter_value(
                                    column,
                                    values.get(column.sql_name, values.get(column.display_name)),
                                )
                        supplied_eventchain_id = values.get("EventChainID", values.get("eventchainid"))
                        import_eventchain_id = (
                            int(supplied_eventchain_id)
                            if supplied_eventchain_id not in (None, "")
                            else next_eventchain_id
                        )
                        pending_inserts.append({
                            "row": index + 1,
                            "eventchain_id": import_eventchain_id,
                            "values": values,
                            "actor": actor,
                        })
                        if supplied_eventchain_id in (None, ""):
                            next_eventchain_id += 1
                        for column in columns:
                            if not column.is_unique:
                                continue
                            normalized = _normalize_import_key(
                                values.get(column.sql_name, values.get(column.display_name))
                            )
                            if normalized:
                                existing_index.setdefault(column.sql_name, {})[normalized] = -1
                        if len(pending_inserts) >= 200:
                            flush_pending_inserts()
                except Exception as exc:
                    error_list = status.get("errors") if isinstance(status.get("errors"), list) else []
                    error_list.append({"row": index + 1, "error": str(exc)})
                    status["errors"] = error_list[:50]
                    status["error_count"] = int(status["error_count"]) + 1
                status["processed"] = index
                if index == total_rows or index % 100 == 0:
                    if index == total_rows:
                        flush_pending_inserts()
                    status["message"] = f"Importing row {index} of {total_rows}..."
                    _write_import_job(job_token, status)
            flush_pending_inserts()
            connection.commit()
        status["done"] = True
        status["failed"] = False
        status["message"] = (
            f"Import completed. Total: {total_rows}; inserted: {status['inserted']}; "
            f"updated: {status['updated']}; skipped: {status['skipped']}; "
            f"errors: {status['error_count']}."
        )
        _write_import_job(job_token, status)
    except Exception as exc:
        status["done"] = True
        status["failed"] = True
        status["fatal_error"] = str(exc)
        status["message"] = (
            f"Import interrupted after {status['processed']} of {status['total_rows']} rows. "
            f"Inserted: {status['inserted']}; updated: {status['updated']}; "
            f"skipped: {status['skipped']}; row errors: {status['error_count']}. "
            f"Technical error: {exc}"
        )
        _write_import_job(job_token, status)
    finally:
        close_old_connections()


def start_import_job(browser: DataBrowser, import_token: str, mapping: dict | None = None) -> dict:
    job_token = uuid4().hex
    mapping = mapping if isinstance(mapping, dict) else {}
    session = _load_import_session(import_token)
    total_rows = len(session.get("rows") if isinstance(session.get("rows"), list) else [])
    _write_import_job(job_token, {
        "job_token": job_token,
        "browser_id": browser.id,
        "token": import_token,
        "total_rows": total_rows,
        "processed": 0,
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "errors": [],
        "error_count": 0,
        "done": False,
        "failed": False,
        "message": "Import queued.",
    })
    thread = threading.Thread(
        target=_run_import_job,
        args=(job_token, browser.id, import_token, mapping),
        daemon=True,
    )
    thread.start()
    return {"job_token": job_token, "total_rows": total_rows}


def get_import_job_status(job_token: str) -> dict:
    return _load_import_job(job_token)
