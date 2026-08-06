from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from reports.data_browsers import (
    BROWSER_DATABASE,
    SYSTEM_SQL_COLUMNS,
    _build_existing_record_index,
    _build_import_row_values,
    _bulk_insert_import_rows,
    _effective_browser_columns,
    _existing_import_record_id_from_index,
    _lookup_validation_cache,
    _next_eventchain_id,
    _update_import_record_with_cursor,
    _validate_lookup_values,
    import_parameter_value,
)
from reports.models import DataBrowser
from reports.sqlserver import connect


def normalize_header(value) -> str:
    return "".join(character for character in str(value or "").lower() if character.isalnum())


class Command(BaseCommand):
    help = "Stream a large Excel file into a configured Data Browser."

    def add_arguments(self, parser):
        parser.add_argument("file_path")
        parser.add_argument("--browser", required=True)
        parser.add_argument("--actor", default="System Import")
        parser.add_argument("--batch-size", type=int, default=500)
        parser.add_argument("--duplicate-mode", choices=["skip", "replace"], default="skip")

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required for Excel imports.") from exc

        path = Path(options["file_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")
        browser = DataBrowser.objects.prefetch_related("columns").filter(
            name=options["browser"]
        ).first()
        if not browser:
            raise CommandError(f"Browser not found: {options['browser']}")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        row_iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(row_iterator)]
        normalized_headers = {normalize_header(header): header for header in headers if header}
        columns = _effective_browser_columns(browser)
        column_map = {}
        for column in columns:
            candidates = [normalize_header(column.display_name), normalize_header(column.sql_name)]
            source_header = next(
                (normalized_headers[candidate] for candidate in candidates if candidate in normalized_headers),
                "",
            )
            column_map[column.sql_name] = {"source_column": source_header}

        mapped = {
            column.display_name: column_map[column.sql_name]["source_column"]
            for column in columns
        }
        self.stdout.write(f"Browser: {browser.name}")
        self.stdout.write(f"Sheet: {sheet.title}")
        self.stdout.write(f"Rows: {max(sheet.max_row - 1, 0)}")
        self.stdout.write(f"Mapping: {mapped}")

        batch_size = max(1, min(int(options["batch_size"]), 1000))
        actor = str(options["actor"] or "System Import")[:150]
        duplicate_mode = options["duplicate_mode"]
        status = {"processed": 0, "inserted": 0, "updated": 0, "skipped": 0, "errors": 0}
        first_errors = []

        with connect(database=BROWSER_DATABASE) as connection:
            cursor = connection.cursor()
            lookup_cache = _lookup_validation_cache(columns, cursor)
            existing_index = _build_existing_record_index(browser, columns, cursor)
            next_eventchain_id = _next_eventchain_id(cursor)
            pending = []

            def flush():
                nonlocal pending, cursor
                if not pending:
                    return
                try:
                    _bulk_insert_import_rows(browser, pending, connection, columns)
                    connection.commit()
                    status["inserted"] += len(pending)
                except Exception:
                    connection.rollback()
                    for prepared in pending:
                        try:
                            _bulk_insert_import_rows(browser, [prepared], connection, columns)
                            connection.commit()
                            status["inserted"] += 1
                        except Exception as exc:
                            connection.rollback()
                            status["errors"] += 1
                            if len(first_errors) < 25:
                                first_errors.append((prepared["row"], str(exc)))
                cursor = connection.cursor()
                pending = []

            for row_number, row in enumerate(row_iterator, start=2):
                status["processed"] += 1
                source_row = dict(zip(headers, row))
                try:
                    values = _build_import_row_values(browser, source_row, column_map)
                    _validate_lookup_values(values, columns, cursor, cache=lookup_cache)
                    existing_id = _existing_import_record_id_from_index(values, columns, existing_index)
                    if existing_id is not None and duplicate_mode == "skip":
                        status["skipped"] += 1
                    elif existing_id is not None:
                        _update_import_record_with_cursor(
                            browser, existing_id, values, cursor, columns,
                            actor=actor, lookup_cache=lookup_cache,
                        )
                        status["updated"] += 1
                    else:
                        for column in columns:
                            if column.sql_name.lower() not in SYSTEM_SQL_COLUMNS:
                                import_parameter_value(
                                    column,
                                    values.get(column.sql_name, values.get(column.display_name)),
                                )
                        pending.append({
                            "row": row_number,
                            "eventchain_id": next_eventchain_id,
                            "values": values,
                            "actor": actor,
                        })
                        next_eventchain_id += 1
                        if len(pending) >= batch_size:
                            flush()
                except Exception as exc:
                    status["errors"] += 1
                    if len(first_errors) < 25:
                        first_errors.append((row_number, str(exc)))

                if status["processed"] % 5000 == 0:
                    flush()
                    self.stdout.write(
                        "Progress: {processed}/{total} | inserted={inserted} | updated={updated} "
                        "| skipped={skipped} | errors={errors}".format(
                            total=max(sheet.max_row - 1, 0), **status
                        )
                    )
            flush()
            connection.commit()

        workbook.close()
        self.stdout.write(self.style.SUCCESS(
            "Completed: processed={processed}, inserted={inserted}, updated={updated}, "
            "skipped={skipped}, errors={errors}".format(**status)
        ))
        if first_errors:
            self.stdout.write("First errors:")
            for row_number, message in first_errors:
                self.stdout.write(f"Row {row_number}: {message}")
